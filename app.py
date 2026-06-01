#左下角搜索并打开 Anaconda Prompt（那个黑框框）。
#输入 D: 并按回车。
#输入 cd D:\桌面备份\毕马威实习\网页实现 并按回车。
#输入 streamlit run app.py 并按回车。

#pip install -U kaleido
#pip install python-pptx
import pdfplumber
import fitz  # PyMuPDF
import json
import re
import io
import time
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed
from openai import OpenAI
import streamlit as st
import requests
from bs4 import BeautifulSoup
import plotly.express as px
import plotly.graph_objects as go
import numpy as np
from pptx import Presentation
from pptx.util import Inches, Pt
from plotly.subplots import make_subplots
import streamlit.components.v1 as components  # 确保引入 components



def get_color_map(all_cos):
    # Create a unique key based on the current selection of companies
    current_selection_key = tuple(sorted(all_cos))
    
    # Check if we need to rebuild the color map (either it doesn't exist, or the selection changed)
    if 'company_color_map' not in st.session_state or st.session_state.get('_last_color_selection') != current_selection_key:
        PRESET_COLORS = ["#C00000", "#0865EE", "#FEAED7", "#92D050", "#7030A0", "#EF9867", "#61CBF4", "#C7A0F7"]
        # Rebuild the map for the new selection
        st.session_state['company_color_map'] = {co: PRESET_COLORS[i % len(PRESET_COLORS)] for i, co in enumerate(all_cos)}
        # Store the current selection key so we know when it changes next time
        st.session_state['_last_color_selection'] = current_selection_key
        
    return st.session_state['company_color_map']

# 然后在 show_step_7_content 内部调用：
# color_map = get_color_map(all_cos)
def show_step_7_content():
    # 🌟 核心修复1：将全局字体配置提到最开头，确保所有子函数、路由引擎都能完美读取，绝不报 NameError
    COMMON_TITLE_FONT = dict(size=18, color="#00338D", family="Microsoft YaHei")

    # ==========================================
    # 1. 样式与前端注入 (紧凑版 CSS & JS)
    # ==========================================

    st.markdown("""
    <style>
    [data-testid="stSidebar"] { background: rgba(255,255,255,0.95) !important; border-right: 1px solid #EAEAEA !important; box-shadow: 2px 0px 15px rgba(0,0,0,0.08) !important; }
    .nav-floating-sign { position: fixed; left: 0; top: 50%; transform: translateY(-50%); background: rgba(0, 51, 141, 0.85); color: white; padding: 20px 8px; border-radius: 0 12px 12px 0; writing-mode: vertical-rl; text-orientation: mixed; font-size: 22px; font-weight: bold; letter-spacing: 3px; z-index: 9999; cursor: pointer; box-shadow: 3px 3px 12px rgba(0,0,0,0.25); transition: all 0.2s; }
    .nav-floating-sign:hover { background: rgba(0, 51, 141, 1); padding-left: 15px; }
   
    /* 放在 @media print { } 块的外面 */
    .stPlotlyChart {
        width: 100% !important;
        min-width: 0 !important;
    }
    .print-only { display: none !important; }
    /* 🌟 新增：专门给封面封底用的样式类 */
    .cover-page {
        position: relative !important;
        width: 338.67mm !important;
        height: 190.5mm !important;
        margin: 0 !important;
        padding: 0 !important;
        page-break-after: always !important;
        overflow: hidden !important;
        background: transparent !important;
    }
    .cover-page img { width: 100% !important; height: 100% !important; object-fit: cover !important; display: block !important; }
    /* Streamlit网页模式去掉container两侧padding */
    .block-container {
        padding-top:0 !important;
        padding-right:10px !important;
        padding-left:10px !important;
        margin-top:0 !important;
    }
    /* 封面封底强制颜色不被浏览器覆盖 */
    .cover-text {
        forced-color-adjust: none !important;
        -webkit-print-color-adjust: exact !important;
        print-color-adjust: exact !important;
        color: white !important;
        -webkit-text-fill-color: white !important;
    }
    .element-container:first-child{
        margin-top:0 !important;
        padding-top:0 !important;
    }
    @media print {
        .print-only { display: block !important; }
        html,body{
            width:338.67mm!important;
            height:190.5mm!important;
            overflow:hidden!important;
            zoom:100%!important;
        }
        .main .block-container{
            max-width:100%!important;
            padding-top:0!important;
            padding-bottom:0!important;
        }   
        .block-container {
            padding-top: 0rem !important;
        }
        /* ===== 隐藏所有交互元素 ===== */
        .no-print, h1, .nav-floating-sign,
        [data-testid="collapsedControl"], header, footer,
        [data-testid="stHeader"], [data-testid="stSidebar"],
        section[data-testid="stSidebar"],
        [data-testid="stToolbar"], button[kind="secondary"],
        input, .stSlider, [data-testid="stSelectbox"],
        [data-testid="stRadio"], [data-testid="stExpander"],
        .stAlert,
        button[role="tab"],
        div[role="tablist"],
        [data-baseweb="tab-list"],
        hr { display: none !important; }

        /* ===== 页面布局撑满纸张 ===== */
        .page-break-container {
            break-inside: avoid !important;
            margin: 0 !important;              /* 清掉容器 margin，防溢出变空白页 */
            padding: 0 !important;
            padding-bottom: 0mm !important;    /* 只留很小的底部间距 */
        }
        .stApp {
            max-width: 100% !important;
            width: 100% !important;
        }
        
        /* 保留两列并排布局的例外 */
        /* ⭐ PDF打印时强制双列 */
        
        .keep-columns [data-testid="stHorizontalBlock"]{
            display:flex!important;
            flex-wrap:nowrap!important;
            align-items:flex-start!important;
            justify-content:space-between!important;
            gap:0!important;
            width:100%!important;
        }
        .keep-columns [data-testid="stHorizontalBlock"]>div{
            width:49%!important;
            min-width:49%!important;
            max-width:49%!important;
            flex:0 0 49%!important;
            overflow:hidden!important;
            page-break-inside:avoid!important;
            break-inside:avoid!important;
        }


        /* ===== 分页标题 ===== */
        .page-break-title {
            break-before: page !important;      /* 用新语法替代 page-break-before */
            padding-top: 10px !important;       /* 原来 40px，缩小防撑出空白 */
            margin-top: 0 !important;
            text-align: left !important;
        }

        /* ===== 标题 ===== */
        h2 {
            display: block !important;
            text-align: left !important;
            color: #00338D !important;
            font-size: 30px !important;
            font-weight: bold !important;
            border-bottom: 2px solid #00338D !important;
            padding-bottom: 6px !important;
            margin: 14px 0 10px 0 !important;
        }
        h3:not(.no-print) {
            display: block !important;
            text-align: left !important;
            color: #00338D !important;
            font-size: 30px !important;
            font-weight: bold !important;
            margin: 10px 0 8px 0 !important;
            page-break-after: avoid !important;
        }

        /* ===== 图表 ===== */
        .plotly-graph-div,
        .stPlotlyChart {
            width: 100% !important;
            max-width: 100% !important;
            height: auto !important;
            page-break-inside: avoid !important;
            display: block !important;
        }

        /* ===== 表格 ===== */
        div[data-testid="stDataFrame"],
        div[data-testid="stTable"] {
            zoom: 0.65 !important;
            margin: 0 auto 20px auto !important;
            max-width: 100% !important;
            page-break-inside: auto !important;
        }
        div[data-testid="stTable"] tr {
            page-break-inside: avoid !important;
        }

        .element-container {
            page-break-inside: avoid !important;
            width: 100% !important;
        }
        .pdf-page-break {
                break-before: page !important;
                page-break-before: always !important;
                height: 0 !important;
                margin: 0 !important;
                padding: 0 !important;
            }
        
            table {
                page-break-inside: auto !important;
            }
            tr {
                page-break-inside: avoid !important;
                page-break-after: auto !important;
            }
            td, th {
                page-break-inside: avoid !important;
            }
            thead {
                display: table-header-group !important;
            }
        }
    }

    /* 纵向打印微调 */
    @media print and (orientation: portrait) {
        .stPlotlyChart { margin-bottom: 10mm !important; }
    }
    /* 横向打印微调 */
    @media print and (orientation: landscape) {
        .stPlotlyChart { margin-bottom: 6mm !important; }
    }

    .stPlotlyChart, div[data-testid="stDataFrame"] { display: flex !important; justify-content: center !important; }
    .highlight-blue-box { border: 1.5px solid rgba(0,51,141,0.85) !important; border-radius: 12px !important; padding: 10px !important; background: rgba(0,51,141,0.02) !important; box-shadow: 0px 4px 12px rgba(0,51,141,0.12) !important; margin-bottom: 25px !important; }
    </style>
    <div class="nav-floating-sign" id="custom-nav-trigger">展开导航栏</div>
    """, unsafe_allow_html=True)

    components.html("""<script>let t = setInterval(() => { const d = window.parent.document; const b = d.getElementById("custom-nav-trigger"); const c = d.querySelector('[data-testid="collapsedControl"]') || d.querySelector('button[kind="header"]'); if(b && c) { b.onclick = () => c.click(); clearInterval(t); } }, 500);</script>""", height=0, width=0)

    # ==========================================
    # 2. 数据检查与年份提取
    # ==========================================
    if 'integrated_data' not in st.session_state or st.session_state['integrated_data'] is None:
        st.warning("⚠️ 请先在 Step 6 完成数据集成。")
        st.stop()
        
    df_raw = st.session_state['integrated_data'].copy()
    if '公司类型' not in df_raw.columns: df_raw['公司类型'] = "未分类"

    valid_years = sorted([y for y in df_raw['报告年份'].dropna().astype(str).str.replace(".0", "", regex=False).unique() if y.isdigit()])
    latest_year = int(valid_years[-1]) if len(valid_years) > 0 else 2025
    prev_year = int(valid_years[-2]) if len(valid_years) > 1 else 2023

    st.markdown("<h3 class='no-print' style='font-weight:700;'>🖼️ 公司级对标报告</h3>", unsafe_allow_html=True)

# ==========================================
    # 3. 必须先加载配置表（从而驱动后续侧边栏）
    # ==========================================
    notes_dict, ordered_modules = {}, []
    def generate_custom_analysis(m_id, df, cos, cy, py):
        """根据m_id和数据自动生成分析内容-自定义，返回字符串"""
        
        def gv(field, year):
            """获取某字段某年所有公司的值，返回 {公司: 值} dict"""
            yr = str(int(year)) if year == int(year) else str(year)
            out = {}
            for co in cos:
                s = df[(df['公司']==co) &
                       (df['报告年份'].astype(str).str.replace('.0','',regex=False)==yr) &
                       (df['字段名']==field)]['(百万)人民币']
                out[co] = s.sum() if not s.empty else None
            return out
    
        def rise_fall(field):
            cv, pv = gv(field, cy), gv(field, py)
            rises, falls = [], []
            for co in cos:
                c, p = cv.get(co), pv.get(co)
                # ✅ None 表示未披露，跳过；p==0 无法计算增幅，也跳过
                if c is None or p is None or p == 0: continue
                chg = (c - p) / abs(p)
                (rises if chg >= 0 else falls).append((co, chg))
            parts = []
            if rises:
                rises.sort(key=lambda x: -x[1])
                parts.append(f"{len(rises)}家公司上升，其中{rises[0][0]}增幅最大（+{rises[0][1]:.1%}）")
            if falls:
                falls.sort(key=lambda x: x[1])
                parts.append(f"{len(falls)}家公司下降，其中{falls[0][0]}降幅最大（{falls[0][1]:.1%}）")
            return "；".join(parts) + "。" if parts else ""
    
        def trend_only(field):
            cv, pv = gv(field, cy), gv(field, py)
            rises, falls, vals = [], [], []
            for co in cos:
                c, p = cv.get(co), pv.get(co)
                if c is None or p is None: continue  # ✅ 未披露跳过
                vals.append(c)
                (rises if c >= p else falls).append(co)
            parts = []
            if rises: parts.append(f"{len(rises)}家上升")
            if falls: parts.append(f"{len(falls)}家下降")
            if vals: parts.append(f"{cy}年均值为{sum(vals)/len(vals)/1e4:.1f}亿元")
            return "；".join(parts) + "。" if parts else ""
    
        def composition(fields_dict, year):
            field_list = list(fields_dict.keys())
            totals = {co: 0 for co in cos}
            vals = {f: {} for f in field_list}
            for f in field_list:
                fv = gv(f, year)
                for co in cos:
                    v = fv.get(co)  # ✅ 不用 or 0，保留 None
                    vals[f][co] = abs(v) if v is not None else 0
                    if v is not None:
                        totals[co] += abs(v)
        
            # ✅ 只统计有真实数据的公司
            disclosed_cos = [co for co in cos if totals[co] > 0]
            if not disclosed_cos:
                return ""
        
            dominant_counts = {f: 0 for f in field_list}
            pcts = {f: {} for f in field_list}
            for co in disclosed_cos:  # ✅ 只遍历披露公司
                tot = totals[co]
                best_f = max(field_list, key=lambda f: vals[f].get(co, 0))
                dominant_counts[best_f] += 1
                for f in field_list:
                    pcts[f][co] = vals[f].get(co, 0) / tot * 100
        
            main_field = max(dominant_counts, key=dominant_counts.get)
            main_name = fields_dict[main_field]
            avg_pct = sum(pcts[main_field].values()) / len(pcts[main_field]) if pcts[main_field] else 0
            max_co = max(pcts[main_field], key=pcts[main_field].get) if pcts[main_field] else ""
            max_pct = pcts[main_field].get(max_co, 0)
        
            return (f"{len(disclosed_cos)}家披露公司中，主要构成部分为{main_name}，"
                    f"该项目占比均值约{avg_pct:.1f}%，"
                    f"其中{max_co}占比最高（{max_pct:.1f}%）。")
    
        # ==================== 分类路由 ====================
        try:
# --- 🌟 专属定制逻辑：综合净资产 ---
            if m_id == "csm_equity":
                # 图表上的“综合净资产” = 期末股东权益 + CSM期末余额 * 0.75
                cv_eq, pv_eq = gv("期末股东权益", cy), gv("期末股东权益", py)
                cv_csm, pv_csm = gv("CSM期末余额", cy), gv("CSM期末余额", py)
                
                rises, falls = [], []
                for co in cos:
                    # 获取当期和前期的权益和CSM，如果缺失则视为0
                    c_e, p_e = cv_eq.get(co) or 0, pv_eq.get(co) or 0
                    c_c, p_c = cv_csm.get(co) or 0, pv_csm.get(co) or 0
                    
                    # 按照图表的逻辑，计算两年的“综合净资产”
                    c_total = c_e + c_c * 0.75
                    p_total = p_e + p_c * 0.75
                    
                    if p_total == 0: continue # 分母为0则跳过
                    
                    chg = (c_total - p_total) / abs(p_total)
                    (rises if chg >= 0 else falls).append((co, chg))
                
                parts = []
                if rises:
                    rises.sort(key=lambda x: -x[1])
                    parts.append(f"{len(rises)}家上升，其中{rises[0][0]}增幅最大（+{rises[0][1]:.1%}）")
                if falls:
                    falls.sort(key=lambda x: x[1])
                    parts.append(f"{len(falls)}家下降，其中{falls[0][0]}降幅最大（{falls[0][1]:.1%}）")
                
                return "综合净资产方面，" + "；".join(parts) + "。" if parts else "综合净资产较去年无明显变动。"
            
            if m_id == "nb_lost":
                cv, pv = gv("新业务亏损合同（LC）——非PAA", cy), gv("新业务亏损合同（LC）——非PAA", py)
                # ✅ 取负数，和图表展示保持一致
                cv = {co: -v if v is not None else None for co, v in cv.items()}
                pv = {co: -v if v is not None else None for co, v in pv.items()}
                rises, falls = [], []
                for co in cos:
                    c, p = cv.get(co), pv.get(co)
                    if c is None or p is None or p == 0: continue
                    chg = (c - p) / abs(p)
                    (rises if chg >= 0 else falls).append((co, chg))
                parts = []
                if rises:
                    rises.sort(key=lambda x: -x[1])
                    parts.append(f"{len(rises)}家公司上升，其中{rises[0][0]}增幅最大（+{rises[0][1]:.1%}）")
                if falls:
                    falls.sort(key=lambda x: x[1])
                    parts.append(f"{len(falls)}家公司下降，其中{falls[0][0]}降幅最大（{falls[0][1]:.1%}）")
                return "；".join(parts) + "。" if parts else ""
            # 涨跌型
            if m_id == "inc_total":   return rise_fall("保险服务收入合计")
            if m_id == "exp_total":   return rise_fall("保险服务费用合计")
            if m_id == "perf_total":  return rise_fall("保险服务业绩")
            if m_id == "inv_return":  return rise_fall("净投资回报")
            if m_id == "uw_profit":   return rise_fall("承保财务净损益")
            if m_id == "inv_profit":  return rise_fall("投资利润")
            if m_id == "net_profit":  return rise_fall("净利润")
            if m_id == "oci_profit":  return rise_fall("其他综合收益")
            if m_id == "total_profit":return rise_fall("综合收益总额")
            if m_id == "asset_trend": return rise_fall("总资产")
            if m_id == "equity_trend":return rise_fall("期末股东权益")
            if m_id == "csm_bal":     return rise_fall("CSM期末余额")
            if m_id == "nb_csm":      return rise_fall("新业务CSM（集团口径）")
    
            # 构成型（最新年）
            if m_id == "comp_1":
                return composition({
                    "采用保费分配法计量的保险合同保险服务收入": "PAA合同收入",
                    "未采用保费分配法计量的保险合同保险服务收入": "非PAA合同收入"
                }, cy)
            if m_id == "comp_2":
                return composition({
                    "合同服务边际的摊销": "合同服务边际释放",
                    "非金融风险调整的变动": "非金融风险调整变动",
                    "预计当期发生的保险服务费用": "预期保险服务费用",
                    "保险获取现金流的摊销（保险服务收入）": "保险获取现金流摊销",
                }, cy)
            if m_id == "exp_1":
                return composition({
                    "保险获取现金流的摊销（保险服务费用）": "保险获取现金流摊销",
                    "亏损部分的确认及转回": "亏损部分确认及转回",
                    "当期发生的赔款及其他相关费用": "当期赔款及费用",
                    "已发生赔款负债相关的履约现金流量变动": "已发生赔款负债变动",
                }, cy)
            if m_id == "asset_struct":
                return composition({
                    "债权投资": "债权投资(AC)",
                    "其他债权投资": "其他债权投资(FVOCI)",
                    "交易性金融资产": "交易性金融资产(FVTPL)",
                    "其他权益工具投资": "其他权益工具(指定FVOCI)",
                }, cy)
            if m_id == "csm_trans":
                return composition({
                    f"采用公允价值法计量的合同期末负债": "公允价值法",
                    f"采用修正追溯调整法计量的合同期末负债": "修正追溯法",
                    f"其他合同期末负债": "其他合同",
                }, cy)
            if m_id in ["csm_comp_lat", "csm_comp_pre"]:
                y = cy if m_id == "csm_comp_lat" else py
                return composition({
                    "新业务CSM（集团口径）": "新业务CSM",
                    "CSM计息": "CSM计息",
                    "CSM调整": "CSM调整",
                }, y)
    
            # 趋势型（两年比较，说上升/下降）
            if m_id == "csm_ratio":
                cv = {co: None for co in cos}
                pv = {co: None for co in cos}
                for co in cos:
                    df_co = df[df['公司']==co]
                    def get_ratio(yr):
                        def s(k): return df_co[(df_co['报告年份'].astype(str).str.replace('.0','',regex=False)==str(yr)) & (df_co['字段名'].str.contains(k,na=False))]['(百万)人民币'].sum()
                        a,b,c = s('CSM期末'),s('LRC非亏损部分期末'),s('LRC亏损部分期末')
                        denom = b+c-a
                        return a/denom if denom!=0 else None
                    cv[co] = get_ratio(cy)
                    pv[co] = get_ratio(py)
                rises, falls, vals = [], [], []
                for co in cos:
                    c, p = cv.get(co), pv.get(co)
                    if c is None or p is None: continue
                    vals.append(c)
                    (rises if c >= p else falls).append(co)
                avg = sum(vals)/len(vals) if vals else 0
                parts = []
                if rises: parts.append(f"{'、'.join(rises)}占比上升")
                if falls: parts.append(f"{'、'.join(falls)}占比下降")
                parts.append(f"{cy}年均值约{avg:.1%}")
                return "；".join(parts) + "。"
    
            if m_id == "nb_struct":
                # 新业务CSM利润率趋势
                cy_v = gv("新业务CSM（集团口径）", cy)
                cy_pv = gv("新业务未来现金流入现值（盈利）", cy)
                py_v = gv("新业务CSM（集团口径）", py)
                py_pv = gv("新业务未来现金流入现值（盈利）", py)
                rises, falls, vals = [], [], []
                for co in cos:
                    c = cy_v.get(co,0) / cy_pv.get(co,1) if cy_pv.get(co) else None
                    p = py_v.get(co,0) / py_pv.get(co,1) if py_pv.get(co) else None
                    if c is None or p is None: continue
                    vals.append(c)
                    (rises if c >= p else falls).append(co)
                avg = sum(vals)/len(vals) if vals else 0
                parts = []
                if rises: parts.append(f"{'、'.join(rises)}CSM利润率上升")
                if falls: parts.append(f"{'、'.join(falls)}CSM利润率下降")
                parts.append(f"{cy}年均值约{avg:.1%}")
                return "；".join(parts) + "。"
    
            if m_id == "exp_struct":
                # 各项占比最大的公司
                acq = gv("获取费用", cy)
                maint = gv("维持费用", cy)
                non_ = gv("非履约费用", cy)
                max_acq = max((co for co in cos if acq.get(co)), key=lambda c: acq.get(c,0), default="")
                max_maint = max((co for co in cos if maint.get(co)), key=lambda c: maint.get(c,0), default="")
                max_non = max((co for co in cos if non_.get(co)), key=lambda c: non_.get(c,0), default="")
                return (f"获取费用绝对金额最高：{max_acq}；"
                        f"维持费用绝对金额最高：{max_maint}；"
                        f"非履约费用绝对金额最高：{max_non}。")
    
            if m_id == "tax_profit":
                return f"各公司税前利润与净利润对比详见图表，关注有效税率变动情况。"
    
            if m_id == "prof_mix":
                ins = gv("保险利润", cy)
                inv = gv("投资利润", cy)
                ins_dom, inv_dom, no_data = [], [], []
                for co in cos:
                    c_ins = ins.get(co) or 0
                    c_inv = inv.get(co) or 0
                    tot = c_ins + c_inv
                    if tot == 0:
                        no_data.append(co)  # ✅ 没数据的单独列出，不乱归类
                        continue
                    ins_pct = c_ins / tot
                    (ins_dom if ins_pct >= 0.5 else inv_dom).append(
                        f"{co}（保险利润占比{ins_pct:.0%}）"
                    )
                parts = []
                if ins_dom: parts.append(f"保险利润占主导：{'、'.join(ins_dom)}")
                if inv_dom: parts.append(f"投资利润占主导：{'、'.join(inv_dom)}")
                if no_data: parts.append(f"未披露：{'、'.join(no_data)}")
                return "；".join(parts) + "。" if parts else ""
    
        except Exception as e:
            return f"（自动生成失败：{e}）"
    
        return ""  # 无话术的m_id返回空
    st.markdown("<div class='no-print'>", unsafe_allow_html=True)
    with st.expander("📥 全局内容分析与注释输入", expanded=False):
        st.markdown("""
            <div style='background:linear-gradient(135deg,#00338D,#0865EE); 
            border-radius:10px; padding:14px 18px; margin-bottom:16px;
            display:flex; align-items:center; justify-content:space-between;'>
                <span style='color:white; font-size:15px; font-weight:bold;'>
                    获取官方注释表模板
                </span>
                <span style='color:rgba(255,255,255,0.8); font-size:12px;'>
                    需要安全码验证
                </span>
            </div>
        """, unsafe_allow_html=True)
    
        col_pwd, col_btn = st.columns([3, 1])
        with col_pwd:
            pwd_input = st.text_input("", placeholder="请输入安全码...", 
                                       type="password", key="template_pwd",
                                       label_visibility="collapsed")
        with col_btn:
            check_btn = st.button("🔓验证下载", use_container_width=True)
    
        if check_btn:
            # 🌟 核心修复 1：一点击按钮，立刻清空之前的旧文件缓存
            if 'filled_notes_excel' in st.session_state:
                del st.session_state['filled_notes_excel']
                
            if pwd_input == "KPMG666":
                try:
                    import requests
                    from io import BytesIO
                    import openpyxl
        
                    url = "https://github.com/z-xylym/my-actuary-tool/raw/refs/heads/main/step7%E5%86%85%E5%AE%B9%E5%88%86%E6%9E%90%E5%92%8C%E6%B3%A8%E9%87%8A%E6%A8%A1%E6%9D%BF%E5%8F%AA%E5%90%AB%E9%BB%98%E8%AE%A4.xlsx"
                    r = requests.get(url, timeout=15)
        
                    if r.status_code == 200:
                        wb = openpyxl.load_workbook(BytesIO(r.content))
                        ws = wb.active
                        header = {cell.value: cell.column for cell in ws[1]}
                        mid_col = header.get('模块ID')
                        custom_col = header.get('分析内容-自定义')
        
                        if mid_col and custom_col and 'integrated_data' in st.session_state:
                            df_for_gen = st.session_state['integrated_data'].copy()
                            _cos = st.session_state.get('selected_cos_cache', [])
                            _valid_years = sorted([y for y in df_for_gen['报告年份'].dropna().astype(str).str.replace(".0","",regex=False).unique() if y.isdigit()])
                            _cy = int(_valid_years[-1]) if _valid_years else 2025
                            _py = int(_valid_years[-2]) if len(_valid_years) > 1 else 2024
        
                            for row in ws.iter_rows(min_row=2):
                                m_id_val = row[mid_col - 1].value
                                if not m_id_val: continue
                                generated = generate_custom_analysis(
                                    str(m_id_val).strip(), df_for_gen, _cos, _cy, _py
                                )
                                if generated:
                                    row[custom_col - 1].value = generated
        
                        output = BytesIO()
                        wb.save(output)
                        output.seek(0)
                        
                        # 🌟 核心修复 2：只有走到这一步，完全成功了，才把文件放入缓存
                        st.session_state['filled_notes_excel'] = output.getvalue()
                        st.success("✅ 验证成功，分析内容已自动填充，请点击下方按钮下载")
                    else:
                        st.error("❌ 文件获取失败")
                except Exception as e:
                    st.error(f"❌ 错误：{e}")
            else:
                st.error("❌ 安全码错误")
        
        # ✅ 下载按钮：严格依赖本次验证成功放入的 session_state
        if 'filled_notes_excel' in st.session_state:
            st.download_button(
                label="🔓点击下载（已自动填充分析内容）",
                data=st.session_state['filled_notes_excel'],
                file_name=f"注释表_{st.session_state.get('selected_cos_cache', [''])[0] if st.session_state.get('selected_cos_cache') else ''}_{2025}年.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
        st.divider()        
        use_default = st.toggle("使用默认注释表", value=True, key="use_default_notes")
        df_notes = None
        if use_default:
            try:
                df_notes = pd.read_excel("https://github.com/z-xylym/my-actuary-tool/raw/refs/heads/main/step7%E5%86%85%E5%AE%B9%E5%88%86%E6%9E%90%E5%92%8C%E6%B3%A8%E9%87%8A%E6%A8%A1%E6%9D%BF%E5%8F%AA%E5%90%AB%E9%BB%98%E8%AE%A4.xlsx")
                st.success("✅ 内置默认注释表加载成功")
            except Exception as e:
                st.error(f"❌ 加载失败：{e}")
        else:
            st.info("💡 请上传包含【图片文件名】【模块ID】【一级分类】【二级分类】【对应图表名称】【分析内容】【注释内容】的 Excel")
            notes_file = st.file_uploader("上传 Excel", type=['xlsx', 'xls'])
            if notes_file: df_notes = pd.read_excel(notes_file)

        if df_notes is not None:
            # 统一清洗文本前后的空格
            for col in ['一级分类', '二级分类', '对应图表名称', '模块ID']:
                if col in df_notes.columns:
                    df_notes[col] = df_notes[col].astype(str).str.strip().replace(['nan', 'NaN', 'NAN', 'None'], '')
            
            # 🌟 核心智能归类补丁：如果二级分类为空白，强制塞入“全部”
            if '二级分类' in df_notes.columns:
                df_notes['二级分类'] = df_notes['二级分类'].apply(lambda x: "全部" if str(x).strip() == "" else str(x).strip())
            
            has_img = '图片文件名' in df_notes.columns
            for _, r in df_notes.iterrows():
                m_id = str(r.get('模块ID', '')).strip()
                if not m_id: continue
                img_val = str(r.get('图片文件名', '')).strip() if has_img and pd.notna(r.get('图片文件名')) else ''
                notes_dict[m_id] = {
                    'title': str(r.get('对应图表名称', '')).strip(),
                    # ✅ 拆成两个字段分别读取
                    'analysis_default': str(r.get('分析内容-默认', '')).strip() if pd.notna(r.get('分析内容-默认')) else '',
                    'analysis_custom': str(r.get('分析内容-自定义', '')).strip() if pd.notna(r.get('分析内容-自定义')) else '',
                    'note': str(r.get('注释内容', '')).strip() if pd.notna(r.get('注释内容')) else '',
                    'image_file': img_val if img_val.lower() != 'nan' else ''
                }
                if m_id not in ordered_modules: ordered_modules.append(m_id)
            # 存入全局缓存
            st.session_state['df_notes'] = df_notes
    st.markdown("</div>", unsafe_allow_html=True)
    # ==========================================
    # 4. 侧边栏：智能多级联动导航（防越界护盾版）
    # ==========================================
    print_mode, active_m_id, active_level_context = False, None, ""
    with st.sidebar:
        st.markdown("<h3 style='color: #00338D; font-size: 18px;'>报告目录导航</h3>", unsafe_allow_html=True)
        if 'df_notes' in st.session_state and st.session_state['df_notes'] is not None and not st.session_state['df_notes'].empty:
            df_n = st.session_state['df_notes']
            
            # 过滤提取有效的一级分类
            first_levels = [x for x in df_n['一级分类'].unique() if str(x).strip() != ""]
            main_nav = st.radio("📁 一级模块", first_levels + ["🖨️ 一键显示全部 (打印/导出)"], key="s7_m")

            if main_nav == "🖨️ 一键显示全部 (打印/导出)":
                print_mode = True
                if not st.session_state.get("pdf_export_mode", False):
                    st.info('竖版适合文字多的页面，横版适合宽图表。勾选"背景图形"以保留颜色。')
                    components.html("""
                    <div style="display:flex; flex-direction:column; gap:8px;">
                        <button onclick="printAs('portrait')" style="width:100%; padding:11px; background:#00338D; color:white; border:none; border-radius:6px; cursor:pointer; font-weight:bold; font-size:13px;">
                            🖨️ 导出竖版 A4 PDF
                        </button>
                        <button onclick="printAs('widescreen')" style="width:100%; padding:11px; background:#008578; color:white; border:none; border-radius:6px; cursor:pointer; font-weight:bold; font-size:13px;">
                            🖨️ 导出横版 16:9 PDF
                        </button>
                    </div>
                    <script>
                    // 页面加载时就注入默认横版 16:9，不等用户点
                    (function() {
                        const doc = window.parent.document;
                        const old = doc.getElementById('dynamic-print-style');
                        if (old) old.remove();
                        const style = doc.createElement('style');
                        style.id = 'dynamic-print-style';
                        // 把 widescreen 那个改成
                        style.innerHTML = '@page { size: 338.67mm 190.5mm; margin: 8mm 12mm; } @page :first { margin: 0mm; } @page :last { margin: 0mm; }';
                        doc.head.appendChild(style);
                    })();
                    
                    function printAs(mode) {
                        const doc = window.parent.document;
                        const old = doc.getElementById('dynamic-print-style');
                        if (old) old.remove();
                        const style = doc.createElement('style');
                        style.id = 'dynamic-print-style';
                        if (mode === 'widescreen') {
                            style.innerHTML =
                            '@page { size: landscape; margin: 5mm; }';
                        } else {
                            style.innerHTML =
                            '@page { size: A4 portrait; margin: 10mm; }';
                        }
                        doc.head.appendChild(style);
                        setTimeout(() => window.parent.print(), 150);
                    }
                    </script>
                    """, height=100)         
            else:
                df_sub1 = df_n[df_n['一级分类'] == main_nav]
                
                # 🌟 获取真实的二级分类列表（去除“全部”这个名字，稍后我们手动加上去，保证“全部”永远在第一位）
                raw_sec_levels = [x for x in df_sub1['二级分类'].unique() if x != "全部"]
                
                # 如果这个一级分类下，所有的图表都没有二级分类（即 raw_sec_levels 为空）
                if len(raw_sec_levels) == 0:
                    charts = [x for x in df_sub1['对应图表名称'].unique() if x]
                    chart_nav = st.radio("具体图表", charts, key="s7_c")
                    matched = df_sub1[df_sub1['对应图表名称'] == chart_nav]
                    active_m_id = matched.iloc[0]['模块ID'] if not matched.empty else None
                else:
                    # 如果有正常的二级分类，则强行加上“全部”作为第一选项
                    sub_nav = st.radio("📂 二级模块", ["全部"] + raw_sec_levels, key="s7_s")

                    if sub_nav != "全部":
                        df_sub2 = df_sub1[df_sub1['二级分类'] == sub_nav]
                        charts = [x for x in df_sub2['对应图表名称'].unique() if x]
                        chart_nav = st.radio("具体图表", charts, key="s7_c")
                        matched = df_sub2[df_sub2['对应图表名称'] == chart_nav]
                        active_m_id = matched.iloc[0]['模块ID'] if not matched.empty else None
                    else:
                        # 🌟 选择了“全部”，展示该一级分类下的所有图表（包括那些原本没有二级分类而被归入“全部”的图）
                        charts = [x for x in df_sub1['对应图表名称'].unique() if x]
                        chart_nav = st.radio("具体图表 (当前二级：全部)", charts, key="s7_c_all")
                        matched = df_sub1[df_sub1['对应图表名称'] == chart_nav]
                        active_m_id = matched.iloc[0]['模块ID'] if not matched.empty else None
        else:
            st.warning("⚠️ 请先加载包含层级信息的注释表")

    # ==========================================
    # 5. 全局配置与图片上传 (UI极致压缩)
    # ==========================================
    st.markdown("<div class='no-print'>", unsafe_allow_html=True)
    with st.expander("⚙️ 全局图表设置与图片覆盖", expanded=False):
        c0, c1, c2, c3, c4 = st.columns([1, 2, 1, 1, 1])   
        with c0:
            all_types = ["全部"] + sorted([str(x) for x in df_raw['公司类型'].unique() if str(x) != 'nan'])
            selected_type = st.selectbox("公司类型", all_types)
            df_filtered = df_raw[df_raw['公司类型'] == selected_type].copy() if selected_type != "全部" else df_raw.copy()
        with c1: 
            raw_ordered_cos = list(dict.fromkeys(df_filtered['公司'].dropna().tolist()))
            selected_cos = st.multiselect("展示公司", options=raw_ordered_cos, default=raw_ordered_cos)
            st.session_state['selected_cos_cache'] = selected_cos
        with c2: 
            unit_label = st.selectbox("显示单位", ["十亿元", "亿元", "百万元", "十万元"])
            divisor = {"十亿元": 1000, "亿元": 100, "百万元": 1, "十万元": 0.1}[unit_label]
        with c3: highlight_co = st.selectbox("特定追踪", ["无"] + selected_cos)
        with c4: enable_ai = st.toggle("一键AI分析", value=False)
        
        HIGHLIGHT_COLOR, HL_BOX_LINE, HL_BOX_FILL = "#00338D", "rgba(0,51,141,0.35)", "rgba(0,51,141,0.03)"
        
        st.markdown("---")
        st.caption("📸 手动上传图片（png或jpg）")
        if 'manual_upload_images' not in st.session_state: st.session_state.manual_upload_images = {}
        uploaded_files = st.file_uploader("拖入截图文件", type=['png', 'jpg'], accept_multiple_files=True)
        
        if uploaded_files and ordered_modules:
            # 🌟 核心修改：定义一个你写了代码渲染的 m_id 黑名单集合
            code_rendered_mids = {"summary_table", "csm_growth_chart", "another_code_chart"} # 在这里填入所有写了代码的模块ID
            
            # 过滤出那些“没有写代码渲染”的纯图片/文字模块，供上传匹配
            non_code_modules = [m for m in ordered_modules if m not in code_rendered_mids]
            
            cols = st.columns(2)
            for i, file in enumerate(uploaded_files):
                with cols[i % 2]:
                    get_name = lambda m: "不匹配/跳过" if m == "不匹配/跳过" else notes_dict.get(m, {}).get('title', m)
                    
                    # 🌟 核心修改：将 options 里的 ordered_modules 替换为过滤后的 non_code_modules
                    sel_mid = st.selectbox(
                        f"图片 {file.name} 对应：", 
                        options=["不匹配/跳过"] + non_code_modules, 
                        format_func=get_name, 
                        key=f"up_{i}"
                    )
                    if sel_mid != "不匹配/跳过": 
                        st.session_state.manual_upload_images[sel_mid] = file
                    st.image(file, use_column_width=True)
    st.markdown("</div>", unsafe_allow_html=True)
    # ==========================================
    # 6. AI 引擎与辅助显示函数 (去重并压缩逻辑)
    # ==========================================
    @st.cache_data(show_spinner=False)
    def _call_llm_api_cached(data_str, field_name, latest_year, unit_str, user_api_key, api_base, api_model):
        try:
            import openai
            client = openai.OpenAI(api_key=user_api_key, base_url=api_base)
            prompt = f"你是资深四大精算专家。几家险企{latest_year}年【{field_name}】数据(单位:{unit_str}):{data_str}。请一句话点评(指出极值与趋势)，不超60字。"
            res = client.chat.completions.create(model=api_model, messages=[{"role": "user", "content": prompt}], temperature=0.2)
            return f"<b><span style='color:#00338D'></span></b> {res.choices[0].message.content}"
        except Exception as e: return f"<span style='color:#C00000; font-size:12px;'>⚠️ AI报错: {str(e)}</span>"

    def generate_ai_insight(df, field_name, is_pct=False):
        if not enable_ai or df is None or df.empty or not field_name: return ""
        try:
            d_sub = df[(df.get('字段名', df.get('指标名称', pd.Series(''))).str.contains(field_name, na=False)) & (df['报告年份'].astype(str) == str(latest_year)) & (df['公司'].isin(selected_cos))] if '字段名' in df.columns or '指标名称' in df.columns else df[df['公司'].isin(selected_cos)]
            if d_sub.empty: return ""
            val_col = "(百万)人民币" if "(百万)人民币" in d_sub.columns else d_sub.columns[-1]
            
            ukey, ubase, umodel = st.session_state.get('api_key', "").strip(), st.session_state.get('base_url', "https://api.deepseek.com"), st.session_state.get('model_name', "deepseek-chat")
            if ukey:
                return _call_llm_api_cached(", ".join([f"{r['公司']}:{r[val_col]:.2f}" for _, r in d_sub.iterrows()]), field_name, latest_year, "百分比" if is_pct else unit_label, ukey, ubase, umodel)
            avg, max_r, min_r, fmt, mlt = d_sub[val_col].mean(), d_sub.loc[d_sub[val_col].idxmax()], d_sub.loc[d_sub[val_col].idxmin()], "{:.1f}%" if is_pct else "{:.1f}", 100 if is_pct else (1/divisor)
            return f"<b> </b>{latest_year}年【{field_name}】均值 <b>{fmt.format(avg*mlt)}</b>。<b>{max_r['公司']}</b> 最高，<b>{min_r['公司']}</b> 最低。"
        except Exception as e: return f"<span style='color:#C00000;'>⚠️ 数据报错: {e}</span>"

    def display_notes(module_id, ai_df=None, ai_field=None, is_pct=False):
        md = notes_dict.get(module_id, {})
        an_default = md.get('analysis_default', "")
        an_custom  = md.get('analysis_custom', "")
        nt         = md.get('note', "")
        ai_txt     = generate_ai_insight(ai_df, ai_field, is_pct)
    
        # ✅ 核心：只有用默认表且自定义列为空时，才自动生成
        using_default = st.session_state.get('use_default_notes', True)
        if using_default and not an_custom:
            an_custom = generate_custom_analysis(
                module_id, df_filtered, selected_cos, latest_year, prev_year
            )
    
        if an_default or an_custom or ai_txt:
            html = '<div style="text-align:left; background:#F0F4FA; border-left:4px solid #00338D; padding:3px 10px; margin-bottom:6px; border-radius:4px; font-family:Microsoft YaHei, 微软雅黑, sans-serif;">'
            if an_default:
                html += f'<p style="margin:2px 0; color:#0A1F5C; font-size:14px; line-height:1.4;">{an_default}</p>'
            if an_custom:
                html += f'<p style="margin:2px 0; color:#7A9CC5; font-size:14px; line-height:1.4;">{an_custom}</p>'
            if ai_txt:
                html += f'<p style="margin:2px 0; color:#D84315; font-size:12px; line-height:1.4;">{ai_txt}</p>'
            st.markdown(html + "</div>", unsafe_allow_html=True)
    
        return (an_default or an_custom), nt

    def display_bottom_note(nt_text):
        if nt_text and str(nt_text).lower() != 'nan':
            st.markdown(f'<div style="text-align: left;margin-top: 1px; margin-bottom: 0px; padding-left: 5px;text-align: left;"><p style="margin: 0; color: #888; font-size: 12px; font-style: italic;">* 注释：{nt_text}</p></div>', unsafe_allow_html=True)

    def show_chart(fig,p_mode,m_id=None):
        if not fig:return
        if p_mode:
            H={"csm_trans":440,"oci_deep":280,"nb_struct":125,"csm_ratio_trend":450,"six_dimensional_charts":180,"nb_margin_trend":450,"csm_maturity_table":460}
            h=H.get(m_id,380)
            fig.update_layout(
                autosize=False,
                width=1500,
                height=h,
                margin=dict(t=35,b=15,l=15,r=15)
            )
            st.plotly_chart(fig,use_container_width=False,config={"displayModeBar":False})
        else:
            fig.update_layout(autosize=True)
            st.plotly_chart(fig,use_container_width=True,config={"displayModeBar":False})


    # ==========================================
    # 📊 图表生成核心逻辑区 (已剥离内置标题，完全适配动态引擎)
    # ==========================================

# --- 1.新业务相关比例 ---
    def create_financial_summary_table(df, cos, highlight_co="无"):
        cy, py = latest_year, prev_year
        col_names, c1, c2, c3, c4, c5, c6, c7, c8 = [], [], [], [], [], [], [], [], []
        for co in cos:
            df_co = df[df['公司'] == co]
            get_val = lambda y, kw: df_co[(df_co['报告年份'].astype(str)==str(y)) & (df_co['字段名']==kw)].drop_duplicates(subset=['公司', '报告年份', '字段名'])['(百万)人民币'].sum()
            calc = lambda n, d, ig=False: "-" if d==0 or pd.isna(d) or pd.isna(n) else f"{(n/d)-1 if ig else (n/d):.0%}"
            
            eq_cy, eq_py, np_cy, np_py = get_val(cy, '期末股东权益'), get_val(py, '期末股东权益'), get_val(cy, '净利润'), get_val(py, '净利润')
            csm_end, csm_beg, nb_cy, nb_py = get_val(cy, 'CSM期末余额'), get_val(cy, 'CSM期初余额'), get_val(cy, '新业务CSM（集团口径）'), get_val(py, '新业务CSM（集团口径）')
            
            # 🌟 获取当期和上期的 CSM摊销、保险服务收入、保险合同负债
            amort_cy, rev_cy, rev_py = get_val(cy, 'CSM摊销'), get_val(cy, '保险服务收入合计'), get_val(py, '保险服务收入合计')
            liab_cy, liab_py = get_val(cy, '期末保险合同负债总额'), get_val(py, '期末保险合同负债总额')
            
            # 🌟 新增：获取当期和上期的 PAA期末余额合计
            paa_cy, paa_py = get_val(cy, 'PAA期末余额合计'), get_val(py, 'PAA期末余额合计')
            
            col_names.append(co)
            c1.append(calc(eq_cy, eq_py, True))
            c2.append(calc(np_cy, np_py, True))
            c3.append(calc(csm_end, csm_beg, True))
            c4.append(calc(nb_cy, nb_py, True))
            c5.append(calc(-amort_cy, csm_end - amort_cy, False))
            c6.append(calc(-nb_cy, amort_cy, False))
            c7.append(calc(rev_cy, rev_py, True))
            
            # 🌟 核心修改：c8 加入 PAA期末余额合计 的加总
            c8.append(calc(liab_cy + paa_cy, liab_py + paa_py, True))
    
        headers = ["公司名称", f"净资产<br>%变化<br>{str(cy)[-2:]}YE/{str(py)[-2:]}YE-1", f"净利润<br>%变化<br>{str(cy)[-2:]}YE/{str(py)[-2:]}YE-1", f"{cy}年12月31日<br>%CSM增长率<br>{str(cy)[-2:]}YE/{str(py)[-2:]}YE-1", f"%NB CSM<br>增长率<br>{str(cy)[-2:]}YE/{str(py)[-2:]}YE-1", f"CSM摊销比例<br>(CSM摊销/<br>摊销前的期末CSM)", f"CSM持续率<br>(新业务CSM/<br>CSM摊销)", f"保险服务收入<br>%变化<br>{str(cy)[-2:]}YE/{str(py)[-2:]}YE-1", f"保险合同净负债余额<br>%变化<br>{str(cy)[-2:]}YE/{str(py)[-2:]}YE-1"]
        current_hl = str(highlight_co).strip()
        
        html = "<table style='width:100%; border-collapse: collapse; font-family: sans-serif; font-size: 11px; margin-bottom: 15px;'>"
        html += "<tr style='background-color: #00338D; color: white; text-align: center; font-weight: bold;'>"
        for idx, h in enumerate(headers):
            b_bottom = "border-bottom: none;" if (idx > 0 and (idx - 1) < len(col_names) and str(col_names[idx-1]).strip() == current_hl) else "border-bottom: 1.5px solid white;"
            html += f"<th style='padding: 6px 4px; { 'text-align: left;' if idx == 0 else 'text-align: center;' } border: 1.5px solid white; {b_bottom}'>{h}</th>"
        html += "</tr>"
        
        for i, co in enumerate(col_names):
            row_vals = [c1[i], c2[i], c3[i], c4[i], c5[i], c6[i], c7[i], c8[i]]
            is_hl = (str(co).strip() == current_hl)
            bg = HL_BOX_FILL   if is_hl else ("white" if i % 2 == 0 else "#F8F9FA")
            base = f"background-color: {bg}; padding: 4px 4px; font-size: 11px; " + ("border-top: 1.5px solid #00338D; border-bottom: 1.5px solid #00338D; font-weight: bold;" if is_hl else "border: 1px solid #EAEAEA;")
            s_first, s_mid, s_last = base + f"text-align: left; color: #333333; { 'border-left: 1.5px solid #00338D;' if is_hl else '' }", base + f"text-align: center; { 'color: #00338D; border-left: none; border-right: none;' if is_hl else 'color: #444444;' }", base + f"text-align: center; { 'color: #00338D; border-right: 1.5px solid #00338D; border-left: none;' if is_hl else 'color: #444444;' }"
            html += f"<tr><td style='{s_first}'>{co}</td>" + "".join([f"<td style='{s_last if idx==len(row_vals)-1 else s_mid}'>{v}</td>" for idx, v in enumerate(row_vals)]) + "</tr>"        
        return html + "</table>"

    # --- 2.柱状图和比例环 ---
    def create_profit_mixed_chart(df, cos, title_prefix, show_lbl, gap, div, unit_str, highlight_co="无"):
        c, cy, py, n = {'PI': '#FFD6EB', 'PS': '#00B8F5', 'CI': '#FD349C', 'CS': '#00338D'}, latest_year, prev_year, len(cos)
        sub_titles = [f"<span style='color:#00338D;'><b>{co}</b></span>" for co in cos]
        # 🌟 动态计算安全间距：如果公司数量过多，自动压缩间距，强制总间距最多占用总宽度的 80%
        safe_h_gap = 0.03 if n <= 1 else min(0.03, 0.8 / (n - 1))  
        fig = make_subplots(
            rows=2, cols=n, 
            row_heights=[0.18, 0.82], 
            subplot_titles=sub_titles, 
            shared_yaxes=True, 
            vertical_spacing=0.15, 
            horizontal_spacing=safe_h_gap,  # 🌟 用安全变量替换原来写死的 0.03
            specs=[[{"type": "domain"}]*n, [{"type": "xy"}]*n]
        )
        txt = dict(textposition='outside', textfont=dict(size=12), constraintext='none', cliponaxis=False) if show_lbl else dict()        
        
        for i, co in enumerate(cos):
            col, df_c = i + 1, df[df['公司'] == co]
            v = lambda y, f: df_c[(df_c['报告年份'].astype(str)==str(y)) & (df_c['字段名']==f)]['(百万)人民币'].sum() / div
            pi, ps, ci, cs = v(py, '投资利润'), v(py, '保险利润'), v(cy, '投资利润'), v(cy, '保险利润')
            tot_c = ci + cs
            pct_ci = (ci / tot_c * 100) if tot_c != 0 else 0
            pct_cs = (cs / tot_c * 100) if tot_c != 0 else 0           
            # 1. 设定圆环的“物理显示大小” (只为画图，不代表真实数据)
            if ci >= 0 and cs >= 0:
                v_ci, v_cs = ci, cs
            elif ci > 0 and cs <= 0:
                v_ci, v_cs = 1, 0  # 投资为正，保险为负：投资在视觉上占满 100%
            elif ci <= 0 and cs > 0:
                v_ci, v_cs = 0, 1  # 保险为正，投资为负：保险在视觉上占满 100%
            else:
                v_ci, v_cs = abs(ci), abs(cs) # 都是亏损：用绝对值展示“亏损的构成比例”
                
            # 2. 定制悬浮提示框，把真实的数值和奇异的比例写进去
            htext = [
                f"投资利润: {ci:.1f}<br>实际占比: {pct_ci:.1f}%",
                f"保险利润: {cs:.1f}<br>实际占比: {pct_cs:.1f}%"
            ]
            
            fig.add_trace(go.Pie(
                labels=['投资', '保险'], 
                values=[v_ci, v_cs], 
                marker_colors=[c['CI'], c['CS']], 
                hole=0.75, 
                textinfo='none', 
                hovertext=htext,       # 🌟 挂载真实的悬浮数据
                hoverinfo='label+text',# 🌟 强制显示我们写好的文本
                showlegend=False, 
                sort=False
            ), row=1, col=col)

            if i == 0: 
                for nm, cl in [(f"{py}YE 投资利润", c['PI']), (f"{py}YE 保险利润", c['PS']), (f"{cy}YE 投资利润", c['CI']), (f"{cy}YE 保险利润", c['CS'])]:
                    fig.add_trace(go.Scatter(x=[None], y=[None], mode='markers', marker=dict(color=cl, symbol='square', size=15), name=nm, showlegend=True), row=2, col=col)
            
            x_lbl = [f"{py}YE", f"{cy}YE"]
            fig.add_trace(go.Bar(x=x_lbl, y=[pi, ci], marker_color=[c['PI'], c['CI']], text=[f"{pi:.0f}", f"{ci:.0f}"] if show_lbl else None, showlegend=False, **txt), row=2, col=col)
            fig.add_trace(go.Bar(x=x_lbl, y=[ps, cs], marker_color=[c['PS'], c['CS']], text=[f"{ps:.0f}", f"{cs:.0f}"] if show_lbl else None, showlegend=False, **txt), row=2, col=col)
            
            is_hl = (str(co).strip() == str(highlight_co).strip())
            bg_fill = HL_BOX_FILL   if is_hl else "rgba(0,0,0,0)"
            line_dict = dict(color=HL_BOX_LINE, width=1.5) if is_hl else dict(color="lightgray", width=1)            
            fig.add_shape(type="rect", xref=f"x{col if col > 1 else ''} domain", yref="paper", x0=-0.05, x1=1.05, y0=0, y1=1.13, fillcolor=bg_fill, line=line_dict, layer="below")       
        
        for ann in fig.layout.annotations: 
            if "<b>" in str(ann.text): ann.update(y=1.05, font=dict(size=13, weight="bold"))
        
        fig.update_layout(barmode='group', bargap=gap, bargroupgap=0.0, height=500, margin=dict(t=50, b=100, l=40, r=20), paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', legend=dict(orientation="h", x=0.5, xanchor="center", y=-0.15))
        fig.update_xaxes(type='category', showgrid=False, tickangle=0, tickfont=dict(color='gray'))
        fig.update_yaxes(showgrid=False, zeroline=True, zerolinecolor='lightgray', showticklabels=False)            
        return fig
   
    # --- 3.保险服务收入等单指标柱状图 ---
    def create_kpmg_chart(df, field_name, title_prefix, show_labels, pct_font_size, global_gap):
        d = df[df['公司'].isin(selected_cos)].copy()
        d['报告年份'] = d['报告年份'].astype(str).str.replace(".0", "", regex=False)
        if field_name == "保险服务业绩":
            df_inc = d[d['字段名'] == "保险服务收入合计"].set_index(['公司', '报告年份'])['(百万)人民币']
            df_exp = d[d['字段名'] == "保险服务费用合计"].set_index(['公司', '报告年份'])['(百万)人民币']
            res_series = df_inc - df_exp
            df_plot = res_series.reset_index(); df_plot.columns = ['公司', '报告年份', 'value']
        else:
            df_plot = d[d['字段名'] == field_name].copy().rename(columns={'(百万)人民币': 'value'})

        df_plot['value'] = df_plot['value'] / divisor
        y_old, y_new = str(prev_year), str(latest_year)
        fig = go.Figure()
        color_map = {y_old: "#1E49E2", y_new: "#00338D"}
        all_max = df_plot['value'].max() if not df_plot['value'].empty else 100
        fixed_offset = all_max * 0.2 

        for yr in [y_old, y_new]:
            df_yr = df_plot[df_plot['报告年份'] == yr].set_index('公司').reindex(selected_cos).reset_index()
            fig.add_trace(go.Bar(name=f"{yr}YE", x=df_yr['公司'], y=df_yr['value'], marker_color=color_map[yr], text=[f"{v:.1f}" if pd.notna(v) else "" for v in df_yr['value']] if show_labels else None, textposition='outside', textfont=dict(size=12)))

        df_old, df_new = df_plot[df_plot['报告年份'] == y_old].set_index('公司'), df_plot[df_plot['报告年份'] == y_new].set_index('公司')
        for co in selected_cos:
            if co in df_old.index and co in df_new.index:
                v_old, v_new = df_old.loc[co, 'value'], df_new.loc[co, 'value']
                if pd.notna(v_old) and pd.notna(v_new) and v_old != 0:
                    pct = (v_new - v_old) / abs(v_old)
                    color, arrow = ("#FD349C", "↗") if pct >= 0 else ("#269924", "↘")
                    fig.add_annotation(x=co, y=max(v_old, v_new) + fixed_offset, text=f"<b>{arrow} {pct:.1%}</b>", showarrow=False, font=dict(color=color, size=pct_font_size), xshift=15)
                    
        if highlight_co in selected_cos:
            idx = selected_cos.index(highlight_co)
            fig.add_shape(type="rect", xref="x", yref="paper", x0=idx - 0.45, x1=idx + 0.45, y0=-0.08, y1=1, fillcolor=HL_BOX_FILL, line=dict(color=HL_BOX_LINE, width=1.5), layer="below")
            
        fig.update_layout(barmode='group', paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', bargroupgap=0.0, bargap=global_gap, legend=dict(orientation="h", yanchor="bottom", y=1.03, xanchor="right", x=1), margin=dict(t=40, b=40, l=20, r=20), height=450)
        fig.update_xaxes(showgrid=False, zeroline=False, showline=False)
        fig.update_yaxes(showgrid=False, zeroline=True, zerolinecolor="#E0E0E0", zerolinewidth=1, showline=False)
        return fig

    # --- 4.保险业务构成1 ---
    def create_kpmg_composition_chart(df, fields, title_prefix, show_labels, label_size, bar_width, co_font_size, highlight_co="无"):
        d = df[df['公司'].isin(selected_cos)].copy()
        d_struct = d[d['字段名'].isin(fields)].copy()
        d_struct['报告年份'] = d_struct['报告年份'].astype(str).str.replace(".0", "", regex=False) + "YE"
        available_cos = [co for co in selected_cos if co in d_struct['公司'].unique()]
        if not available_cos: return go.Figure().add_annotation(text="无对应数据", showarrow=False)
        titles = [f"<b>{co}</b>" for co in available_cos]
        fig = make_subplots(rows=1, cols=len(available_cos), shared_yaxes=True, column_titles=titles, horizontal_spacing=0.015)
        short_names = {fields[0]: "采用保费分配法", fields[1]: "未采用保费分配法"}
        
        for i, co in enumerate(available_cos):
            d_co = d_struct[d_struct['公司'] == co].pivot(index='报告年份', columns='字段名', values='(百万)人民币')
            for f in fields:
                if f not in d_co.columns: d_co[f] = 0
            d_co = d_co.reindex(sorted(d_co.index.tolist()))
            d_co['Total'] = d_co.sum(axis=1).replace(0, 1)
            paa_val, non_paa_val = d_co[fields[0]] / d_co['Total'] * 100, d_co[fields[1]] / d_co['Total'] * 100
            
            fig.add_trace(go.Bar(x=d_co.index, y=non_paa_val, name=short_names[fields[1]] if i == 0 else None, marker_color="rgb(227,207,251)", text=[f"{v:.0f}%" if v > 0 else "" for v in non_paa_val] if show_labels else None, textposition='inside', insidetextanchor='middle', textfont=dict(size=label_size), constraintext='none', textangle=0, width=bar_width, showlegend=(i == 0), legendgroup="group2", hoverinfo="none"), row=1, col=i+1)
            fig.add_trace(go.Bar(x=d_co.index, y=paa_val, name=short_names[fields[0]] if i == 0 else None, marker_color="#510DBC", text=[f"{v:.0f}%" if v > 0 else "" for v in paa_val] if show_labels else None, textposition='inside', insidetextanchor='middle', textfont=dict(size=label_size), constraintext='none', textangle=0, width=bar_width, showlegend=(i == 0), legendgroup="group1", hoverinfo="none"), row=1, col=i+1)

            is_hl = (str(co).strip() == str(highlight_co).strip())
            bg_fill = HL_BOX_FILL   if is_hl else "rgba(0,0,0,0)"
            line_dict = dict(color=HL_BOX_LINE, width=1.5) if is_hl else dict(color="#E0E0E0", width=1)
            fig.add_shape(type="rect", xref="x domain", yref="y domain", x0=-0.06, x1=1.06, y0=-0.1, y1=1.12, fillcolor=bg_fill, line=line_dict, layer="above", row=1, col=i+1)

        fig.update_layout(barmode='stack', height=400, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', margin=dict(t=50, b=80, l=10, r=10), legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5))
        for ann in fig.layout.annotations:
            if "<b>" in str(ann.text): ann.update(y=1.03, font=dict(size=co_font_size, color="#00338D"))
        return fig


    # --- 5.保险业务构成2 ---
    def create_kpmg_multi_composition_chart(df, field_map, color_map, title_prefix, show_labels, label_size, bar_width, co_font_size, highlight_co="无"):
        fields = list(field_map.keys())
        d = df[df['公司'].isin(selected_cos)].drop_duplicates(subset=['公司', '报告年份', '字段名'], keep='first').copy()
        d['报告年份'] = d['报告年份'].astype(str).str.replace(".0", "", regex=False)
        
        d_p = d[d['字段名'].isin(fields)].pivot_table(index=['box', '报告年份'], columns='字段名', values='(百万)人民币', aggfunc='first').fillna(0) if d.empty else d[d['字段名'].isin(fields)].pivot_table(index=['公司', '报告年份'], columns='字段名', values='(百万)人民币', aggfunc='first').fillna(0)
        
        # 安全计算 Total，防止某些字段在 df 中完全不存在而报错
        exist_dp_fs = [f for f in fields if f in d_p.columns]
        d_p['Total'] = d_p[exist_dp_fs].abs().sum(axis=1).replace(0, 1) if exist_dp_fs else 1
        for f in fields: 
            d_p[field_map[f]] = d_p[f] / d_p['Total'] * 100 if f in d_p.columns else 0
        
        all_yrs = sorted(d['报告年份'].unique())
        df_avg = d_p[[field_map[f] for f in fields]].groupby('报告年份').mean().reindex(all_yrs[-2:] if len(all_yrs)>=2 else all_yrs)
        df_avg.index = [f"{y}YE" for y in df_avg.index] 
    
        av_cos = [co for co in selected_cos if co in d['公司'].unique()]
        if not av_cos: return go.Figure(), pd.DataFrame()
        titles = [f"<b>{co}</b>" for co in av_cos]
        fig = make_subplots(rows=1, cols=len(av_cos), shared_yaxes=True, horizontal_spacing=0.015, column_titles=titles)
        
        for i, co in enumerate(av_cos):
            d_co = d[d['公司']==co].pivot(index='报告年份', columns='字段名', values='(百万)人民币').reindex(all_yrs).fillna(0)
            
            # 🌟 核心修改 1：计算真实的绝对值 Total，用来判断这一年是不是全为 0（未披露）
            exist_co_fs = [f for f in fields if f in d_co.columns]
            raw_total = d_co[exist_co_fs].abs().sum(axis=1) if exist_co_fs else pd.Series(0, index=d_co.index)
            d_co['Total'] = raw_total.replace(0, 1)
            
            for f, d_n in field_map.items():
                val = d_co.get(f, 0) / d_co['Total'] * 100
                txt_c = "white" if any(x in color_map[f] for x in ["30, 73, 226", "114, 19, 234", "0, 163, 161"]) else "black"
                fig.add_trace(go.Bar(x=[f"{y}YE" for y in d_co.index], y=val, name=d_n if i==0 else None, marker_color=color_map[f], text=[f"{v:.0f}%" if abs(v) >= 1 else "" for v in val] if show_labels else None, textangle=0, textposition='inside', insidetextanchor='middle', textfont=dict(size=label_size, color=txt_c), constraintext='none', cliponaxis=False, width=bar_width, showlegend=(i==0), legendgroup=f), row=1, col=i+1)
            
            # 🌟 核心修改 2：新增灰色占位柱！如果没有数据，高度设为 100 (填满柱子)，有数据则设为 0 (自动隐藏)
            missing_y = [100 if v == 0 else 0 for v in raw_total]
            missing_t = ["未披露" if v == 0 else "" for v in raw_total]
            fig.add_trace(go.Bar(
                x=[f"{y}YE" for y in d_co.index], 
                y=missing_y, 
                marker_color="#CDCDCD", 
                text=missing_t, 
                textangle=0, textposition='inside', insidetextanchor='middle', 
                textfont=dict(size=12, color="white"), 
                constraintext='none', cliponaxis=False, width=bar_width, 
                showlegend=False, hoverinfo='skip'
            ), row=1, col=i+1)
            
            is_hl = (str(co).strip() == str(highlight_co).strip())
            bg_fill = HL_BOX_FILL   if is_hl else "rgba(0,0,0,0)"
            line_dict = dict(color=HL_BOX_LINE, width=1.5) if is_hl else dict(color="rgba(0,0,0,0)", width=0)
            fig.add_shape(type="rect", xref="x domain", yref="y domain", x0=-0.06, x1=1.06, y0=-0.1, y1=1.08, fillcolor=bg_fill, line=line_dict, layer="above", row=1, col=i+1)
                    
        fig.update_layout(barmode='relative', height=400, margin=dict(t=50, b=80, l=20, r=20), legend=dict(orientation="h", yanchor="top", y=-0.17, xanchor="center", x=0.5, font=dict(size=10)), plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        for ann in fig.layout.annotations:
            if "<b>" in str(ann.text): ann.update(y=1, font=dict(size=co_font_size, color="#00338D"))
        return fig, df_avg
    
    
    
    def create_kpmg_exp_chart(df, field_map, color_map, title_prefix, show_labels, label_size, bar_width, co_font_size, highlight_co="无"):
        fields = list(field_map.keys())
        d = df[df['公司'].isin(selected_cos)].drop_duplicates(subset=['公司', '报告年份', '字段名'], keep='first').copy()
        d['报告年份'] = d['报告年份'].astype(str).str.replace(".0", "", regex=False)
        
        d_p = d[d['字段名'].isin(fields)].pivot_table(index=['box', '报告年份'], columns='字段名', values='(百万)人民币', aggfunc='first').fillna(0) if d.empty else d[d['字段名'].isin(fields)].pivot_table(index=['公司', '报告年份'], columns='字段名', values='(百万)人民币', aggfunc='first').fillna(0)
        d_p['Total'] = d_p[fields].abs().sum(axis=1).replace(0, 1)
        for f in fields: d_p[field_map[f]] = d_p[f] / d_p['Total'] * 100
        
        all_yrs = sorted(d['报告年份'].unique())
        df_avg = d_p[[field_map[f] for f in fields]].groupby('报告年份').mean().reindex(all_yrs[-2:] if len(all_yrs)>=2 else all_yrs)
        df_avg.index = [f"{y}YE" for y in df_avg.index] 
    
        av_cos = [co for co in selected_cos if co in d['公司'].unique()]
        if not av_cos: return go.Figure(), pd.DataFrame()
        titles = [f"<b>{co}</b>" for co in av_cos]
        fig = make_subplots(rows=1, cols=len(av_cos), shared_yaxes=True, horizontal_spacing=0.015, column_titles=titles)
        
        for i, co in enumerate(av_cos):
            d_co = d[d['公司']==co].pivot(index='报告年份', columns='字段名', values='(百万)人民币').reindex(all_yrs).fillna(0)
            d_co['Total'] = d_co[fields].abs().sum(axis=1).replace(0, 1)
            for f, d_n in field_map.items():
                val = d_co.get(f, 0) / d_co['Total'] * 100
                txt_c = "white" if any(x in color_map[f] for x in ["30, 73, 226", "114, 19, 234", "0, 163, 161"]) else "black"
                fig.add_trace(go.Bar(x=[f"{y}YE" for y in d_co.index], y=val, name=d_n if i==0 else None, marker_color=color_map[f], text=[f"{v:.0f}%" if abs(v) >= 1 else "" for v in val] if show_labels else None, textangle=0, textposition='inside', insidetextanchor='middle', textfont=dict(size=label_size, color=txt_c), constraintext='none', cliponaxis=False, width=bar_width, showlegend=(i==0), legendgroup=f), row=1, col=i+1)
            
            is_hl = (str(co).strip() == str(highlight_co).strip())
            bg_fill = HL_BOX_FILL   if is_hl else "rgba(0,0,0,0)"
            line_dict = dict(color=HL_BOX_LINE, width=1.5) if is_hl else dict(color="rgba(0,0,0,0)", width=0)
            fig.add_shape(type="rect", xref="x domain", yref="y domain", x0=-0.06, x1=1.06, y0=-0.1, y1=1.08, fillcolor=bg_fill, line=line_dict, layer="above", row=1, col=i+1)
                    
        # 原有的 layout 更新
        fig.update_layout(barmode='relative', height=400, margin=dict(t=50, b=80, l=10, r=10), legend=dict(orientation="h", yanchor="top", y=-0.17, xanchor="center", x=0.5, font=dict(size=10)), plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        
        # 👇 加上这行代码：在图表上方画一条“拉穿全场”的水平基准线
        fig.add_hline(y=0, line_color="orange", line_width=1.5, layer="below")
        for ann in fig.layout.annotations:
            if "<b>" in str(ann.text): ann.update(y=1, font=dict(size=co_font_size, color="#00338D"))
        return fig, df_avg
  
    #--------利润贡献--------
    def create_profit_composition_chart(df, selected_cos, target_year, show_labels, label_size, bar_width, co_font_size, highlight_co="无"):
        source_fields = ["合同服务边际的摊销","非金融风险调整的变动","亏损部分的确认及转回","采用保费分配法计量的保险合同保险业绩","保险利润","再保净损益"]
        year_str = str(target_year).replace(".0","")
        d_sub = df[(df['报告年份'].astype(str).str.contains(year_str)) & df['公司'].isin(selected_cos) & df['字段名'].isin(source_fields)].copy()
        if d_sub.empty: return None, None

        d_pivot = d_sub.pivot(index='公司', columns='字段名', values='(百万)人民币')
        contrib_val = (d_pivot.get('合同服务边际的摊销', 0) / d_pivot.get('保险利润', 1) * 100)
        contribution_df = pd.DataFrame(contrib_val).transpose()
        contribution_df.index = ["合同服务边际释放的贡献"]

        for f in source_fields:
            if f not in d_pivot.columns: d_pivot[f] = np.nan  # ✅ 用 nan 而不是 0，保留未披露信息

        d_pivot = d_pivot.reindex(selected_cos)

        # 🌟 修复：把 .all() 改成了 .any()，实现“只要缺一个，全盘皆灰”的严格一票否决逻辑
        no_data_cos = set(co for co in selected_cos if co in d_pivot.index and d_pivot.loc[co, source_fields].isna().any())

        d_pivot = d_pivot.fillna(0)  # 填0用于后续计算

        d_pivot['营运偏差及其他_展示'] = d_pivot['保险利润']-d_pivot['合同服务边际的摊销']-d_pivot['非金融风险调整的变动']+d_pivot['亏损部分的确认及转回']-d_pivot['采用保费分配法计量的保险合同保险业绩']-d_pivot['再保净损益']
        d_pivot['合同服务边际的释放_展示'] = d_pivot['合同服务边际的摊销']
        d_pivot['非金融风险调整的变动_展示'] = d_pivot['非金融风险调整的变动']
        d_pivot['亏损部分的确认及转回_展示'] = -d_pivot['亏损部分的确认及转回']
        d_pivot['保费分配法业务净损益_展示'] = d_pivot['采用保费分配法计量的保险合同保险业绩']
        d_pivot['再保净损益_展示'] = d_pivot['再保净损益']

        display_mapping = [
            ("亏损部分的确认及转回_展示","亏损部分的确认及转回","rgb(190,190,190)"),
            ("合同服务边际的释放_展示","合同服务边际的释放","rgb(30,73,226)"),
            ("非金融风险调整的变动_展示","非金融风险调整的变动","rgb(118,210,255)"),
            ("营运偏差及其他_展示","营运偏差及其他","rgb(114,19,214)"),
            ("保费分配法业务净损益_展示","保费分配法业务净损益","rgb(253,52,156)"),
            ("再保净损益_展示","再保净损益","rgb(9,142,126)"),
        ]
        d_pivot['Total'] = d_pivot[[i[0] for i in display_mapping]].abs().sum(axis=1).replace(0,1)

        fig = go.Figure()
        x_indices = list(range(len(d_pivot.index)))
        no_data_x = [list(d_pivot.index).index(co) for co in no_data_cos if co in d_pivot.index]

        # ✅ 先画灰色未披露占位柱。高度设为 100（因为 Y 轴范围是 -45 到 +105）
        if no_data_x:
            fig.add_trace(go.Bar(x=no_data_x, y=[100]*len(no_data_x), marker_color="#CDCDCD", width=bar_width,
                showlegend=False, text=["未披露"]*len(no_data_x), textposition='inside',
                insidetextanchor='middle', textangle=0, textfont=dict(size=label_size, color="white"), constraintext='none'))

        for col_name, legend_name, color in display_mapping:
            vals_pct = (d_pivot[col_name] / d_pivot['Total'] * 100)
            # ✅ 未披露公司值强制为0，不覆盖灰柱
            vals_pct = [0 if co in no_data_cos else v for co, v in zip(d_pivot.index, vals_pct)]
            txt_color = "white" if any(x in color for x in ["30, 73, 226","30,73,226","114, 19, 214","114,19,214","9, 142, 126","9,142,126"]) else "black"
            fig.add_trace(go.Bar(name=legend_name, x=x_indices, y=vals_pct, width=bar_width, marker_color=color,
                text=[f"{v:.0f}%" if abs(v)>=2 else "" for v in vals_pct] if show_labels else None,
                textposition='inside', textangle=0, insidetextanchor='middle',
                textfont=dict(size=label_size, color=txt_color), constraintext='none'))

        fig.update_layout(barmode='relative', height=550, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            margin=dict(t=20, b=80, l=220, r=80),
            legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="right", x=-0.05, traceorder="reversed", font=dict(size=11, color="#00338D")),
            yaxis=dict(side='right', showgrid=False, range=[-45,105], tickmode='array', showticklabels=True,
                tickvals=[-40,-20,0,20,40,60,80,100], ticktext=["-40%","-20%","0%","20%","40%","60%","80%","100%"],
                zeroline=True, zerolinecolor="#F7860C"))

        if highlight_co and highlight_co != "无" and highlight_co in d_pivot.index:
            idx = list(d_pivot.index).index(highlight_co)
            # 🌟 修复常量的缺失，直接写入颜色值
            fig.add_shape(type="rect", xref="x", yref="paper", x0=idx-0.46, x1=idx+0.46, y0=0, y1=1.08, fillcolor="rgba(0, 51, 141, 0.05)", line=dict(color="rgba(0, 51, 141, 0.85)", width=1.5), layer="above")

        fig.update_xaxes(showgrid=False, zeroline=False, tickvals=x_indices,
            ticktext=[f"<span style='font-size:{co_font_size}px;color:#00338D;'><b>{co}</b></span>" for co in d_pivot.index], side="top")
        
        return fig, contribution_df

    # --- 数据预处理补充 (7,8用) ---
    target_fields = ['净投资回报', '承保财务损益', '分出再保险财务损益']
    df_clean = df_filtered[df_filtered['字段名'].isin(target_fields)].drop_duplicates(subset=['公司', '报告年份', '字段名'], keep='first')
    df_pivot_5 = df_clean.pivot_table(index=['公司', '报告年份'], columns='字段名', values='(百万)人民币').fillna(0).reset_index()
    df_pivot_5['承保财务净损益'] = -df_pivot_5.get('承保财务损益',0) - df_pivot_5.get('分出再保险财务损益',0)
    df_pivot_5['投资利润'] = df_pivot_5.get('净投资回报',0) - df_pivot_5['承保财务净损益']
    df_plot_final = df_pivot_5.melt(id_vars=['公司', '报告年份'], value_vars=['净投资回报', '承保财务净损益', '投资利润'], var_name='指标名称', value_name='value')
    df_profit_raw = df_filtered[(df_filtered['字段名'].isin(['投资利润', '保险利润'])) & (df_filtered['公司'].isin(selected_cos))].drop_duplicates(subset=['公司', '报告年份', '字段名'], keep='first').copy()
    df_profit_raw.rename(columns={'字段名': '指标名称', '(百万)人民币': 'value'}, inplace=True)

    # --- 7.投资收益、承保损益单柱图 ---
    def create_simple_kpmg_chart(df_source, field_name, title, show_labels, p_size, g_gap, highlight_co="无"):
        d = df_source[df_source['指标名称'] == field_name].copy()
        d['报告年份'] = d['报告年份'].astype(str).str.replace(".0", "", regex=False)
        d['value'] = d['value'] / divisor 
        fig = go.Figure()
        x_indices = list(range(len(selected_cos)))
        
        for yr, col in zip([str(prev_year), str(latest_year)], ["#FD349C", "#97014F"]):
            df_yr = d[d['报告年份'] == yr].set_index('公司').reindex(selected_cos).reset_index()
            fig.add_trace(go.Bar(name=f"{yr}YE", x=x_indices, y=df_yr['value'], marker_color=col, text=[f"{v:.1f}" if pd.notna(v) else "" for v in df_yr['value']] if show_labels else None, textposition='outside', textfont=dict(size=12)))
            
        all_max = d['value'].max() if not d['value'].empty else 100
        off = all_max * 0.3
        df_old, df_new = d[d['报告年份'] == str(prev_year)].set_index('公司'), d[d['报告年份'] == str(latest_year)].set_index('公司')
        
        for i, co in enumerate(selected_cos):
            if co in df_old.index and co in df_new.index:
                v0, v1 = df_old.loc[co, 'value'], df_new.loc[co, 'value']
                if pd.notna(v0) and v0 != 0:
                    pct = (v1 - v0) / abs(v0)
                    arr, col = ("↗", "#FD349C") if pct >= 0 else ("↘", "#269924")
                    fig.add_annotation(x=i, y=max(v0, v1) + off, text=f"<b>{arr} {pct:.1%}</b>", showarrow=False, font=dict(color=col, size=p_size), xshift=10)            
        
        fig.update_layout(barmode='group', bargroupgap=0.0, bargap=g_gap, margin=dict(t=40, b=60, l=20, r=20), height=500, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', legend=dict(orientation="h", yanchor="bottom", y=1.05, xanchor="right", x=1))
        
        current_hl = str(highlight_co).strip()
        clean_cos = [str(c).strip() for c in selected_cos]
        if current_hl in clean_cos:
            idx = clean_cos.index(current_hl)
            fig.add_shape(type="rect", xref="x", yref="paper", x0=idx - 0.46, x1=idx + 0.46, y0=-0.12, y1=1.05, fillcolor=HL_BOX_FILL  , line=dict(color=HL_BOX_LINE, width=1.5), layer="above")
            
        x_labels = [f"<span style='color:#00338D;'><b>{co}</b></span>" for co in selected_cos]
        fig.update_xaxes(showgrid=False, zeroline=False, tickvals=x_indices, ticktext=x_labels)
        fig.update_yaxes(showgrid=False, zeroline=True, zerolinecolor="#E0E0E0", zerolinewidth=1)
        return fig

    # --- 9.税前和净利润补充计算 ---
    df_tax_sub = df_filtered[(df_filtered['字段名'].isin(['税前利润总额', '净利润'])) & (df_filtered['公司'].isin(selected_cos))].drop_duplicates(subset=['公司', '报告年份', '字段名']).copy()
    df_tax_pivot = df_tax_sub.pivot_table(index=['公司', '报告年份'], columns='字段名', values='(百万)人民币').fillna(0).reset_index()
    df_tax_pivot['有效税率'] = np.where(df_tax_pivot['税前利润总额'] != 0, (df_tax_pivot['税前利润总额'] - df_tax_pivot['净利润']) / df_tax_pivot['税前利润总额'], 0)
    df_tax_pivot['报告年份'] = df_tax_pivot['报告年份'].astype(str).str.replace(".0", "", regex=False) + "YE"

    def create_tax_subplot_chart(df_pivot, selected_cos, show_labels, bar_width, label_size, co_font_size, co_y_offset, highlight_co="无"):
        av_cos = [c for c in selected_cos if c in df_pivot['公司'].unique()]
        if not av_cos: return go.Figure()
        
        cols = ['税前利润总额', '净利润']
        g_max, g_min = df_pivot[cols].max().max()/divisor, df_pivot[cols].min().min()/divisor
        # 🌟 修复 1：动态计算 Y 轴底部边界。如果有负数，底部延伸；否则为 0
        y_top, y_bot, lbl_y, ph_h = g_max*1.25, min(0, g_min*1.2), g_max*1.15, g_max*0.4
        all_yrs = sorted(df_pivot['报告年份'].astype(str).unique()) # 提取全局年份兜底
        
        n = len(av_cos)
        fig = make_subplots(rows=1, cols=n, shared_yaxes=True, column_titles=[f"<b><span style='color:#00338D;'>{c}</span></b>" for c in av_cos], horizontal_spacing=0.03 if n<=1 else min(0.03, 0.8/(n-1)))
        
        for nm, c in zip(cols, ['#1E49E2', '#C7A0F7']): fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", marker=dict(symbol="square", size=12, color=c), name=nm, showlegend=True), row=1, col=1)

        for i, co in enumerate(av_cos):
            ci, d = i+1, df_pivot[df_pivot['公司']==co].sort_values('报告年份')
            x, yp, yn, tr = d['报告年份'].astype(str).tolist(), (d[cols[0]]/divisor).tolist(), (d[cols[1]]/divisor).tolist(), d['有效税率'].tolist()
            pr, nr, pt, nt, my, mt = [], [], [], [], [], []
            
            for vp, vn in zip(yp, yn):
                if (pd.isna(vp) or vp==0) and (pd.isna(vn) or vn==0):
                    pr.append(0); nr.append(0); pt.append(""); nt.append(""); my.append(ph_h); mt.append("未披露")
                else:
                    pr.append(vp); nr.append(vn); pt.append(f"{vp:.1f}" if show_labels and pd.notna(vp) else ""); nt.append(f"{vn:.1f}" if show_labels and pd.notna(vn) else ""); my.append(0); mt.append("")
                    
            for yy, c, t, og in [(pr, '#1E49E2', pt, 1), (nr, '#C7A0F7', nt, 2)]: fig.add_trace(go.Bar(x=x, y=yy, marker_color=c, text=t, textposition='outside', textfont=dict(size=label_size), width=bar_width, offsetgroup=og, showlegend=False, cliponaxis=False), row=1, col=ci)
            fig.add_trace(go.Bar(x=x, y=my, marker_color="#CDCDCD", text=mt, textposition='inside', insidetextanchor='middle', textfont=dict(size=12, color="white"), width=bar_width*2, showlegend=False, cliponaxis=False, hoverinfo='skip'), row=1, col=ci)
            
            for j, yr in enumerate(x):
                if pd.notna(tr[j]) and mt[j] == "": fig.add_annotation(x=yr, y=lbl_y, text=f"<b>{tr[j]:.0%}</b>", showarrow=False, font=dict(size=label_size+2, color="#97014F" if tr[j]>=0 else "#269924"), xref=f"x{ci}" if ci>1 else "x", yref="y1")
            
            hl = (str(co).strip() == str(highlight_co).strip())
            fig.add_shape(type="rect", xref=f"x{ci} domain" if ci>1 else "x domain", yref="paper", x0=-0.05, x1=1.05, y0=-0.12, y1=1.1, line=dict(color=HL_BOX_LINE if hl else "rgba(200, 200, 200, 0.3)", width=1.5 if hl else 1), fillcolor=HL_BOX_FILL   if hl else "rgba(0,0,0,0)", layer="above")

        fig.update_layout(height=550, margin=dict(t=40, b=100, l=20, r=20), plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5))
        for i in range(1, n+1):
            # 🌟 修复 2：强行写入 categoryorder 和 categoryarray，锁死年份显示
            fig.update_xaxes(type='category', categoryorder='array', categoryarray=all_yrs, showline=True, linecolor='lightgray', linewidth=1, showgrid=False, tickfont=dict(size=10), zeroline=False, ticks="", ticklen=0, row=1, col=i)
            # 应用包含负数的新范围，开启 zeroline 分界线
            fig.update_yaxes(showline=False, showgrid=False, showticklabels=False, zeroline=True, zerolinecolor="#E0E0E0", range=[y_bot, y_top], row=1, col=i)
        for a in fig.layout.annotations:
            if "span" in a.text: a.update(y=co_y_offset, font_size=co_font_size)
            
        return fig

    # --- 10.综合收益变动趋势补充 ---
    df_fin_raw = df_filtered[(df_filtered['字段名'].isin(['净利润', '其他综合收益', '综合收益总额'])) & (df_filtered['公司'].isin(selected_cos))].drop_duplicates(subset=['公司', '报告年份', '字段名'], keep='first').copy()
    df_fin_raw.rename(columns={'字段名': '指标名称', '(百万)人民币': 'value'}, inplace=True)
    
    def create_financial_trend_chart_v5(df_source, field_name, title, show_labels, p_size, g_gap, highlight_co="无"):
        d = df_source[df_source['指标名称'] == field_name].copy()
        d['报告年份'], d['value'] = d['报告年份'].astype(str).str.replace(".0", "", regex=False), d['value'] / divisor 
        fig, x_idx, hl_co = go.Figure(), list(range(len(selected_cos))), str(highlight_co).strip()
        
        for yr, col in zip([str(prev_year), str(latest_year)], ["rgb(15, 101, 253)", "rgb(0, 51, 141)"]):
            df_yr = d[d['报告年份'] == yr].set_index('公司').reindex(selected_cos).reset_index()
            fig.add_trace(go.Bar(name=f"{yr}YE", x=x_idx, y=df_yr['value'], marker_color=col, text=[f"{v:.1f}" if pd.notna(v) and v!=0 else "" for v in df_yr['value']] if show_labels else None, textposition='outside', textfont=dict(size=12), cliponaxis=False))
            
        max_val = d['value'].max() - d['value'].min() if not d['value'].dropna().empty else 100
        off, df_old, df_new = max_val * 0.12, d[d['报告年份'] == str(prev_year)].set_index('公司'), d[d['报告年份'] == str(latest_year)].set_index('公司')
        
        for i, co in enumerate(selected_cos):
            if co in df_old.index and co in df_new.index and pd.notna(df_old.loc[co, 'value']) and df_old.loc[co, 'value'] != 0:
                v0, v1 = df_old.loc[co, 'value'], df_new.loc[co, 'value']
                pct = (v1 - v0) / abs(v0)
                fig.add_annotation(x=i, y=(max(v0, v1)+off if v1>=0 else min(v0, v1)-off), text=f"<b>{'↗' if pct>=0 else '↘'} {pct:.1%}</b>", showarrow=False, font=dict(color="#FD349C" if pct>=0 else "#269924", size=p_size), xshift=10, valign="bottom" if v1>=0 else "top")
                
        fig.update_layout(barmode='group', bargroupgap=0, bargap=g_gap, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', margin=dict(t=40, b=60, l=40, r=40), height=500, legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        
        if hl_co in [str(c).strip() for c in selected_cos]:
            idx = [str(c).strip() for c in selected_cos].index(hl_co)
            fig.add_shape(type="rect", xref="x", yref="paper", x0=idx-0.46, x1=idx+0.46, y0=-0.12, y1=1.05, fillcolor=HL_BOX_FILL  , line=dict(color=HL_BOX_LINE, width=1.5), layer="above")
    
        fig.update_xaxes(showline=False, showgrid=False, zeroline=False, tickvals=x_idx, ticktext=[f"<span style='color:#00338D;'><b>{co}</b></span>" for co in selected_cos], ticks="", ticklen=0)
        fig.update_yaxes(showgrid=False, zeroline=True)
        return fig

    # --- 11.资产端分类 ---
    def create_asset_composition_chart(df, selected_cos, field_map, color_map, title_prefix, show_labels, label_size, bar_width, co_font_size, highlight_co="无"):
        fields = list(field_map.keys())
        d_struct = df[(df['公司'].isin(selected_cos)) & (df['字段名'].isin(fields))].copy()
        d_struct['报告年份'] = d_struct['报告年份'].astype(str).str.replace(".0", "", regex=False) + "YE"
        all_yrs, av_cos = [f"{prev_year}YE", f"{latest_year}YE"], [co for co in selected_cos if co in d_struct['公司'].unique()]
        if not av_cos: return go.Figure()

        hl_co = str(highlight_co).strip()
        fig = make_subplots(rows=1, cols=len(av_cos), shared_yaxes=True, column_titles=[f"<span style='color:#00338D;'><b>{co}</b></span>" for co in av_cos], horizontal_spacing=0.015)
        for i, co in enumerate(av_cos):
            d_co = d_struct[d_struct['公司'] == co].pivot(index='报告年份', columns='字段名', values='(百万)人民币').reindex(all_yrs).fillna(0)
            for f in fields:
                if f not in d_co.columns: d_co[f] = 0
            raw_total = d_co.sum(axis=1)
            d_co['T'] = raw_total.replace(0, 1) 
            for fn in fields:
                val_pct = d_co[fn] / d_co['T'] * 100
                fig.add_trace(go.Bar(x=d_co.index, y=val_pct, name=field_map[fn] if i==0 else None, marker_color=color_map[fn], text=[f"{v:.0f}%" if show_labels and raw_total.iloc[idx]>0 else "" for idx, v in enumerate(val_pct)], textposition='inside', insidetextanchor='middle', textfont=dict(size=label_size, color="white" if fn in ["FVOCI", "指定FVOCI", "AC"] else "black"), constraintext='none', textangle=0, cliponaxis=False, width=bar_width, showlegend=(i==0), legendgroup=fn, hoverinfo="skip"), row=1, col=i+1)
            fig.add_trace(go.Bar(x=d_co.index, y=[100 if t==0 else 0 for t in raw_total], name="未披露", marker_color="#CDCDCD", text=["未披露" if t==0 else "" for t in raw_total], textposition='inside', insidetextanchor='middle', textfont=dict(size=label_size, color="white"), constraintext='none', textangle=0, cliponaxis=False, width=bar_width, showlegend=False, hoverinfo="skip"), row=1, col=i+1)
            
            is_hl = (str(co).strip() == hl_co)
            fig.add_shape(type="rect", xref="x domain", yref="y domain", x0=-0.06, x1=1.06, y0=-0.12, y1=1.12, fillcolor=HL_BOX_FILL   if is_hl else "rgba(0,0,0,0)", line=dict(color=HL_BOX_LINE, width=1.5) if is_hl else dict(color="rgba(0,0,0,0)", width=0), layer="above", row=1, col=i+1)

        fig.update_layout(barmode='stack', height=550, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', uniformtext=dict(minsize=label_size, mode='show'), margin=dict(t=50, b=120, l=40, r=40), legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5, font=dict(size=11), itemsizing="constant"))
        for i in range(1, len(av_cos) + 1):
            fig.update_xaxes(showgrid=False, showline=False, zeroline=False, tickfont=dict(size=10), ticks="", ticklen=0, row=1, col=i)
            fig.update_yaxes(showgrid=False, range=[0, 101], showline=False, zeroline=False, tickvals=[0, 25, 50, 75, 100], ticktext=["0%", "25%", "50%", "75%", "100%"], showticklabels=(i==1), row=1, col=i)
        for ann in fig.layout.annotations:
            ann.update(y=ann.y + 0.05, font_size=co_font_size)
        return fig
  
    # --- 12. OCI变动分析 ---
    def create_oci_chart(df_raw, year, title, show_labels, co_font_size, bar_gap, selected_cos, highlight_co="无"):
        oci_fields = ['可转损益OCI合计', '不可转损益OCI合计', '净利润', '综合收益总额']
        d_sub = df_raw[(df_raw['字段名'].isin(oci_fields)) & (df_raw['报告年份'].astype(str) == str(year).replace(".0", ""))].drop_duplicates(subset=['公司', '报告年份', '字段名'], keep='last').copy()
        if d_sub.empty: return go.Figure().update_layout(paper_bgcolor='rgba(0,0,0,0)')
        
        df_pivot = d_sub.pivot_table(index='公司', columns='字段名', values='(百万)人民币', aggfunc='sum').fillna(0)
        for f in oci_fields:
            if f not in df_pivot.columns: df_pivot[f] = 0.0
        df_pivot = df_pivot / divisor
        df_pivot['other'] = df_pivot['综合收益总额'] - df_pivot['净利润'] - df_pivot['可转损益OCI合计'] - df_pivot['不可转损益OCI合计']
        mc = {"净利润": {"c": "rgb(0,176,240)", "n": "净利润"}, "可转损益OCI合计": {"c": "rgb(253,52,156)", "n": "可转损益OCI变动+FVOCI债券公允价值变动"}, "不可转损益OCI合计": {"c": "rgb(114,19,234)", "n": "不可转损益负债OCI变动+FVOCI股权公允价值变动"}, "other": {"c": "rgb(127,127,127)", "n": "其他"}, "综合收益总额": {"c": "rgb(172,234,255)", "n": "综合收益总额"}}
        
        av_cos, hl_co = [c for c in selected_cos if c in df_pivot.index], str(highlight_co).strip()
        if not av_cos: return go.Figure()
        all_vals = df_pivot[['净利润', '可转损益OCI合计', '不可转损益OCI合计', 'other', '综合收益总额']].values.flatten()
        max_val, min_val = np.nanmax(all_vals), np.nanmin(all_vals)
        buffer = (max_val - min_val) * 0.8
        y_range = [min_val - buffer if min_val < 0 else min_val - abs(min_val)*0.3, max_val + buffer]
        
        fig = make_subplots(rows=1, cols=len(av_cos), shared_yaxes=True, horizontal_spacing=0.015, subplot_titles=[f"<b><span style='color:#00338D;'>{co}</span></b>" for co in av_cos])
        for col_idx, co in enumerate(av_cos):
            for m_key, m_info in mc.items():
                val = df_pivot.loc[co].get(m_key, 0)
                fig.add_trace(go.Bar(name=m_info["n"], x=[m_key], y=[val], text=[f"{val:.0f}" if (show_labels and val!=0) else ""], textposition='outside', textfont=dict(size=11, color='#00338D'), marker_color=m_info["c"], width=0.8, legendgroup=m_key, showlegend=(col_idx==0), cliponaxis=False, constraintext='none'), row=1, col=col_idx+1)
            
            is_hl = (str(co).strip() == hl_co)
            bg_fill = HL_BOX_FILL   if is_hl else "rgba(240, 240, 240, 0.35)"
            line_dict = dict(color=HL_BOX_LINE, width=1.5) if is_hl else dict(color="rgba(210, 210, 210, 0.6)", width=1)
            fig.add_shape(type="rect", xref="x domain", yref="y domain", x0=-0.04, x1=1.04, y0=-0.04, y1=1.13, fillcolor=bg_fill, line=line_dict, layer="above" if is_hl else "below", row=1, col=col_idx+1)
            fig.update_xaxes(showticklabels=False, showline=False, zeroline=False, showgrid=False, ticks="", ticklen=0, row=1, col=col_idx+1)
            
        fig.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', barmode='group', bargap=bar_gap, height=420, margin=dict(t=50, b=40, l=40, r=30), legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5, font=dict(size=11)))
        fig.update_yaxes(range=y_range, showline=False, zeroline=False, showgrid=False, gridcolor='rgba(0,0,0,0)', gridwidth=0, row=1, col="all")
        for ann in fig.layout.annotations: ann.update(y=1.03, font_size=co_font_size)
        return fig

    # --- 13. OCI变动分析2 ---    
    def create_asset_liab_oci_chart(df_raw, selected_cos, bar_gap, co_font_size, show_labels, highlight_co="无"):
        d, yrs, fns = df_raw[df_raw['字段名'].isin(['可转损益的负债OCI', 'FVOCI债券公允价值'])].copy(), [str(prev_year), str(latest_year)], ['可转损益的负债OCI', 'FVOCI债券公允价值']
        d['v'] = d['(百万)人民币'] / divisor
        av_cos = [c for c in selected_cos if c in d['公司'].unique()]
        if not av_cos: return go.Figure()
        
        gv = lambda c, y, f: d[(d['公司']==c) & (d['报告年份'].astype(str)==y) & (d['字段名']==f)]['v'].sum()
        all_v = [gv(c, y, f) for c in av_cos for y in yrs for f in fns]
        g_max, g_min = max(all_v+[0.1]), min(all_v+[-0.1])
        ph_h, n, colors = max(abs(g_max), abs(g_min))*0.4, len(av_cos), {fns[0]: "rgb(0, 184, 245)", fns[1]: "rgb(253, 52, 156)"}
        
        fig = make_subplots(rows=1, cols=n, shared_yaxes=True, horizontal_spacing=0.01 if n<=1 else min(0.01, 0.8/(n-1)), subplot_titles=[f"<b><span style='color:#00338D;'>{c}</span></b>" for c in av_cos])
        for fn, c in colors.items(): fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", marker=dict(symbol="square", size=12, color=c), name=fn, showlegend=True), row=1, col=1)
        
        for i, co in enumerate(av_cos):
            ci, x_lbl, v1, v2, t1, t2, my = i+1, [f"{y}YE" for y in yrs], [], [], [], [], []
            for y in yrs:
                val1, val2 = gv(co, y, fns[0]), gv(co, y, fns[1])
                m = (val1==0 and val2==0) 
                v1.append(0 if m else val1); v2.append(0 if m else val2); my.append(ph_h if m else 0)
                t1.append("" if m else (f"{val1:.0f}" if show_labels and val1!=0 else ""))
                t2.append("" if m else (f"{val2:.0f}" if show_labels and val2!=0 else ""))
                if m: fig.add_annotation(x=f"{y}YE", y=ph_h/2, text="未披露", showarrow=False, font=dict(color="white", size=12), xref=f"x{ci}" if ci>1 else "x", yref="y1", xanchor="center", yanchor="middle")

            fig.add_trace(go.Bar(x=x_lbl, y=v1, marker_color=colors[fns[0]], text=t1, textposition='outside', textfont=dict(size=11, color='#00338D'), width=0.4, offsetgroup=1, showlegend=False, cliponaxis=False, constraintext='none'), row=1, col=ci)
            fig.add_trace(go.Bar(x=x_lbl, y=v2, marker_color=colors[fns[1]], text=t2, textposition='outside', textfont=dict(size=11, color='#00338D'), width=0.4, offsetgroup=2, showlegend=False, cliponaxis=False, constraintext='none'), row=1, col=ci)
            fig.add_trace(go.Bar(x=x_lbl, y=my, marker_color="#CDCDCD", width=0.8, offset=-0.4, showlegend=False, cliponaxis=False, hoverinfo='skip'), row=1, col=ci) 
            
            # 🌟 核心修复：纯粹的 "x domain"，加深边框线颜色为 #B0B0B0，y0拉长到底部平齐
            hl = (str(co).strip() == str(highlight_co).strip())
            xref_str = f"x{ci} domain" if ci > 1 else "x domain"
            yref_str = "y domain"
            fig.add_shape(type="rect",
                xref=xref_str, yref=yref_str,
                x0=-0.06, x1=1.06, y0=0, y1=1.08,
                fillcolor=HL_BOX_FILL if hl else "rgba(245,245,245,0.5)",
                line=dict(color=HL_BOX_LINE   if hl else "#CCCCCC", width=1.5 if hl else 1),
                layer="below")       
        fig.update_layout(barmode='group', bargap=bar_gap, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', height=420, margin=dict(t=50, b=80, l=20, r=20), legend=dict(orientation="h", yanchor="top", y=-0.28, x=0.5, xanchor="center"))
        y_rng = [g_min - abs(g_min)*0.3, g_max + abs(g_max)*0.3]
        
        for i in range(1, n+1):
            fig.update_xaxes(type='category', categoryorder='array', categoryarray=x_lbl, showline=False, zeroline=False, showgrid=False, ticks="", ticklen=0, row=1, col=i)
            fig.update_yaxes(range=y_rng, showline=False, zeroline=True, zerolinecolor="#E0E0E0", zerolinewidth=1.02, showgrid=False, gridcolor='rgba(0,0,0,0)', gridwidth=0, row=1, col=i)
        for ann in fig.layout.annotations:
            if "span" in str(ann.text): ann.update(y=1.08, font_size=co_font_size)
            
        return fig
    
    # --- 14. 计算 OCI 分析表 ---
    def calculate_oci_analysis_table(df_raw, selected_cos, highlight_co="无"):
        f = ['AC', 'FVOCI', 'FVTPL', '指定FVOCI', 'FVOCI债券公允价值', '可转损益的负债OCI']
        df = df_raw.copy()
        df['报告年份'], df['公司'], df['字段名'] = df['报告年份'].astype(str).str.replace('.0','',regex=False).str.strip(), df['公司'].astype(str).str.strip(), df['字段名'].astype(str).str.strip()
        val_col = next((c for c in ["(百万)人民币", "(百万)原币", "数值", "值"] if c in df.columns), df.columns[-1])
        cy, py = str(latest_year), str(prev_year)
        
        # 🌟 先收集所有公司的数据，准备按列生成
        metrics = {k: [] for k in ["py_r", "cy_r", "f_c", "l_c", "tot"]}
        for co in selected_cos:
            d_co, data = df[(df['公司'].str.contains(co, na=False)) | (df['公司']==co)], {}
            for y in [cy, py]:
                for fn in f:
                    vs = d_co[(d_co['报告年份']==y) & (d_co['字段名']==fn)][val_col]
                    if vs.empty: vs = d_co[(d_co['报告年份']==y) & (d_co['字段名'].str.contains(fn, na=False))][val_col]
                    data[f"{fn}_{y}"] = vs.iloc[0] if not vs.empty and pd.notna(vs.iloc[0]) else 0

            gr = lambda y: data[f"FVOCI_{y}"] / sum(data[f"{k}_{y}"] for k in ['AC', 'FVOCI', 'FVTPL', '指定FVOCI']) if sum(data[f"{k}_{y}"] for k in ['AC', 'FVOCI', 'FVTPL', '指定FVOCI']) != 0 else "未披露"
            v_fc, v_fp, v_lc, v_lp = data[f"FVOCI债券公允价值_{cy}"], data[f"FVOCI债券公允价值_{py}"], data[f"可转损益的负债OCI_{cy}"], data[f"可转损益的负债OCI_{py}"]
            
            r_f = "未披露" if v_fc==0 and v_fp==0 else (v_fc/v_fp if v_fp!=0 else 0)
            r_l = "未披露" if v_lc==0 and v_lp==0 else (v_lc/v_lp if v_lp!=0 else 0)
            r_tot = "未披露" if r_f=="未披露" or r_l=="未披露" else (r_f/r_l if r_l!=0 else 0)

            metrics["py_r"].append(gr(py)); metrics["cy_r"].append(gr(cy))
            metrics["f_c"].append(r_f); metrics["l_c"].append(r_l); metrics["tot"].append(r_tot)

        # 🌟 开始生成 HTML：列名为公司，行名为指标。极简紧凑样式
        hl_co = str(highlight_co).strip()
        html = "<table style='width:100%; border-collapse:collapse; font-family:sans-serif; font-size:11px; margin-top:-20px; margin-bottom:15px; text-align:center;'>"
        # 表头：首列是"项目"，后面全是"公司名"
        html += "<tr style='background-color:#00338D; color:white; font-weight:bold;'><th style='padding:4px 3px; border:1px solid white; text-align:left; white-space:nowrap;'>项目</th>"
        for co in selected_cos:
            is_hl = (str(co).strip() == hl_co)
            # 高亮公司的表头稍微深一点
            html += f"<th style='padding:4px 3px; border:1px solid white; background-color:{'#002266' if is_hl else '#00338D'};'>{co}</th>"
        html += "</tr>"

        # 逐行写入指标数据
        rows = [(f"FVOCI占比({py})", "py_r"), (f"FVOCI占比({cy})", "cy_r"), ("FVOCI变动", "f_c"), ("负债OCI变动", "l_c"), ("资负匹配率", "tot")]
        for idx, (lbl, key) in enumerate(rows):
            bg = "#F8F9FA" if idx % 2 == 0 else "white"
            html += f"<tr style='background-color:{bg}; color:#333;'><td style='padding:4px 3px; border:1px solid #EAEAEA; text-align:left; font-weight:bold; white-space:nowrap;'>{lbl}</td>"
            
            for i, co in enumerate(selected_cos):
                val, is_hl = metrics[key][i], (str(co).strip() == hl_co)
                # 复杂的样式控制：未披露优先 > 高亮优先 > 普通
                td_bg = "#CDCDCD" if val == "未披露" else ("rgba(0,51,141,0.08)" if is_hl else bg)
                tc = "white" if val == "未披露" else ("#00338D" if is_hl else "#333")
                fw = "bold" if is_hl and val != "未披露" else "normal"
                brd = "1.5px solid #00338D" if is_hl else "1px solid #EAEAEA"
                v_str = "未披露" if val == "未披露" else f"{val:.1%}"
                
                html += f"<td style='padding:3px 2px; border:{brd}; background-color:{td_bg}; color:{tc}; font-weight:{fw};'>{v_str}</td>"
            html += "</tr>"

        return html + "</table>", cy, py


    # --- 15.净资产与总资产 ---
    def create_asset_trend_chart(df_source, field_name, title, show_labels, p_size, g_gap, highlight_co="无"):
        d = df_source[df_source['指标名称'] == field_name].copy()
        d['报告年份'], d['value'] = d['报告年份'].astype(str).str.replace(".0", "", regex=False), d['value'] / divisor 
        fig, x_idx, hl_co = go.Figure(), list(range(len(selected_cos))), str(highlight_co).strip()
        
        for yr, col in zip([str(prev_year), str(latest_year)], ["rgb(15, 101, 253)", "rgb(0, 51, 141)"]):
            df_yr = d[d['报告年份'] == yr].set_index('公司').reindex(selected_cos).reset_index()
            fig.add_trace(go.Bar(name=f"{yr}YE", x=x_idx, y=df_yr['value'], marker_color=col, text=[f"{v:.1f}" if pd.notna(v) and v!=0 else "" for v in df_yr['value']] if show_labels else None, textposition='outside', textfont=dict(size=12), cliponaxis=False))
            
        all_vals = d['value'].dropna()
        off = (all_vals.max() - all_vals.min()) * 0.12 if not all_vals.empty else 12
        df_old, df_new = d[d['报告年份'] == str(prev_year)].set_index('公司'), d[d['报告年份'] == str(latest_year)].set_index('公司')
        
        for i, co in enumerate(selected_cos):
            if co in df_old.index and co in df_new.index:
                v0, v1 = df_old.loc[co, 'value'], df_new.loc[co, 'value']
                if pd.notna(v0) and v0 != 0:
                    pct = (v1 - v0) / abs(v0)
                    fig.add_annotation(x=i, y=(max(v0, v1)+off if v1>=0 else min(v0, v1)-off), text=f"<b>{'↗' if pct>=0 else '↘'} {pct:.1%}</b>", showarrow=False, font=dict(color="#FD349C" if pct>=0 else "#269924", size=p_size), xshift=10, valign="bottom" if v1>=0 else "top")
                    
        if hl_co in [str(c).strip() for c in selected_cos]:
            idx = [str(c).strip() for c in selected_cos].index(hl_co)
            fig.add_shape(type="rect", xref="x", yref="paper", x0=idx-0.46, x1=idx+0.46, y0=-0.12, y1=1.05, fillcolor=HL_BOX_FILL  , line=dict(color=HL_BOX_LINE, width=1.5), layer="above")

        fig.update_layout(barmode='group', bargroupgap=0, bargap=g_gap, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', margin=dict(t=40, b=60, l=40, r=40), height=500, legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        fig.update_xaxes(showline=False, showgrid=False, zeroline=False, tickvals=x_idx, ticktext=[f"<span style='color:#00338D;'><b>{co}</b></span>" for co in selected_cos], ticks="", ticklen=0)
        fig.update_yaxes(showgrid=False, zeroline=True, zerolinecolor="#E0E0E0", zerolinewidth=1.02)
        return fig
    
    # --- 16. CSM 趋势 ---
    def create_csm_trend_chart(df_source, field_name, title, show_labels, p_size, g_gap, highlight_co="无"):   
        d = df_source[df_source['指标名称'] == field_name].copy()
        d['value'] = d['value'] / divisor 
        yrs, hl_co = sorted(d['报告年份'].dropna().astype(str).str.replace(".0", "", regex=False).unique()), str(highlight_co).strip()
        
        c_map = {str(latest_year): "#00338D", str(prev_year): "#0F65FD"}
        if len(yrs) > 2: c_map[yrs[-3]] = "#ADD8E6"
        
        fig, x_idx = go.Figure(), list(range(len(selected_cos)))
        all_vals = d['value'].dropna()
        all_vals = all_vals[all_vals != 0]
        y_max, y_min = (all_vals.max(), all_vals.min()) if not all_vals.empty else (100, 0)
        off = (y_max - y_min) * 0.12 if not all_vals.empty else 12
        placeholder_h = y_max * 0.15 if y_max > 0 else 10  

        # 隐形散点锁定图例
        for yr in yrs:
            fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", marker=dict(symbol="square", size=10, color=c_map.get(yr, "gray")), name=f"{yr}YE"))

        # 🌟 核心修复：预先按年份整理好所有数据桶
        y_data, t_data, c_data, p_data, fc_data = {y:[] for y in yrs}, {y:[] for y in yrs}, {y:[] for y in yrs}, {y:[] for y in yrs}, {y:[] for y in yrs}
        
        for i, co in enumerate(selected_cos):
            co_vals = [d[(d['公司'] == co) & (d['报告年份'].astype(str) == yr)]['value'].iloc[0] if not d[(d['公司'] == co) & (d['报告年份'].astype(str) == yr)].empty else np.nan for yr in yrs]
            
            # 如果这家公司所有年份全是空的或 0
            if all(pd.isna(v) or v == 0 for v in co_vals):
                for yr in yrs:
                    y_data[yr].append(placeholder_h); t_data[yr].append(""); c_data[yr].append("#CDCDCD"); p_data[yr].append("inside"); fc_data[yr].append("white")
                # 🌟 绝对居中打上标签
                fig.add_annotation(x=i, y=placeholder_h/2, text="未披露", showarrow=False, font=dict(color="white", size=11), xanchor="center", yanchor="middle")
            else:
                # 正常情况或只有部分年份缺失，文字就在各自的柱子内部或外部
                for idx, yr in enumerate(yrs):
                    v = co_vals[idx]
                    if pd.isna(v) or v == 0:
                        y_data[yr].append(placeholder_h); t_data[yr].append("未披露"); c_data[yr].append("#CDCDCD"); p_data[yr].append("inside"); fc_data[yr].append("white")
                    else:
                        y_data[yr].append(v); t_data[yr].append(f"{v:.0f}" if show_labels else ""); c_data[yr].append(c_map.get(yr, "gray")); p_data[yr].append("outside"); fc_data[yr].append("#333")

        # 将整理好的数据画成真实柱子
        for yr in yrs:
            fig.add_trace(go.Bar(name=f"{yr}YE", x=x_idx, y=y_data[yr], marker_color=c_data[yr], text=t_data[yr], textposition=p_data[yr], textfont=dict(size=12, color=fc_data[yr]), cliponaxis=False, showlegend=False))

        df_old, df_new = d[d['报告年份'].astype(str) == str(prev_year)].set_index('公司'), d[d['报告年份'].astype(str) == str(latest_year)].set_index('公司')
        for i, co in enumerate(selected_cos):
            if co in df_old.index and co in df_new.index:
                v0, v1 = df_old.loc[co, 'value'], df_new.loc[co, 'value']
                if pd.notna(v0) and v0 != 0 and pd.notna(v1) and v1 != 0:
                    pct = (v1 - v0) / abs(v0)
                    fig.add_annotation(x=i, y=(max(v0, v1)+off if v1>=0 else min(v0, v1)-off), text=f"<b>{'↗' if pct>=0 else '↘'} {pct:.1%}</b>", showarrow=False, font=dict(color="#FD349C" if pct>=0 else "#269924", size=p_size), xshift=15, valign="bottom" if v1>=0 else "top")
                    
        if hl_co in [str(c).strip() for c in selected_cos]:
            idx = [str(c).strip() for c in selected_cos].index(hl_co)
            fig.add_shape(type="rect", xref="x", yref="paper", x0=idx-0.46, x1=idx+0.46, y0=-0.12, y1=1.05, fillcolor=HL_BOX_FILL  , line=dict(color=HL_BOX_LINE, width=1.5), layer="above")

        r_top = max(y_max + (y_max - y_min) * 0.15, placeholder_h * 1.3)
        fig.update_layout(barmode='group', bargroupgap=0, bargap=g_gap, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', margin=dict(t=40, b=60, l=40, r=40), height=450, legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        fig.update_xaxes(showline=False, showgrid=False, zeroline=False, tickvals=x_idx, ticktext=[f"<span style='color:#00338D;'><b>{co}</b></span>" for co in selected_cos], ticks="", ticklen=0)
        fig.update_yaxes(showgrid=False, zeroline=True, zerolinecolor="#E0E0E0", zerolinewidth=1.02, range=[0 if y_min>=0 else y_min*1.2, r_top])
        return fig


# --- CSM占比 ---
    def create_csm_ratio_chart(df, cos, show_lbl, gap, highlight_co="无"):
        c_bar = {prev_year: '#C7A0F7', latest_year: '#1E49E2'}
        c_line = {prev_year: '#7213EA', latest_year: '#00338D'}
        hl_co, fig, x_idx = str(highlight_co).strip(), go.Figure(), list(range(len(cos)))
    
        def get_r(y):
            data = []
            for co in cos:
                c_df = df[(df['公司'] == co) & (df['报告年份'].astype(str) == str(y))]
                v = lambda k: c_df[c_df['字段名'].str.contains(k, regex=False, na=False)]['(百万)人民币'].sum()
                a, b, c = v('CSM期末'), v('LRC非亏损部分期末'), v('LRC亏损部分期末')
                denom = b + c - a
                data.append(a / denom if denom != 0 else np.nan) 
            return data
    
        y_py, y_cy = get_r(prev_year), get_r(latest_year)
        valid_py, valid_cy = [v for v in y_py if pd.notna(v)], [v for v in y_cy if pd.notna(v)]
        avg_py, avg_cy = np.mean(valid_py) if valid_py else 0, np.mean(valid_cy) if valid_cy else 0
        
        # 🌟 修复：同时获取最大值和最小值，用于设定动态的 Y 轴范围
        all_valid_vals = valid_py + valid_cy
        y_max_val = max(all_valid_vals) if all_valid_vals else 0
        y_min_val = min(all_valid_vals) if all_valid_vals else 0
        
        y_range_top = y_max_val * 1.15 + 0.05
        # 🌟 如果有负数，底部向下延伸 15%；否则就在 0
        y_range_bot = y_min_val * 1.15 - 0.05 if y_min_val < 0 else 0 
        
        # 占位柱高度：取绝对值最大值的 15%
        placeholder_h = max(abs(y_max_val), abs(y_min_val)) * 0.15 if all_valid_vals else 0.1 

        for yr in [prev_year, latest_year]:
            fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", marker=dict(symbol="square", size=10, color=c_bar[yr]), name=f"{yr}YE"))

        y_vals_py, t_vals_py, c_vals_py, p_vals_py, fc_vals_py = [], [], [], [], []
        y_vals_cy, t_vals_cy, c_vals_cy, p_vals_cy, fc_vals_cy = [], [], [], [], []
        
        for i, co in enumerate(cos):
            v_p, v_c = y_py[i], y_cy[i]
            m_p, m_c = (pd.isna(v_p) or v_p == 0), (pd.isna(v_c) or v_c == 0)
            
            if m_p and m_c:
                y_vals_py.append(placeholder_h); y_vals_cy.append(placeholder_h)
                t_vals_py.append(""); t_vals_cy.append("")
                c_vals_py.append("#CDCDCD"); c_vals_cy.append("#CDCDCD")
                p_vals_py.append("inside"); p_vals_cy.append("inside")
                fc_vals_py.append("white"); fc_vals_cy.append("white")
                fig.add_annotation(x=i, y=placeholder_h/2, text="未披露", showarrow=False, font=dict(color="white", size=11), xanchor="center", yanchor="middle")
            else:
                if m_p: y_vals_py.append(placeholder_h); t_vals_py.append("未披露"); c_vals_py.append("#CDCDCD"); p_vals_py.append("inside"); fc_vals_py.append("white")
                else: y_vals_py.append(v_p); t_vals_py.append(f"{v_p:.1%}" if show_lbl else ""); c_vals_py.append(c_bar[prev_year]); p_vals_py.append("outside"); fc_vals_py.append("#333")
                
                if m_c: y_vals_cy.append(placeholder_h); t_vals_cy.append("未披露"); c_vals_cy.append("#CDCDCD"); p_vals_cy.append("inside"); fc_vals_cy.append("white")
                else: y_vals_cy.append(v_c); t_vals_cy.append(f"{v_c:.1%}" if show_lbl else ""); c_vals_cy.append(c_bar[latest_year]); p_vals_cy.append("outside"); fc_vals_cy.append("#333")

        fig.add_trace(go.Bar(name=f"{prev_year}YE", x=x_idx, y=y_vals_py, marker_color=c_vals_py, text=t_vals_py, textposition=p_vals_py, textfont=dict(size=12, color=fc_vals_py), cliponaxis=False, showlegend=False))
        fig.add_trace(go.Bar(name=f"{latest_year}YE", x=x_idx, y=y_vals_cy, marker_color=c_vals_cy, text=t_vals_cy, textposition=p_vals_cy, textfont=dict(size=12, color=fc_vals_cy), cliponaxis=False, showlegend=False))
    
        too_close = abs(avg_py - avg_cy) < y_range_top * 0.05
        for i, (avg, yr) in enumerate([(avg_py, prev_year), (avg_cy, latest_year)]):
                fig.add_shape(type="line", x0=0, x1=1, y0=avg, y1=avg, xref="paper", yref="y", line=dict(color=c_line[yr], dash="dash"))
                y_shift = y_max_val * 0.20 if (too_close and i == 0) else (-y_range_top * 0.03 if too_close else 0)
                fig.add_annotation(x=0.99, y=avg + y_shift, xref="paper", yref="y", xanchor="right", text=f"{yr}均值: {avg:.1%}", showarrow=False, bgcolor="rgba(255,255,255,0.15)", bordercolor=c_line[yr], borderwidth=1, borderpad=2, font=dict(color=c_line[yr], size=11))
    
        if hl_co in [str(c).strip() for c in cos]:
            idx = [str(c).strip() for c in cos].index(hl_co)
            fig.add_shape(type="rect", xref="x", yref="paper", x0=idx-0.46, x1=idx+0.46, y0=-0.10, y1=1.05, fillcolor=HL_BOX_FILL, line=dict(color=HL_BOX_LINE  , width=1.5), layer="above")
    
        fig.update_layout(barmode='group', bargroupgap=0, bargap=gap, height=450, margin=dict(t=50, b=40, l=40, r=40), plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', legend=dict(orientation="h", yanchor="bottom", y=1.05, xanchor="right", x=1))
        fig.update_xaxes(showline=False, showgrid=False, zeroline=False, tickvals=x_idx, ticktext=[f"<span style='color:#00338D;'><b>{c}</b></span>" for c in cos], ticks="", ticklen=0)
        
        # 🌟 修复：应用包含负数的动态 y_range_bot
        fig.update_yaxes(showgrid=False, tickformat='.0%', zeroline=True, zerolinecolor='lightgray', range=[y_range_bot, max(y_range_top, placeholder_h * 1.3)])
        return fig

# --- 18. CSM概览表 (升级为纯HTML渲染) ---
    def show_csm_summary_table(df, target_year, selected_cos, highlight_co="无"):
        f_map = {"CSM期初余额": "CSM余额", "新业务CSM（集团口径）": "新业务CSM", "CSM计息": "CSM计息", "CSM调整": "CSM调整", "CSM摊销": "CSM摊销", "CSM其他": "其他变动", "CSM期末余额": "CSM期末余额"}
        f_keys, f_names = list(f_map.keys()), list(f_map.values())
        y_str = str(target_year).replace(".0", "")
        
        d_sub = df[(df['报告年份'].astype(str).str.contains(y_str)) & (df['字段名'].isin(f_keys))]
        if d_sub.empty: return ""
        
        hl = str(highlight_co).strip()
        # 🌟 字体缩小至 11px，间距压缩
        html = "<table style='width:100%; border-collapse:collapse; font-family:sans-serif; font-size:11px; margin-bottom:10px; text-align:center;'><tr style='background-color:#00338D; color:white; font-weight:bold;'><th style='padding:4px; border:1px solid white; text-align:left;'>公司</th>"
        for fn in f_names: html += f"<th style='padding:4px; border:1px solid white;'>{fn}</th>"
        html += "</tr>"
        
        for i, co in enumerate(selected_cos):
            is_hl = (str(co).strip() == hl)
            bg = "rgba(0,51,141,0.08)" if is_hl else ("white" if i % 2 == 0 else "#F8F9FA")
            html += f"<tr style='background-color:{bg}; color:{'#00338D' if is_hl else '#333'}; font-weight:{'bold' if is_hl else 'normal'};'><td style='padding:4px; border:1px solid #EAEAEA; text-align:left; font-weight:bold;'>{co}</td>"
            
            for fk in f_keys:
                vs = d_sub[(d_sub['公司'] == co) & (d_sub['字段名'] == fk)]['(百万)人民币']
                v = vs.iloc[0] if not vs.empty else 0
                
                # 🌟 核心拦截：缺失或为 0 则变灰写“未披露”
                if pd.isna(v) or v == 0:
                    html += "<td style='padding:4px; border:1px solid #EAEAEA; background-color:#CDCDCD; color:white;'>未披露</td>"
                else:
                    v_div = v / divisor
                    v_str = f"({abs(v_div):.1f})" if v_div < 0 else f"{v_div:.1f}"
                    html += f"<td style='padding:4px; border:1px solid #EAEAEA;'>{v_str}</td>"
            html += "</tr>"
            
        return html + "</table>"

    # --- 19. CSM过渡期拆 ---
    def create_csm_transition_chart(df, selected_cos, show_labels, bar_width, highlight_co="无"):
        cols, m = 3, [("采用公允价值法计量的合同", "rgb(0, 51, 141)", "采用公允价值法计量的合同"), ("采用修正追溯法计量的合同", "rgb(1, 176, 234)", "采用修正追溯调整法计量的合同"), ("其他保险合同", "#7213Ea", "其他合同")]
        rows = (len(selected_cos) + cols - 1) // cols
        titles = [f"<b><span style='color:#00338D;'>{co}</span></b>" for co in selected_cos]
        fig = make_subplots(rows=rows, cols=cols, subplot_titles=titles, horizontal_spacing=0.08, vertical_spacing=0.15)
        ym, hl_co = latest_year, str(highlight_co).strip()
        
        # 准备统一的 Y 轴标签（强行锁定，防止变成 0, 2, 4）
        y_labels = [f"{ym-2}YE", f"{ym-1}YE", f"{ym}YE"]
        
        # 用隐形散点强制锁定全局图例
        for name, clr, _ in m:
            fig.add_trace(go.Scatter(
                x=[None], y=[None], mode="markers", 
                marker=dict(symbol="square", size=10, color=clr), 
                name=name, showlegend=True
            ))

        for i, co in enumerate(selected_cos):
            r, c, df_c = (i // cols) + 1, (i % cols) + 1, df[df['公司'] == co]
            for yt in [ym-2, ym-1, ym]:
                dy, suf = (ym, "期末负债") if yt == ym else ((ym, "期初负债") if yt == ym-1 else (ym-1, "期初负债"))
                df_y = df_c[df_c['报告年份'].astype(str) == str(dy)]
                vs = [df_y.loc[df_y['字段名'] == f"{f}{suf}", '(百万)人民币'].sum() for _, _, f in m]
                tot = sum(vs)
                
                if tot <= 0:
                    fig.add_trace(go.Bar(x=[100], textangle=0, constraintext='none',y=[f"{yt}YE"], orientation='h', marker_color='#CDCDCD', width=bar_width, showlegend=False, text=["未披露"], textposition='inside', insidetextanchor='middle', textfont=dict(size=11, color="white")), row=r, col=c)
                else:
                    for j, (name, clr, _) in enumerate(m):
                        pct = vs[j] / tot * 100
                        fig.add_trace(go.Bar(x=[pct],textangle=0, constraintext='none',y=[f"{yt}YE"], orientation='h', name=name, marker_color=clr, width=bar_width, showlegend=False, text=[f"{pct:.0f}%" if show_labels and pct>0 else ""], textposition='inside', insidetextanchor='middle', textfont=dict(size=11, color="white")), row=r, col=c)
            
            is_hl = (str(co).strip() == hl_co)
            bg_fill = HL_BOX_FILL   if is_hl else "rgba(0,0,0,0)"
            line_dict = dict(color=HL_BOX_LINE, width=1.5) if is_hl else dict(color="rgba(0,0,0,0)", width=0)
            fig.add_shape(type="rect", xref="x domain", yref="y domain", x0=-0.08, x1=1.08, y0=-0.10, y1=1.33, fillcolor=bg_fill, line=line_dict, layer="above", row=r, col=c)
            fig.update_yaxes(
                type='category',             # 强制设为分类文本轴
                categoryorder='array',       # 按我们指定的数组排序
                categoryarray=y_labels,      # 锁定标签为 
                showgrid=False, ticks="", ticklen=0, tickcolor="rgba(0,0,0,0)",
                row=r, col=c                 # 精准打击当前子图
            )
            fig.update_xaxes(
                range=[0, 100], 
                tickvals=[0, 20, 40, 60, 80, 100], 
                ticksuffix="%",               # 也可以加 "%"
                showgrid=False,
                row=r, col=c                 # 精准打击当前子图
            )
            
        fig.update_layout(barmode='stack', height=300 * rows, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', margin=dict(t=60, b=80, l=50, r=20), legend=dict(orientation="h", x=0.5, xanchor="center", y=-0.08))
        fig.update_annotations(yshift=5, font_size=12)
        return fig

    # --- 20.摊销前 CSM ---
    def create_csm_composition_chart(df, selected_cos, target_year, show_labels, label_size, bar_width, highlight_co="无", title_text=""):
        field_map = {"新业务CSM（集团口径）": "新业务 CSM", "CSM计息": "CSM 计息", "CSM调整": "CSM 调整"}
        color_map = {"新业务CSM（集团口径）": "rgb(0, 51, 140)", "CSM计息": "rgb(147, 157, 253)", "CSM调整": "rgb(253, 52, 156)"}
        fields, year_str = list(field_map.keys()), str(target_year).replace(".0", "")
        d_sub = df[(df['报告年份'].astype(str).str.contains(year_str)) & (df['公司'].isin(selected_cos)) & (df['字段名'].isin(fields))].copy()
        if d_sub.empty: return None
        d_pivot = d_sub.pivot(index='公司', columns='字段名', values='(百万)人民币')
        for f in fields:
            if f not in d_pivot.columns: d_pivot[f] = 0
        d_pivot = d_pivot.reindex(selected_cos).fillna(0)
        d_pivot['Total'] = d_pivot[fields].abs().sum(axis=1).replace(0, 1)
        no_data_cos = d_pivot[d_pivot[fields].abs().sum(axis=1) == 0].index.tolist()
        fig, hl_co = go.Figure(), str(highlight_co).strip()
        x_idx = list(range(len(selected_cos)))
        no_data_x = [i for i, co in enumerate(selected_cos) if co in no_data_cos]
        if no_data_x:
            fig.add_trace(go.Bar(
                x=no_data_x, y=[100] * len(no_data_x),
                width=bar_width,
                marker_color="#D9D9D9",
                showlegend=False,
                text=["未披露"] * len(no_data_x),
                textposition='inside',
                insidetextanchor='middle',
                textfont=dict(size=label_size, color="white"),
                hoverinfo='skip'
            ))
        
        for f_name in fields:
            vals_pct = (d_pivot[f_name] / d_pivot['Total']) * 100
            # ✅ 有数据的公司正常显示，无数据的公司值设为0不显示
            vals_pct = [0 if co in no_data_cos else v for co, v in zip(selected_cos, vals_pct)]
            t_color = "white" if f_name == "新业务CSM（集团口径）" else "black"
            fig.add_trace(go.Bar(
                name=field_map[f_name], x=x_idx, y=vals_pct,
                width=bar_width, marker_color=color_map[f_name],
                text=[f"{v:.0f}%" if abs(v) >= 1 else "" for v in vals_pct] if show_labels else None,
                textposition='inside', insidetextanchor='middle',
                textfont=dict(size=label_size, color=t_color),
                constraintext='none'
            ))  
        if hl_co in [str(c).strip() for c in selected_cos]:
            idx = [str(c).strip() for c in selected_cos].index(hl_co)
            fig.add_shape(type="rect", xref="x", yref="paper", x0=idx-0.46, x1=idx+0.46, y0=-0.12, y1=1.05, fillcolor=HL_BOX_FILL  , line=dict(color=HL_BOX_LINE, width=1.5), layer="above")

        tick_txt = [f"<span style='font-size:14px; color:#00338D;'><b>{co}</b></span>" for co in selected_cos]
        
        # 🌟 修复点：动态判断是否传入了 title_text
        layout_args = dict(barmode='relative', height=500, margin=dict(t=80 if title_text else 40, b=100, l=60, r=40), paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5))
        if title_text:
            layout_args['title'] = dict(text=f"<b>{title_text}</b>", x=0.5, y=0.95, xanchor='center', font=COMMON_TITLE_FONT)
            
        fig.update_layout(**layout_args)
        fig.update_xaxes(showgrid=False, zeroline=False, ticktext=tick_txt, tickvals=x_idx, ticks="", ticklen=0)
        fig.update_yaxes(showgrid=False, range=[-25, 105], tickvals=[-20, 0, 20, 40, 60, 80, 100], ticktext=["-20%", "0%", "20%", "40%", "60%", "80%", "100%"], zeroline=True, zerolinecolor='#FFB07C', zerolinewidth=1.5)
        return fig

    # --- 21. CSM 比率折线图 ---  
    def get_company_color_mapping(companies, key_prefix="cp"):
        PRESET_COLORS_FIXED = ["#C00000", "#0865EE", "#FEAED7", "#92D050", "#7030A0", "#EF9867", "#61CBF4", "#C7A0F7"]
        cols = st.columns(len(companies))
        mapping = {}
        for i, co in enumerate(companies):
            with cols[i]:
                st.caption(f"**{co}**")
                mapping[co] = st.color_picker(label=co, value=PRESET_COLORS_FIXED[i % len(PRESET_COLORS_FIXED)], key=f"{key_prefix}_{co}", label_visibility="collapsed")
        return mapping

    #22.CSM摊销、持续比例图
    def render_shared_legend(selected_cos, color_map, highlight_co="无"):
            items = "".join([f'<div style="display:flex;align-items:center;"><div style="width:10px;height:10px;border-radius:50%;background:{HIGHLIGHT_COLOR if co == highlight_co else color_map.get(co, "#333")};margin-right:5px;flex-shrink:0;"></div><span>{co}</span></div>' for co in selected_cos])
            st.markdown(f'<div style="display:flex;flex-wrap:wrap;align-items:center;gap:14px;margin-bottom:6px;padding-left:6px;font-size:12px;">{items}</div>', unsafe_allow_html=True)
    
    def create_ratio_line_chart_v3(df_plot, title, color_map, show_labels, marker_size, y_axis_format, selected_cos, highlight_co="无", print_mode=False):
            fig = go.Figure()
            latest_yr = df_plot['报告年份'].max()
            df_latest = df_plot[df_plot['报告年份'] == latest_yr].dropna(subset=['value'])
            
            max_co = df_latest.loc[df_latest['value'].idxmax(), '公司'] if not df_latest.empty else None
            min_co = df_latest.loc[df_latest['value'].idxmin(), '公司'] if not df_latest.empty else None
    
            for co in selected_cos:
                d_co = df_plot[df_plot['公司'] == co].sort_values('报告年份')
                if d_co.empty: continue
                
                is_hl, is_ext = (co == highlight_co), (co in [max_co, min_co])
                fig.add_trace(go.Scatter(
                    x=d_co['报告年份'], y=d_co['value'], name=co,
                    mode='lines+markers+text' if show_labels else 'lines+markers',
                    line=dict(color=HIGHLIGHT_COLOR if is_hl else color_map.get(co, "#333"), width=4 if is_hl else (2.5 if is_ext else 2), dash="solid" if (is_hl or is_ext) else "dot"),
                    marker=dict(size=marker_size * 1.5 if is_hl else marker_size, symbol='circle'),
                    text=[f"{v*100:.1f}%" if (not print_mode or i == len(d_co)-1) else "" for i, v in enumerate(d_co['value'])],
                    textposition="middle right", cliponaxis=False
                ))
    
            fig.update_layout(title=dict(text=title, x=0, font=dict(size=12)), paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', height=160, margin=dict(t=25, b=0, l=0, r=0), showlegend=False)
            fig.update_xaxes(showgrid=False, showline=False, zeroline=False)
            fig.update_yaxes(showgrid=False, zeroline=False, tickformat=y_axis_format)
            
            return fig

    # --- 22.综合净资产指标 --- 
    def create_csm_equity_analysis(df, selected_cos, show_labels, label_size, bar_width, co_font_size, pct_height_adjust, highlight_co="无",title_text=""):
        d_sub = df[(df['报告年份'].astype(str).isin([str(prev_year), str(latest_year)])) & (df['公司'].isin(selected_cos)) & (df['字段名'].isin(["CSM期末余额", "期末股东权益"]))].copy()
        d_sub['value'] = d_sub['(百万)人民币'] / divisor            
        
        # 获取全局最大值，用于设置 Y 轴高度和占位柱高度
        _all_totals = [d_sub[(d_sub['公司'] == _co) & (d_sub['报告年份'].astype(str) == str(_yr))]['value'].sum() for _co in selected_cos for _yr in [prev_year, latest_year]]
        global_max_bar = max(_all_totals + [1.0])
        global_base_y, global_y_top = global_max_bar * 1.03, global_max_bar * 1.12
        placeholder_h = global_max_bar * 0.4  # 🌟 占位灰柱高度
        
        hl_co = str(highlight_co).strip()
        n = len(selected_cos)
        titles = [f"<span style='color:#00338D; font-size:{co_font_size}px;'><b>{co}</b></span>" for co in selected_cos]
        fig = make_subplots(rows=1, cols=n, horizontal_spacing=0.015 if n<=1 else min(0.015, 0.8/(n-1)), subplot_titles=titles)            
        
        # 隐形散点锁定全局图例
        for leg_name, leg_color in [("再保前净CSM", "rgb(0, 184, 245)"), ("净资产", "rgb(30, 73, 226)"), ("再保前税后CSM/净资产", "rgb(114, 19, 214)")]:
            fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", marker=dict(symbol="square", size=12, color=leg_color), name=leg_name, showlegend=True), row=1, col=1)                
        
        for i, co in enumerate(selected_cos):
            col_idx, cd = i + 1, d_sub[d_sub['公司'] == co]
            gv = lambda y, f: cd[(cd['报告年份'].astype(str)==str(y)) & (cd['字段名']==f)]['value'].sum()
            v24b, v25b, v24e, v25e = gv(prev_year,"CSM期末余额"), gv(latest_year,"CSM期末余额"), gv(prev_year,"期末股东权益"), gv(latest_year,"期末股东权益")
            
            # 🌟 核心拦截逻辑：判断该年是否完全没数据
            m24, m25 = (v24b==0 and v24e==0), (v25b==0 and v25e==0)
            
            v24c, v25c = (0 if m24 else v24b*0.75), (0 if m25 else v25b*0.75)
            v24e, v25e = (0 if m24 else v24e), (0 if m25 else v25e)
            
            p24 = (v24c / v24e * 100) if v24e != 0 and not m24 else 0
            p25 = (v25c / v25e * 100) if v25e != 0 and not m25 else 0
            x_axis = [f"{prev_year}YE", f"{latest_year}YE"]
            
            # 画真实数据柱子（如果是未披露，传进去的是 0，自动隐
            lbl = lambda v, m: f"{v:.0f}" if show_labels and pd.notna(v) and v!=0 and not m else ""
            fig.add_trace(go.Bar(x=x_axis, y=[v24c, v25c], marker_color="rgb(0, 184, 245)", width=bar_width, showlegend=False, text=[lbl(v24c,m24), lbl(v25c,m25)], textposition='inside', textangle=0, textfont=dict(size=label_size, color="white"), constraintext='none'), row=1, col=col_idx)                
            fig.add_trace(go.Bar(x=x_axis, y=[v24e, v25e], marker_color="rgb(30, 73, 226)", width=bar_width, showlegend=False, text=[lbl(v24e,m24), lbl(v25e,m25)], textposition='inside', textangle=0, textfont=dict(size=label_size, color="white"), constraintext='none'), row=1, col=col_idx)                
            
            # 🌟 画专门的灰色占位柱子
            fig.add_trace(go.Bar(x=x_axis, y=[placeholder_h if m24 else 0, placeholder_h if m25 else 0], marker_color="#CDCDCD", width=bar_width, showlegend=False, text=["未披露" if m24 else "", "未披露" if m25 else ""], textposition='inside', insidetextanchor='middle', textfont=dict(size=12, color="white"), constraintext='none', cliponaxis=False), row=1, col=col_idx)
            
            # 折线图与比例标签逻辑：只有在真实数据存在时才绘制
            pct_max, delta = max(abs(p24), abs(p25), 1.0), global_max_bar * 0.04
            y24_line = global_base_y + (p24 / pct_max) * delta * 0.5 if not m24 else None
            y25_line = global_base_y + (p25 / pct_max) * delta * 0.5 if not m25 else None
            
            fig.add_trace(go.Scatter(x=x_axis, y=[y24_line, y25_line], mode="lines", line=dict(color="rgb(114, 19, 214)", width=1.2), showlegend=False, hoverinfo="skip"), row=1, col=col_idx)
            # 对于散点和文本，如果是未披露，把文本设为空字符串
            fig.add_trace(go.Scatter(x=x_axis, y=[y24_line, y25_line], mode="markers+text", marker=dict(symbol="triangle-up", size=7, color="rgb(114, 19, 214)"), text=[f"{p24:.0f}%" if not m24 else "", f"{p25:.0f}%" if not m25 else ""], textposition="top center", textfont=dict(color="rgb(114, 19, 214)", size=label_size), showlegend=False, hoverinfo="skip"), row=1, col=col_idx)                
            
            fig.update_yaxes(range=[0, global_y_top], showticklabels=False, showgrid=False, zeroline=False, row=1, col=col_idx)                
            
            # 🌟 恢复高亮逻辑：高亮公司带深色框和浅蓝蒙版（layer="above"），普通公司透明/交替浅灰
            is_hl = (str(co).strip() == hl_co)
            bg_fill = HL_BOX_FILL   if is_hl else ("rgba(200, 200, 200, 0.12)" if i % 2 == 1 else "rgba(255,255,255,0)")
            border_dict = dict(color=HL_BOX_LINE, width=1.5) if is_hl else dict(color="#EAEAEA", width=1)
            xref_str = f"x{col_idx} domain" if col_idx > 1 else "x domain"
            fig.add_shape(type="rect", xref=xref_str, yref="paper", x0=-0.04, x1=1.04, y0=-0.12, y1=1.12, line=border_dict, fillcolor=bg_fill, layer="above" if is_hl else "below")

        fig.update_layout(barmode='stack', height=550, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', margin=dict(t=50, b=100, l=20, r=20), legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5, font=dict(size=12), itemsizing="constant"))                
        
        for ann in fig.layout.annotations:
            if "span" in str(ann.text): ann.update(y=1.03)
                
        for i in range(1, n + 1):
            fig.update_xaxes(type="category", showgrid=False, zeroline=False, ticks="", ticklen=0, row=1, col=i)
        return fig

# --- 23.新业务盈利合同 ---   
    def create_new_biz_csm_chart(df_source, field_name, title, show_labels, p_size, g_gap, highlight_co="无"):
        d = df_source[df_source['指标名称'] == field_name].copy()
        d['报告年份'], d['value'] = d['报告年份'].astype(str).str.replace(".0", "", regex=False), d['value'] / divisor 
        y_old, y_new, hl_co = str(prev_year), str(latest_year), str(highlight_co).strip()
        fig, x_idx = go.Figure(), list(range(len(selected_cos)))
        
        all_vals = d['value'].dropna()
        all_vals = all_vals[all_vals != 0]
        y_max, y_min = (all_vals.max(), all_vals.min()) if not all_vals.empty else (100, 0)
        off = (y_max - y_min) * 0.11 if not all_vals.empty else 12
        placeholder_h = y_max * 0.2 if y_max > 0 else 10  # 未披露灰色柱子的高度
        
        df_old, df_new = d[d['报告年份'] == y_old].set_index('公司'), d[d['报告年份'] == y_new].set_index('公司')
        vals_old, vals_new, txt_old, txt_new = [], [], [], []
        col_old, col_new, pos_old, pos_new, fc_old, fc_new = [], [], [], [], [], []

        for i, co in enumerate(selected_cos):
            v0 = df_old.loc[co, 'value'] if co in df_old.index else np.nan
            v1 = df_new.loc[co, 'value'] if co in df_new.index else np.nan
            m0, m1 = pd.isna(v0), pd.isna(v1)

            # 情况1：两年都未披露 -> 双倍宽灰柱，文字居中悬浮
            if m0 and m1:
                vals_old.append(placeholder_h); vals_new.append(placeholder_h)
                txt_old.append(""); txt_new.append("")
                col_old.append("#CDCDCD"); col_new.append("#CDCDCD")
                pos_old.append("inside"); pos_new.append("inside")
                fc_old.append("white"); fc_new.append("white")
                fig.add_annotation(x=i, y=placeholder_h/2, text="未披露", showarrow=False, font=dict(color="white", size=11), xanchor="center", yanchor="middle")
            
            # 情况2：只有往年未披露 -> 往年灰柱(字在内)，最新年蓝柱(字在外)
            elif m0 and not m1:
                vals_old.append(placeholder_h); vals_new.append(v1)
                txt_old.append("未披露"); txt_new.append(f"{v1:.1f}" if show_labels else "")
                col_old.append("#CDCDCD"); col_new.append("rgb(0, 51, 141)")
                pos_old.append("inside"); pos_new.append("outside")
                fc_old.append("white"); fc_new.append("#333")
                
            # 情况3：只有最新年未披露 -> 往年蓝柱(字在外)，最新年灰柱(字在内)
            elif not m0 and m1:
                vals_old.append(v0); vals_new.append(placeholder_h)
                txt_old.append(f"{v0:.1f}" if show_labels else ""); txt_new.append("未披露")
                col_old.append("rgb(15, 101, 253)"); col_new.append("#CDCDCD")
                pos_old.append("outside"); pos_new.append("inside")
                fc_old.append("#333"); fc_new.append("white")
                
            # 情况4：两年都有数据 -> 正常蓝柱，计算百分比箭头
            else:
                vals_old.append(v0); vals_new.append(v1)
                txt_old.append(f"{v0:.1f}" if show_labels else ""); txt_new.append(f"{v1:.1f}" if show_labels else "")
                col_old.append("rgb(15, 101, 253)"); col_new.append("rgb(0, 51, 141)")
                pos_old.append("outside"); pos_new.append("outside")
                fc_old.append("#333"); fc_new.append("#333")
                
                # 🌟 修复：只有当分母 (v0) 不等于 0 时，才计算和绘制百分比箭头
                if v0 != 0:
                    pct = (v1 - v0) / abs(v0)
                    arr_color, arr_symbol = ("#FD349C", "↗") if pct >= 0 else ("#269924", "↘")
                    y_pos, v_align = (max(v0, v1) + off, "bottom") if v1 >= 0 else (min(v0, v1) - off, "top")
                    fig.add_annotation(x=i, y=y_pos, text=f"<b>{arr_symbol} {pct:.1%}</b>", showarrow=False, font=dict(color=arr_color, size=p_size), xshift=8, valign=v_align)
        # 🌟 增加隐形散点图例，锁定为方形和标准蓝
        fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", marker=dict(symbol="square", size=12, color="rgb(15, 101, 253)"), name=f"{y_old}年新业务盈利合同（CSM）"))
        fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", marker=dict(symbol="square", size=12, color="rgb(0, 51, 141)"), name=f"{y_new}年新业务盈利合同（CSM）"))

        # 🌟 真实柱子，关闭自身图例
        fig.add_trace(go.Bar(name=f"{y_old}年新业务盈利合同（CSM）", x=x_idx, y=vals_old, marker_color=col_old, text=txt_old, textposition=pos_old, insidetextanchor="middle",textfont=dict(size=12, color=fc_old), textangle=0, cliponaxis=False, showlegend=False))
        fig.add_trace(go.Bar(name=f"{y_new}年新业务盈利合同（CSM）", x=x_idx, y=vals_new, marker_color=col_new, text=txt_new, textposition=pos_new, insidetextanchor="middle",textfont=dict(size=12, color=fc_new), textangle=0, cliponaxis=False, showlegend=False))
            
        if hl_co in [str(c).strip() for c in selected_cos]:
            idx = [str(c).strip() for c in selected_cos].index(hl_co)
            fig.add_shape(type="rect", xref="x", yref="paper", x0=idx-0.46, x1=idx+0.46, y0=-0.12, y1=1.05, fillcolor=HL_BOX_FILL  , line=dict(color=HL_BOX_LINE, width=1.5), layer="above")

        fig.update_layout(barmode='group', bargroupgap=0, bargap=g_gap, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', margin=dict(t=40, b=60, l=40, r=40), height=500, legend=dict(orientation="v", yanchor="top", y=1.1, xanchor="right", x=1.0))
        r_top, r_bot = y_max + (y_max - y_min) * 0.18, y_min - (y_max - y_min) * 0.15 if y_min < 0 else 0
        r_top = max(r_top, placeholder_h * 1.3) # 保证未披露柱子不会顶破天花板
        fig.update_yaxes(showgrid=False, zeroline=True, zerolinecolor="#E0E0E0", zerolinewidth=1.02, range=[r_bot, r_top])
        fig.update_xaxes(showline=False, showgrid=False, zeroline=False, tickvals=x_idx, ticktext=[f"<span style='color:#00338D;'><b>{co}</b></span>" for co in selected_cos], ticks="", ticklen=0)
        return fig



# --- 24.新业务亏损 ---   
    def create_new_lost_csm_chart(df_source, field_name, title, show_labels, p_size, g_gap, highlight_co="无"):
        d = df_source[df_source['指标名称'] == field_name].copy()
        d['报告年份'], d['value'] = d['报告年份'].astype(str).str.replace(".0", "", regex=False), d['value'] / divisor 
        y_old, y_new, hl_co = str(prev_year), str(latest_year), str(highlight_co).strip()
        fig, x_idx = go.Figure(), list(range(len(selected_cos)))
        
        all_vals = d['value'].dropna()
        all_vals = all_vals[all_vals != 0]
        if not all_vals.empty:
            y_max, y_min = all_vals.max(), all_vals.min()
            off = (y_max - y_min) * 0.15
            
            # 判断整体方向：如果极小值的绝对值 > 极大值的绝对值，说明主要是负数（向下画）
            if abs(y_min) > abs(y_max):
                placeholder_h = y_min * 0.5  # 取最深的一半
            else:
                placeholder_h = y_max * 0.5  # 取最高的一半
                
        else:
            # 如果全军覆没都没数据，给个默认的向下占位
            y_max, y_min = 0, -100
            off = 12
            placeholder_h = -50 
            
        df_old, df_new = d[d['报告年份'] == y_old].set_index('公司'), d[d['报告年份'] == y_new].set_index('公司')
        vals_old, vals_new, txt_old, txt_new = [], [], [], []
        col_old, col_new, pos_old, pos_new, fc_old, fc_new = [], [], [], [], [], []

        for i, co in enumerate(selected_cos):
            v0 = df_old.loc[co, 'value'] if co in df_old.index else np.nan
            v1 = df_new.loc[co, 'value'] if co in df_new.index else np.nan
            m0, m1 = pd.isna(v0), pd.isna(v1)

            # 🌟 修复了文字居中：如果两年都未披露，只在居中位置打一个标签
            if m0 and m1:
                vals_old.append(placeholder_h); vals_new.append(placeholder_h)
                txt_old.append(""); txt_new.append("")
                col_old.append("#CDCDCD"); col_new.append("#CDCDCD")
                pos_old.append("inside"); pos_new.append("inside")
                fc_old.append("white"); fc_new.append("white")
                fig.add_annotation(x=i, y=placeholder_h/2, text="未披露", showarrow=False, font=dict(color="white", size=11), xanchor="center", yanchor="middle")
            
            elif m0 and not m1:
                vals_old.append(placeholder_h); vals_new.append(v1)
                txt_old.append("未披露"); txt_new.append(f"{v1:.1f}" if show_labels else "")
                col_old.append("#CDCDCD"); col_new.append("rgb(0, 51, 141)")
                pos_old.append("inside"); pos_new.append("outside")
                fc_old.append("white"); fc_new.append("#333")
                
            elif not m0 and m1:
                vals_old.append(v0); vals_new.append(placeholder_h)
                txt_old.append(f"{v0:.1f}" if show_labels else ""); txt_new.append("未披露")
                col_old.append("rgb(15, 101, 253)"); col_new.append("#CDCDCD")
                pos_old.append("outside"); pos_new.append("inside")
                fc_old.append("#333"); fc_new.append("white")
                
            else:
                vals_old.append(v0); vals_new.append(v1)
                txt_old.append(f"{v0:.1f}" if show_labels else ""); txt_new.append(f"{v1:.1f}" if show_labels else "")
                col_old.append("rgb(15, 101, 253)"); col_new.append("rgb(0, 51, 141)")
                pos_old.append("outside"); pos_new.append("outside")
                fc_old.append("#333"); fc_new.append("#333")
                
                # 🌟 修复：如果基数 v0 为 0，跳过百分比计算，防止 ZeroDivisionError
                if v0 != 0:
                    pct = (v1 - v0) / abs(v0)
                    arr_color, arr_symbol = ("#FD349C", "↗") if pct >= 0 else ("#269924", "↘")
                    y_pos, v_align = (max(v0, v1) + off, "bottom") if v1 >= 0 else (min(v0, v1) - off, "top")
                    fig.add_annotation(x=i, y=y_pos, text=f"<b>{arr_symbol} {pct:.1%}</b>", showarrow=False, font=dict(color=arr_color, size=p_size), xshift=10, valign=v_align)
        fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", marker=dict(symbol="square", size=10, color="rgb(15, 101, 253)"), name=f"{y_old}年新业务亏损（LC）"))
        fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", marker=dict(symbol="square", size=10, color="rgb(0, 51, 141)"), name=f"{y_new}年新业务亏损（LC）"))

        fig.add_trace(go.Bar(name=f"{y_old}年新业务亏损（LC）", x=x_idx, y=vals_old, marker_color=col_old, text=txt_old, textposition=pos_old, textfont=dict(size=12, color=fc_old), textangle=0, cliponaxis=False, showlegend=False))
        fig.add_trace(go.Bar(name=f"{y_new}年新业务亏损（LC）", x=x_idx, y=vals_new, marker_color=col_new, text=txt_new, textposition=pos_new, textfont=dict(size=12, color=fc_new), textangle=0, cliponaxis=False, showlegend=False))
            
        if hl_co in [str(c).strip() for c in selected_cos]:
            idx = [str(c).strip() for c in selected_cos].index(hl_co)
            fig.add_shape(type="rect", xref="x", yref="paper", x0=idx-0.46, x1=idx+0.46, y0=-0.12, y1=1.05, fillcolor=HL_BOX_FILL  , line=dict(color=HL_BOX_LINE, width=1.5), layer="above")

        fig.update_layout(barmode='group', bargroupgap=0, bargap=g_gap, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', margin=dict(t=40, b=60, l=40, r=40), height=500, legend=dict(orientation="v", yanchor="top", y=1.15, xanchor="right", x=1.0))
        
        # 🌟 修复 Y 轴范围保护，同时适应正数和负数的扩展
        r_top = y_max + (y_max - y_min) * 0.18
        r_bot = y_min - (y_max - y_min) * 0.15 if y_min < 0 else 0
        if placeholder_h > 0: r_top = max(r_top, placeholder_h * 1.3)
        if placeholder_h < 0: r_bot = min(r_bot, placeholder_h * 1.3)
            
        fig.update_yaxes(showgrid=False, zeroline=True, zerolinecolor="#E0E0E0", zerolinewidth=1.02, range=[r_bot, r_top])
        fig.update_xaxes(showline=False, showgrid=False, zeroline=False, tickvals=x_idx, ticktext=[f"<span style='color:#00338D;'><b>{co}</b></span>" for co in selected_cos], ticks="", ticklen=0)
        return fig
    
# 24.新业务占比 (终极完美版：修正了图例大小与亏损占位柱方向)
    def create_new_business_metrics_charts(df_raw, selected_cos, show_lab, lab_sz, bar_width, co_sz, highlight_co="无", is_print_mode=False):
        df_clean = df_raw.copy()
        df_clean['报告年份'] = df_clean['报告年份'].astype(str).str.replace('.0', '', regex=False)
        val_col = "(百万)人民币" if "(百万)人民币" in df_clean.columns else df_clean.columns[-1]
        req_f = ["新业务RA", "新业务亏损合同（LC）——非PAA", "新业务未来现金流入现值", "新业务未来现金流入现值（盈利）", "新业务未来现金流入现值（亏损）", "新业务CSM（集团口径）"]
        
        df_pivot = df_clean[df_clean['字段名'].isin(req_f)].pivot_table(index=['公司', '报告年份'], columns='字段名', values=val_col, aggfunc='sum').reset_index()
        if selected_cos: df_pivot = df_pivot[df_pivot['公司'].isin(selected_cos)]
        if df_pivot.empty: return None, None, None

        # 只补齐列名，不强填 0，保留 np.nan 以区分真实 0 和未披露
        for c in req_f: 
            if c not in df_pivot.columns: 
                df_pivot[c] = np.nan
                
# 🌟 终极修复：使用 np.where。只要分母是 0，直接让结果等于 np.nan（空值，触发未披露）
        num1 = df_pivot['新业务CSM（集团口径）'].astype(float)
        den1 = df_pivot['新业务未来现金流入现值（盈利）'].astype(float)
        df_pivot['新业务CSM利润率'] = np.where(den1 == 0, np.nan, num1 / den1) # 👈 0.0 改成 np.nan
        
        num2 = df_pivot['新业务亏损合同（LC）——非PAA'].astype(float)
        den2 = df_pivot['新业务未来现金流入现值（亏损）'].astype(float)
        df_pivot['新业务LC亏损率'] = np.where(den2 == 0, np.nan, num2 / den2) # 👈 0.0 改成 np.nan
        
        num3 = df_pivot['新业务RA'].astype(float)
        den3 = df_pivot['新业务未来现金流入现值'].astype(float)
        df_pivot['新业务RA率'] = np.where(den3 == 0, np.nan, num3 / den3) # 👈 0.0 改成 np.nan
        
        # 兜底：防止任何意外产生的无穷大 (inf) 破坏图表
        df_pivot.replace([np.inf, -np.inf], 0.0, inplace=True)

        configs = [("新业务CSM利润率", "rgb(30, 73, 225)", "rgb(149, 229, 255)"), ("新业务LC亏损率", "rgb(253, 52, 156)", "rgb(255, 214, 235)"), ("新业务RA率", "rgb(114, 19, 234)", "rgb(227, 207, 251)")]
        figs, hl_co, x_idx = [], str(highlight_co).strip(), list(range(len(selected_cos)))

        for m, c_lat, c_pre in configs:
            df_lat, df_pre = df_pivot[df_pivot['报告年份']==str(latest_year)].set_index('公司').reindex(selected_cos), df_pivot[df_pivot['报告年份']==str(prev_year)].set_index('公司').reindex(selected_cos)
            fig = go.Figure()
            
            val_valid = pd.concat([df_lat[m], df_pre[m]]).dropna()
            val_valid = val_valid[val_valid != 0]
            ph = val_valid.abs().max() * 0.3 if not val_valid.empty else 0.05
            
            if m == "新业务LC亏损率": ph = -ph
            
            yp, yl, tp, tl, cp, cl, pp, pl, fp, fl = [],[],[],[],[],[],[],[],[],[]

            for i, co in enumerate(selected_cos):
                vp, vl = df_pre.loc[co, m] if co in df_pre.index else np.nan, df_lat.loc[co, m] if co in df_lat.index else np.nan
                
                # 判断依据严格变为 pd.isna()，彻底没数(NaN)才算未披露
                mp, ml = pd.isna(vp), pd.isna(vl)
                
                if mp and ml:
                    yp.append(ph); yl.append(ph); tp.append(""); tl.append(""); cp.append("#CDCDCD"); cl.append("#CDCDCD"); pp.append("inside"); pl.append("inside"); fp.append("white"); fl.append("white")
                    fig.add_annotation(x=i, y=ph/2, text="未披露", showarrow=False, font=dict(color="white", size=11))
                elif mp and not ml:
                    yp.append(ph); yl.append(vl); tp.append("未披露"); tl.append(f"{vl*100:.1f}%" if show_lab else ""); cp.append("#CDCDCD"); cl.append(c_lat); pp.append("inside"); pl.append("outside"); fp.append("white"); fl.append("#333")
                elif not mp and ml:
                    yp.append(vp); yl.append(ph); tp.append(f"{vp*100:.1f}%" if show_lab else ""); tl.append("未披露"); cp.append(c_pre); cl.append("#CDCDCD"); pp.append("outside"); pl.append("inside"); fp.append("#333"); fl.append("white")
                else:
                    yp.append(vp); yl.append(vl); tp.append(f"{vp*100:.1f}%" if show_lab else ""); tl.append(f"{vl*100:.1f}%" if show_lab else ""); cp.append(c_pre); cl.append(c_lat); pp.append("outside"); pl.append("outside"); fp.append("#333"); fl.append("#333")

            fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", marker=dict(symbol="square", size=10, color=c_pre), name=f"{prev_year}年"))
            fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", marker=dict(symbol="square", size=10, color=c_lat), name=f"{latest_year}年"))
            
            fig.add_trace(go.Bar(x=x_idx, y=yp, marker_color=cp, text=tp, textposition=pp, textfont=dict(size=lab_sz, color=fp), showlegend=False, cliponaxis=False, constraintext='none'))
            fig.add_trace(go.Bar(x=x_idx, y=yl, marker_color=cl, text=tl, textposition=pl, textfont=dict(size=lab_sz, color=fl), showlegend=False, cliponaxis=False, constraintext='none'))

            if hl_co in [str(c).strip() for c in selected_cos]:
                idx = [str(c).strip() for c in selected_cos].index(hl_co)
                fig.add_shape(type="rect", xref="x", yref="paper", x0=idx-0.46, x1=idx+0.46, y0=-0.40, y1=1.15, fillcolor="rgba(0, 51, 141, 0.05)", line=dict(color="rgba(0, 51, 141, 0.85)", width=1.5), layer="above")

            mb, leg_y, leg_a = 17, 1.02, "bottom"
            if not df_lat[m].dropna().empty:
                avg, y_r = df_lat[m].mean(), (max(val_valid) - min(val_valid)) if not val_valid.empty else 1
                if avg < 0: mb, leg_y, leg_a = 60, -0.15, "top"
                
                if m == "新业务LC亏损率" and avg < 0: leg_y -= 0.11; mb += 17
                
                fig.add_hline(y=avg, line_dash="dash", line_color=c_lat, line_width=1.5)
                l_m = max((df_pre[m].iloc[-1] if not pd.isna(df_pre[m].iloc[-1]) else 0), (df_lat[m].iloc[-1] if not pd.isna(df_lat[m].iloc[-1]) else 0))
                dyn_a, dyn_y = ("bottom", 2) if abs(avg - l_m) >= (y_r * 0.15) else (("bottom", 15) if avg >= l_m else ("top", -15))
                fig.add_annotation(x=0.98, xref="paper", y=avg, yref="y", xanchor="right", yanchor=dyn_a, yshift=dyn_y, text=f"{str(latest_year)[-2:]}年平均 {avg*100:.1f}%", showarrow=False, bgcolor="rgba(255,255,255,0.85)", bordercolor=c_lat, borderwidth=1, borderpad=2, font=dict(color=c_lat, size=max(lab_sz-3, 8)))

            fig.update_layout(title=dict(text=f"<b>{m}</b>", x=0.5, xanchor='center', y=0.98 if is_print_mode else 0.95, font=dict(size=11, color="#00338D", family="Microsoft YaHei")), plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", barmode='group', bargroupgap=0, bargap=max(0, 1.0 - bar_width), margin=dict(t=30 if is_print_mode else 50, b=0 if is_print_mode else mb, l=20, r=40),height=210 if is_print_mode else 360, yaxis=dict(tickformat=".0%", showgrid=False, zeroline=True, zerolinecolor="#E0E0E0", zerolinewidth=1.02), legend=dict(orientation="h", yanchor=leg_a, y=leg_y-0.15, xanchor="right", x=1,font=dict(size=11) ))
            fig.update_xaxes(showgrid=False, zeroline=False, ticktext=[f"<span style='font-size:{co_sz}px; color:#00338D;'><b>{c}</b></span>" for c in selected_cos], tickvals=x_idx, ticks="", ticklen=0)
            figs.append(fig)

        return figs[0], figs[1], figs[2]

    # --- 25.新业务利润率指标趋势 ---   
    def create_nb_margin_trend_chart(df, cos, title, color_map, show_labels, marker_size, legend_font_size, highlight_co="无"):
        years = sorted(df['报告年份'].unique())
        x_categories = [f"{y}YE" for y in years]        
        plot_data = []
        for co in cos:
            for y in years:
                df_curr = df[(df['公司'] == co) & (df['报告年份'] == y)]
                v = lambda k: df_curr[df_curr['字段名'].str.contains(k, regex=False, na=False)]['(百万)人民币'].sum()
                csm, lc, pv = v('新业务CSM'), v('新业务亏损合同'), v('新业务未来现金流入') 
                if pv != 0: plot_data.append({'公司': co, '报告年份': f"{y}YE", 'year_num': y, 'value': (csm + lc) / pv})                   
        
        df_plot = pd.DataFrame(plot_data, columns=['公司', '报告年份', 'year_num', 'value'])
        fig = go.Figure()
        max_co, min_co = None, None
        if not df_plot.empty:
            latest_yr = df_plot['year_num'].max()
            df_latest = df_plot[df_plot['year_num'] == latest_yr].dropna(subset=['value'])
            if not df_latest.empty:
                max_co = df_latest.loc[df_latest['value'].idxmax(), '公司']
                min_co = df_latest.loc[df_latest['value'].idxmin(), '公司']

        for co in cos:
            d_co = df_plot[df_plot['公司'] == co].sort_values('year_num')
            is_missing = d_co.empty or d_co['value'].isna().all()
            
            if is_missing:
                fig.add_trace(go.Scatter(
                    x=x_categories, y=[None]*len(x_categories), 
                    name=f"<span style='color:#999999'>{co} (未披露)</span>", 
                    mode='lines+markers', 
                    line=dict(color="#CDCDCD", width=2, dash="dot"), 
                    marker=dict(size=marker_size, symbol='circle'),
                    showlegend=True
                ))
            else:
                is_highlight = (co == highlight_co)
                is_ext = (co in [max_co, min_co]) # 🌟 判断当前公司是否是最高或最低
                
                # 标签显示逻辑：如果是高亮、最高、最低，或者头尾端点，就显示标签
                text_labels = [f"{v:.1%}" if (is_highlight or is_ext or i==0 or i==len(d_co)-1) else "" for i, v in enumerate(d_co['value'])]
                
                fig.add_trace(go.Scatter(
                    x=d_co['报告年份'], y=d_co['value'], 
                    name=co, 
                    mode='lines+markers+text' if show_labels else 'lines+markers', 
                    line=dict(
                        color=HIGHLIGHT_COLOR if is_highlight else color_map.get(co, "#333"), 
                        width=4 if is_highlight else (2.5 if is_ext else 2), # 🌟 极值稍微加粗到 2.5
                        dash="solid" if (is_highlight or is_ext) else "dot"  # 🌟 极值变成实线 (solid)
                    ), 
                    marker=dict(size=marker_size*1.5 if is_highlight else marker_size, symbol='circle'), 
                    text=text_labels if show_labels else None, 
                    textposition="top center", 
                    cliponaxis=False,
                    showlegend=True
                ))
                
        fig.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', margin=dict(t=40, b=40, l=70, r=70), height=450, showlegend=True, legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, font=dict(size=legend_font_size)), width=600,  xaxis=dict(showgrid=False, showline=False, zeroline=False), yaxis=dict(showgrid=False, zeroline=False, tickformat=".1%"))
        fig.update_xaxes(type='category', categoryorder='array', categoryarray=x_categories)
        
        valid_vals = df_plot['value'].dropna()
        if not valid_vals.empty:
            y_max, y_min = valid_vals.max(), valid_vals.min()
            bottom_range = y_min * 0.8 if y_min >= 0 else y_min * 1.2
            fig.update_yaxes(range=[bottom_range, y_max + abs(y_max) * 0.15])
            
        return fig

# --- 26. 费用结构 --- 
    def create_expense_breakdown_chart(df, selected_cos, show_labels, label_size, bar_width, co_font_size, highlight_co="无"):
        y, f = [str(prev_year), str(latest_year)], ["获取费用", "维持费用", "非履约费用"]
        d_sub = df[(df['报告年份'].astype(str).isin(y)) & (df['公司'].isin(selected_cos)) & (df['字段名'].isin(f))].copy()
        d_sub['value'] = d_sub['(百万)人民币'] / divisor
        
        avg_acq, avg_maint, avg_non = d_sub[d_sub['字段名'] == "获取费用"]['value'].sum(), d_sub[d_sub['字段名'] == "维持费用"]['value'].sum(), d_sub[d_sub['字段名'] == "非履约费用"]['value'].sum()
        avg_total = avg_acq + avg_maint + avg_non
        p_acq, p_maint, p_non = (avg_acq/avg_total*100) if avg_total>0 else 0, (avg_maint/avg_total*100) if avg_total>0 else 0, (avg_non/avg_total*100) if avg_total>0 else 0       
        # 计算全局最高柱子高度，用来兜底
        global_max = max([d_sub[(d_sub['公司']==co) & (d_sub['报告年份'].astype(str)==yr)]['value'].sum() for co in selected_cos for yr in y] + [1.0])
        hl_co = str(highlight_co).strip()
        
        # 🌟 设置灰色占位柱的高度：全局最高的一半
        placeholder_h = global_max * 0.5
        
        titles = [f"<span style='color:#00338D; font-size:{co_font_size}px;'><b>{co}</b></span>" for co in selected_cos]
        fig = make_subplots(rows=1, cols=len(selected_cos), horizontal_spacing=0.02, subplot_titles=titles)
        
        # 统一设置图例
        for leg_name, leg_color in [("获取费用", "rgb(30, 73, 226)"), ("维持费用", "rgb(118, 210, 255)"), ("非履约费用", "rgb(114, 19, 234)")]:
            fig.add_trace(go.Scatter(x=[None], y=[None], mode="markers", marker=dict(symbol="square", size=12, color=leg_color), name=leg_name, showlegend=True), row=1, col=1)
            
        for i, co in enumerate(selected_cos):
            col_idx = i + 1
            xref_key, yref_key, xref_d, yref_d = f"x{col_idx}" if col_idx>1 else "x", f"y{col_idx}" if col_idx>1 else "y", f"x{col_idx} domain" if col_idx>1 else "x domain", f"y{col_idx} domain" if col_idx>1 else "y domain"
            cd = d_sub[d_sub['公司'] == co]
            gv = lambda yr, field: cd[(cd['报告年份'].astype(str)==str(yr)) & (cd['字段名']==field)]['value'].sum()
            v24a, v25a, v24m, v25m, v24n, v25n = gv(prev_year,"获取费用"), gv(latest_year,"获取费用"), gv(prev_year,"维持费用"), gv(latest_year,"维持费用"), gv(prev_year,"非履约费用"), gv(latest_year,"非履约费用")
            
            t24, t25 = v24a+v24m+v24n, v25a+v25m+v25n
            m24, m25 = (t24 == 0), (t25 == 0) # 🌟 标记这两年是不是全空（未披露）
            
            lbl = lambda v, t: f"{v/t*100:.0f}%" if t>0 and v>0 else ""
            x_axis = [f"{prev_year}YE", f"{latest_year}YE"]
            
            # 🌟 画真实数据柱子（如果是未披露，传进去的是 0，在堆叠图里会自动隐形）
            fig.add_trace(go.Bar(x=x_axis, y=[v24a, v25a], marker_color="rgb(30, 73, 226)", width=bar_width, showlegend=False, text=[lbl(v24a,t24), lbl(v25a,t25)] if show_labels else None, textposition='inside', textangle=0, textfont=dict(size=label_size, color="white"), constraintext='none'), row=1, col=col_idx)
            fig.add_trace(go.Bar(x=x_axis, y=[v24m, v25m], marker_color="rgb(118, 210, 255)", width=bar_width, showlegend=False, text=[lbl(v24m,t24), lbl(v25m,t25)] if show_labels else None, textposition='inside', textangle=0, textfont=dict(size=label_size, color="#1a1a2e"), constraintext='none'), row=1, col=col_idx)
            fig.add_trace(go.Bar(x=x_axis, y=[v24n, v25n], marker_color="rgb(114, 19, 234)", width=bar_width, showlegend=False, text=[lbl(v24n,t24), lbl(v25n,t25)] if show_labels else None, textposition='inside', textangle=0, textfont=dict(size=label_size, color="white"), constraintext='none'), row=1, col=col_idx)                        
            
            # 🌟 专门画一层隐形的灰色占位柱子
            fig.add_trace(go.Bar(
                x=x_axis, 
                y=[placeholder_h if m24 else 0, placeholder_h if m25 else 0], 
                marker_color="#CDCDCD", width=bar_width, showlegend=False, 
                text=["未披露" if m24 else "", "未披露" if m25 else ""], 
                textposition='inside', insidetextanchor='middle', textangle=0,textfont=dict(size=12, color="white"), constraintext='none', cliponaxis=False
            ), row=1, col=col_idx)

            # 🌟 连线逻辑：只有当两年都有真实数据时，才画段之间的连接线
            if not m24 and not m25:
                for y0, y1 in [(v24a, v25a), (v24a+v24m, v25a+v25m), (t24, t25)]:
                    fig.add_shape(type="line", xref=xref_key, yref=yref_key, x0=bar_width/2, y0=y0, x1=1-bar_width/2, y1=y1, line=dict(color="rgba(170,170,170,0.85)", width=0.9), layer="above")
            
            # 顶部总额标签（未披露时不画顶部数字）
            for x_cat, total_v, is_missing in [(f"{prev_year}YE", t24, m24), (f"{latest_year}YE", t25, m25)]:
                if not is_missing:
                    fig.add_annotation(x=x_cat, y=total_v+global_max*0.02, xref=xref_key, yref=yref_key, text=f"<b>{total_v:.0f}</b>", showarrow=False, font=dict(size=label_size, color="#222"), bgcolor="white", bordercolor="#BBB", borderwidth=1, xanchor="center", yanchor="bottom")
            
            # 🌟 涨幅箭头逻辑：只有当两年都有数据时，才计算和绘制增幅箭头
            if not m24 and not m25:
                pct = ((t25-t24)/t24*100) if t24>0 else 0
                base_arrow_y, slope = max(t24, t25) + global_max*0.12, global_max*0.01
                if pct > 0: y_start, y_end, arr_clr, sign = base_arrow_y-slope, base_arrow_y+slope, "rgb(253, 52, 155)", "+"
                elif pct < 0: y_start, y_end, arr_clr, sign = base_arrow_y+slope, base_arrow_y-slope, "rgb(0, 180, 100)", ""
                else: y_start, y_end, arr_clr, sign = base_arrow_y, base_arrow_y, "rgb(253, 52, 155)", ""
                fig.add_annotation(x=0.65, y=y_end, ax=0.35, ay=y_start, xref=xref_key, yref=yref_key, axref=xref_key, ayref=yref_key, text="", showarrow=True, arrowhead=2, arrowsize=0.7, arrowwidth=1.5, arrowcolor=arr_clr)
                fig.add_annotation(x=0.5, xref=xref_d, y=max(y_start, y_end)+global_max*0.02, yref=yref_key, text=f"<b>{sign}{pct:.0f}%</b>", showarrow=False, font=dict(size=label_size+1, color=arr_clr), xanchor="center", yanchor="bottom")
            
            # 框线与背景
            is_hl = (str(co).strip() == hl_co)
            bg_fill = HL_BOX_FILL   if is_hl else ("rgba(200, 200, 200, 0.18)" if i % 2 == 1 else "rgba(255,255,255,0)")
            border_dict = dict(color=HL_BOX_LINE, width=1.5) if is_hl else dict(color="#CCCCCC", width=1)
            fig.add_shape(type="rect", xref=xref_d, yref=yref_d, x0=0, x1=1, y0=-0.1, y1=1.13, line=border_dict, fillcolor=bg_fill, layer="above" if is_hl else "below", row=1, col=col_idx)

            fig.update_yaxes(range=[0, global_max*1.3], showticklabels=False, showgrid=False, zeroline=False, row=1, col=col_idx)
            fig.update_xaxes(type="category", showgrid=False, zeroline=False, ticks="", ticklen=0, row=1, col=col_idx)
            
        fig.update_layout(barmode='stack', height=550, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', margin=dict(t=60, b=80, l=10, r=10), legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5, font=dict(size=12), itemsizing="constant"))
        for ann in fig.layout.annotations:
            if "span" in str(ann.text): ann.update(y=1.03)

        return fig, (p_acq, p_maint, p_non)

    # --- 27. 折现率假设表 (纯HTML生成不变) ---
    def create_discount_rate_table(df_raw, target_year, cos, highlight_co="无"):
        df = df_raw.copy()
        df['报告年份'] = df['报告年份'].astype(str).str.replace('.0', '', regex=False)
        curr_year_str, prev_year_str = str(target_year), str(int(target_year) - 1)
        df_target = df[(df['报告年份'].isin([curr_year_str, prev_year_str])) & (df['公司'].isin(cos)) & (df['字段名'].str.contains("折现率", na=False))]
        val_col = "(百万)原币" if "(百万)原币" in df_target.columns else df_target.columns[-1]
        data_map = df_target.set_index(['公司', '报告年份'])[val_col].to_dict()
        current_hl = str(highlight_co).strip()
        def format_discount_val(val):
            if pd.isna(val) or str(val).lower() in ['nan', '', '-']: return "未披露", True
            v_str = str(val).strip()
            try:
                if float(v_str.replace('%', '').replace(',', '')) == 0: return "未披露", True
            except:
                pass
            
            return v_str, False

        html = "<table style='width:100%; border-collapse: collapse; font-family: sans-serif; margin-bottom: 20px; font-size: 13px;'>"
        html += f"<tr style='background-color: #00338D; color: white; font-size: 14px; text-align: center; font-weight: bold;'><th style='padding: 10px; text-align: left; border: 1px solid white; width: 30%;'>公司名称</th><th style='padding: 10px; text-align: center; border: 1px solid white; width: 35%;'>{curr_year_str}年12月31日</th><th style='padding: 10px; text-align: center; border: 1px solid white; width: 35%;'>{prev_year_str}年12月31日</th></tr>"

        for row_idx, co in enumerate(cos):
            # 获取数值和是否未披露的标记
            v_c, m_c = format_discount_val(data_map.get((co, curr_year_str), "-"))
            v_p, m_p = format_discount_val(data_map.get((co, prev_year_str), "-"))
            is_hl = (str(co).strip() == current_hl)
            
            row_bg, text_color, font_weight, borders = ("#F8F9FA" if row_idx % 2 == 0 else "white"), "#333", "normal", "border: 1px solid #EAEAEA;"
            if is_hl: row_bg, text_color, font_weight, borders = "rgba(0, 51, 141, 0.08)", "#00338D", "bold", "border-top: 2px solid #00338D; border-bottom: 2px solid #00338D; border-left: 1px solid #EAEAEA; border-right: 1px solid #EAEAEA;"

            style_c = f"background-color: #CDCDCD; color: white;" if m_c else f"color: {text_color};"
            style_p = f"background-color: #CDCDCD; color: white;" if m_p else f"color: {text_color};"
            
            html += f"<tr style='background-color: {row_bg}; font-weight: {font_weight};'><td style='padding: 12px 10px; text-align: left; {borders} font-weight: bold; color: {text_color};'>{co}</td><td style='padding: 12px 10px; text-align: center; {borders} {style_c}'>{v_c}</td><td style='padding: 12px 10px; text-align: center; {borders} {style_p}'>{v_p}</td></tr>"
            
        return html + "</table>"
    
    # --- 28. 非金融风险置信水平表 (纯HTML生成不变) ---
    def create_confidence_level_table(df_raw, target_year, cos, highlight_co="无"):
        df = df_raw.copy()
        df['报告年份'] = df['报告年份'].astype(str).str.replace('.0', '', regex=False)
        curr_year_str, prev_year_str = str(target_year), str(int(target_year) - 1)
        df_target = df[(df['报告年份'].isin([curr_year_str, prev_year_str])) & (df['公司'].isin(cos)) & (df['字段名'].str.contains("非金融风险的置信水平", na=False))]
        val_col = "(百万)原币" if "(百万)原币" in df_target.columns else df_target.columns[-1]
        data_map = df_target.set_index(['公司', '报告年份'])[val_col].to_dict()
        current_hl = str(highlight_co).strip()
        
        def format_conf_val(val):
            # 🌟 拦截 1：没填、NaN、占位符
            if pd.isna(val) or str(val).lower() in ['nan', '', '-']: return "未披露", True
            v_str = str(val).strip()
            # 🌟 拦截 2：算出来等于 0（能完美识别 0, 0.0, 0% 等）
            try:
                if float(v_str.replace('%', '').replace(',', '')) == 0: return "未披露", True
            except: pass
            
            # 正常格式化
            if '%' in v_str: return v_str, False
            try: return (f"{float(v_str) * 100:.0f}%" if 0 < float(v_str) <= 1.0 else f"{float(v_str):.0f}%"), False
            except: return v_str, False

        html = "<table style='width:100%; border-collapse: collapse; font-family: sans-serif; margin-bottom: 20px; font-size: 13px;'>"
        html += f"<tr style='background-color: #00338D; color: white; font-size: 14px; text-align: center; font-weight: bold;'><th style='padding: 10px; text-align: left; border: 1px solid white; width: 30%;'>公司名称</th><th style='padding: 10px; text-align: center; border: 1px solid white; width: 35%;'>{curr_year_str}年12月31日</th><th style='padding: 10px; text-align: center; border: 1px solid white; width: 35%;'>{prev_year_str}年12月31日</th></tr>"

        for row_idx, co in enumerate(cos):
            # 拿到格式化后的值，以及是否“未披露”的布尔判断
            v_c, m_c = format_conf_val(data_map.get((co, curr_year_str), "-"))
            v_p, m_p = format_conf_val(data_map.get((co, prev_year_str), "-"))
            is_hl = (str(co).strip() == current_hl)
            
            row_bg, text_color, font_weight, borders = ("#F8F9FA" if row_idx % 2 == 0 else "white"), "#333", "normal", "border: 1px solid #EAEAEA;"
            if is_hl: row_bg, text_color, font_weight, borders = "rgba(0, 51, 141, 0.08)", "#00338D", "bold", "border-top: 2px solid #00338D; border-bottom: 2px solid #00338D; border-left: 1px solid #EAEAEA; border-right: 1px solid #EAEAEA;"
            
            # 🌟 动态注入单元格样式：如果是未披露，强制灰底灰字
            style_c = f"background-color: #CDCDCD; color: white;" if m_c else f"color: {text_color};"
            style_p = f"background-color: #CDCDCD; color: white;" if m_p else f"color: {text_color};"
            
            html += f"<tr style='background-color: {row_bg}; font-weight: {font_weight};'><td style='padding: 12px 10px; text-align: left; {borders} font-weight: bold; color: {text_color};'>{co}</td><td style='padding: 12px 10px; text-align: center; {borders} {style_c}'>{v_c}</td><td style='padding: 12px 10px; text-align: center; {borders} {style_p}'>{v_p}</td></tr>"
            
        return html + "</table>"


# --- 28. 合同服务边际CSM 预期摊销速度折线图 ---
    def create_csm_maturity_table(df_raw, target_year, selected_cos, color_map, show_labels, marker_size, highlight_co="无", title_text=""):
        COMMON_TITLE_FONT = dict(size=18, color="#00338D", family="Microsoft YaHei")
        
        df = df_raw.copy()
        df['报告年份'] = df['报告年份'].astype(str).str.replace('.0', '', regex=False)
        curr_year_str = str(target_year).replace('.0', '')
        
        fields = [
            "1年及1年以内合同服务边际", "1-5年合同服务边际", 
            "5-10年合同服务边际", "10-20年合同服务边际", "20年合同服务边际"
        ]
        
        df_target = df[(df['报告年份'] == curr_year_str) & (df['公司'].isin(selected_cos)) & (df['字段名'].isin(fields))].copy()
        
        preferred_cols = ["(%)原币", "(%)", "百分比", "比例", "占比", "数值", "值", "(百万)原币"]
        val_col = next((c for c in preferred_cols if c in df_target.columns), None)
        if not val_col:
            non_dim_cols = [c for c in df_target.columns if c not in ['报告年份', '公司', '字段名']]
            val_col = non_dim_cols[-1] if non_dim_cols else None
            
        if not val_col: return go.Figure()

        data_map = df_target.set_index(['公司', '字段名'])[val_col].to_dict()
        
        def parse_pct(v):
            if pd.isna(v): return None
            s = str(v).strip()
            if s in ["", "nan", "-"]: return None
            try:
                if "%" in s: return float(s.replace("%", "").replace(",", ""))
                num = float(s.replace(",", ""))
                if abs(num) <= 1.0: return num * 100
                return num
            except:
                return None

        fig = go.Figure()
        hl_co = str(highlight_co).strip()

        # 🌟 第一步：遍历计算数据，并标记哪些公司是“未披露/数据为0”
        plot_info = {}
        valid_cos = []  # 专门存放有真实有效数据的公司名单
        for co in selected_cos:
            f1 = parse_pct(data_map.get((co, "1年及1年以内合同服务边际")))
            f5 = parse_pct(data_map.get((co, "1-5年合同服务边际")))
            f10 = parse_pct(data_map.get((co, "5-10年合同服务边际")))
            f20 = parse_pct(data_map.get((co, "10-20年合同服务边际")))
            f20p = parse_pct(data_map.get((co, "20年合同服务边际")))
            
            has_0_5 = (f1 is not None and f1 > 0) or (f5 is not None and f5 > 0)
            
            v1  = f1  if f1  is not None else 0
            v5  = f5  if f5  is not None else 0
            v10 = f10 if f10 is not None else 0
            v20 = f20 if f20 is not None else 0
            v20p = f20p if f20p is not None else 0
            
            actual_v10 = ((v20 + v20p) / 2) if (f10 is None or v10 == 0) else v10
            x_vals = [0, 5, 10]
            
            if has_0_5:
                y5  = v1 + v5
                y10 = y5 + actual_v10
            else:
                y10 = actual_v10
                y5  = y10 / 2.0
            
            # 🌟 拦截：如果10年累计依然是0，判定为无数据/未披露
            is_missing = (y10 == 0)
            if not is_missing:
                valid_cos.append(co)
                
            plot_info[co] = {'x_vals': x_vals, 'y_vals': [0, y5, y10], 'y5': y5, 'y10': y10, 'is_missing': is_missing}

        # 🌟 找出 X=10 时最高和最低的公司（只在有效公司里找！）
        max_co = max(valid_cos, key=lambda k: plot_info[k]['y10']) if valid_cos else None
        min_co = min(valid_cos, key=lambda k: plot_info[k]['y10']) if valid_cos else None

        # 🌟 第二步：正式画图并分配虚实线
        for co in selected_cos:
            p_data = plot_info[co]
            
            if p_data['is_missing']:
                # 🌟 生成“隐形图例”：不画真实的线，名字标灰并加(未披露)
                fig.add_trace(go.Scatter(
                    x=[None], y=[None], 
                    name=f"<span style='color:#999999'>{co} (未披露)</span>", 
                    mode='lines+markers', 
                    line=dict(color="#CDCDCD", width=2, dash="dot"), 
                    marker=dict(size=marker_size, symbol='circle'),
                    showlegend=True
                ))
            else:
                # 🌟 正常绘制有数据的公司
                y5, y10 = p_data['y5'], p_data['y10']
                is_hl = (co == hl_co)
                is_ext = (co in [max_co, min_co])
                
                default_color = f"hsl({hash(co) % 360}, 70%, 50%)" 
                line_color = "#00338D" if is_hl else color_map.get(co, default_color)
                
                # 高亮和极值线稍微加粗，高亮为4，极值为2.5
                line_width = 4 if is_hl else (2.5 if is_ext else 2) 
                
                # 🌟 判断线型：最高、最低以及高亮的公司用实线，其余全部用虚线
                line_dash = "solid" if (is_ext or is_hl) else "dash"
                
                fig.add_trace(go.Scatter(
                    x=p_data['x_vals'],
                    y=p_data['y_vals'],
                    name=co,
                    mode='lines+markers+text' if show_labels else 'lines+markers',
                    text=["", f"{y5:.1f}%", f"{y10:.1f}%"], # 起点 0 不显示文字，显得干净
                    textposition="top center",
                    line=dict(color=line_color, width=line_width, dash=line_dash), 
                    marker=dict(size=marker_size * 1.5 if is_hl else marker_size, color=line_color),
                    cliponaxis=False
                ))
                
        layout_args = dict(
            paper_bgcolor='rgba(0,0,0,0)', 
            plot_bgcolor='rgba(0,0,0,0)', 
            margin=dict(t=80 if title_text else 50, b=40, l=40, r=40), 
            height=450,
            # 将图例移到顶部并换行展示，防止重叠
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5, font=dict(size=11), itemwidth=30)
        )
        if title_text:
            layout_args['title'] = dict(text=f"<b>{title_text}</b>", x=0.5, y=0.95, xanchor='center', font=COMMON_TITLE_FONT)
            
        fig.update_layout(**layout_args)
        
        # 强制格式化 X 轴和 Y 轴
        fig.update_xaxes(
            tickvals=[0, 5, 10], 
            ticktext=["0", "5年", "10年"], 
            showgrid=True, gridcolor="rgba(200,200,200,0.3)", zeroline=False
        )
        fig.update_yaxes(
            showgrid=True, gridcolor="rgba(200,200,200,0.3)", 
            zeroline=True, zerolinecolor="#E0E0E0", 
            tickformat=".0f", ticksuffix="%"
        )
        
        return fig



# --- 29. 6维度增长率 散点图 ---
    def create_six_dimensional_charts(df_raw, target_year, cos, divisor=1, unit_label="百万元", highlight_co="无", label_size=11, show_labels=False, dot_size=11):
        import pandas as pd, numpy as np, plotly.graph_objects as go, plotly.express as px
        df = df_raw.copy()
        df[["字段名", "公司", "报告年份"]] = df[["字段名", "公司", "报告年份"]].astype(str).apply(lambda x: x.str.strip().str.replace(".0", "", regex=False))
        needed = ["净利润", "期初股东权益", "期末股东权益", "CSM期初余额", "CSM期末余额", "总资产", "投资收益率", "综合偿付能力充足率"]
        df = df[(df["报告年份"] == str(target_year)) & df["公司"].isin(cos) & df["字段名"].isin(needed)].drop_duplicates(subset=["公司", "报告年份", "字段名"])
        if df.empty: return []

        a_cols, r_cols = ["(百万)人民币", "(亿元)人民币", "(百万)原币", "(亿元)原币", "人民币", "原币", "数值", "值"], ["(%)原币", "(%)", "百分比", "比例", "占比", "比率"]
        def get_v(r, is_r):
            for c in (r_cols + a_cols if is_r else a_cols + r_cols):
                if c in r.index and pd.notna(r[c]) and str(r[c]).strip().lower() not in ["", "nan", "none", "null", "-"]:
                    s = str(r[c]).strip().replace(",", "")
                    try:
                        v = float(s[:-1])/100.0 if s.endswith("%") else float(s)
                        return v/100.0 if is_r and ("%" in str(r[c]) or abs(v)>1) else v
                    except: continue
            return np.nan

        records = [{"公司": r["公司"], "字段名": r["字段名"], "数值": get_v(r, r["字段名"] in ["投资收益率", "综合偿付能力充足率"])} for _, r in df.iterrows()]
        df_p = pd.DataFrame(records).dropna(subset=["数值"]).pivot_table(index="公司", columns="字段名", values="数值", aggfunc="first").reindex(cos)
        
        # 🌟 修复：抹去索引的名字，防止后续 sort_values 发生歧义冲突！
        df_p.index.name = None 
        
        c = lambda n: pd.to_numeric(df_p[n], errors="coerce") if n in df_p.columns else pd.Series(np.nan, index=df_p.index)

        np_val, eq_b, eq_e, csm_b, csm_e, ta, ir, sr = c("净利润"), c("期初股东权益"), c("期末股东权益"), c("CSM期初余额"), c("CSM期末余额"), c("总资产"), c("投资收益率"), c("综合偿付能力充足率")
        pd_data = pd.DataFrame({"公司": df_p.index, "净利润": np_val, "CSM期末余额": csm_e, "期末股东权益": eq_e, "投资收益率": ir, "综合偿付能力充足率": sr})
        
        pd_data["利润率"] = np.where((eq_b+eq_e)!=0, np_val / ((eq_b+eq_e)/2), np.nan)
        pd_data["CSM增长率"] = np.where(csm_b!=0, (csm_e-csm_b)/csm_b, np.nan)
        pd_data["财务杠杆率"] = np.where(ta!=0, eq_e/ta, np.nan)
        pd_data["净资产增长率"] = np.where(eq_b!=0, (eq_e-eq_b)/eq_b, np.nan)

        cfgs = [("股东回报","利润率","净利润",".1%"), ("盈利潜力","CSM增长率","CSM期末余额",".1%"), ("财务杠杆","财务杠杆率","期末股东权益",".1%"), ("投资能力","投资收益率","期末股东权益",".1%"), ("财务稳定","净资产增长率","期末股东权益",".1%"), ("偿付能力","综合偿付能力充足率","期末股东权益",".0%")]
        cmap, hl, rd = {co: px.colors.qualitative.Plotly[i % 10] for i, co in enumerate(cos)}, str(highlight_co).strip(), divisor or 1
        figs = []

        for t, y_c, x_c, y_fmt in cfgs:
            d_plt = pd_data[["公司", y_c, x_c]].copy().replace([np.inf, -np.inf], np.nan).dropna()
            fig = go.Figure()
            title_html = f"<span style='font-size:14px'><b>{t}</b></span><br><span style='font-size:11px;color:#666'>Y轴={y_c}，X轴={x_c}（单位：{unit_label}）</span>"
            
            if d_plt.empty:
                fig.update_layout(title=dict(text=title_html, x=0.02), height=250, margin=dict(l=20, r=15, t=40, b=10), paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
                figs.append(fig); continue

            d_plt["x_p"] = pd.to_numeric(d_plt[x_c], errors="coerce") / rd
            for _, r in d_plt.sort_values("公司").iterrows():
                co, is_hl = str(r["公司"]), (str(r["公司"]) == hl)
                fig.add_trace(go.Scatter(
                    x=[r["x_p"]], y=[r[y_c]], mode="markers+text" if show_labels else "markers", name=co, text=[co] if show_labels else None, textposition="top center", textfont=dict(size=label_size, color="#333"),
                    marker=dict(size=dot_size*1.45 if is_hl else dot_size, color=cmap.get(co,"#1f77b4"), line=dict(color="white", width=1.8 if is_hl else 1.2), opacity=0.95),
                    customdata=[[f"{r['x_p']:,.2f}" if pd.notna(r["x_p"]) else "-", f"{r[y_c]:{y_fmt}}" if pd.notna(r[y_c]) else "-", f"{r[x_c]:,.2f}" if pd.notna(r[x_c]) else "-"]],
                    hovertemplate=f"<b>{co}</b><br>{x_c}（{unit_label}）: %{{customdata[0]}}<br>{y_c}: %{{customdata[1]}}<br>{x_c}（原值）: %{{customdata[2]}}<extra></extra>"
                ))

            fig.update_layout(title=dict(text=title_html, x=0.02), width=260, height=250, margin=dict(l=20, r=20, t=45, b=15), paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", showlegend=False, hovermode="closest", xaxis=dict(showgrid=True, gridcolor="rgba(180,180,180,0.2)", zeroline=True, zerolinecolor="rgba(150,150,150,0.3)"), yaxis=dict(tickformat=y_fmt, showgrid=True, gridcolor="rgba(180,180,180,0.2)", zeroline=True, zerolinecolor="rgba(150,150,150,0.3)"))
            fig.add_hline(y=0, line_dash="dash", line_color="rgba(120,120,120,0.7)", line_width=1)
            fig.add_vline(x=0, line_dash="dash", line_color="rgba(120,120,120,0.7)", line_width=1)
            figs.append(fig)
            
        return figs

    def create_six_dimensional_legend(cos, highlight_co="无"):
        import plotly.graph_objects as go, plotly.express as px
        cmap, hl = {co: px.colors.qualitative.Plotly[i % 10] for i, co in enumerate(cos)}, str(highlight_co).strip()
        fig = go.Figure([go.Scatter(x=[None], y=[None], mode="markers", name=co, marker=dict(size=12 if co==hl else 10, color=cmap.get(co,"#1f77b4"), line=dict(color="white", width=1.8 if co==hl else 1.2), opacity=0.95)) for co in cos])
        fig.update_layout(height=50, margin=dict(l=0, r=0, t=0, b=0), paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", showlegend=True, legend=dict(orientation="h", x=0.5, xanchor="center", y=0.5, yanchor="middle", bgcolor="rgba(0,0,0,0)", font=dict(size=11)), xaxis=dict(visible=False), yaxis=dict(visible=False))
        return fig






# --- 29. 业绩明细表 (纯HTML生成不变) ---
    def create_financial_report_table(df_raw, target_year, cos, div, unit_str, highlight_co="无"):
        df = df_raw.copy()
        df['报告年份'] = df['报告年份'].astype(str).str.replace('.0', '', regex=False)
        df_year = df[(df['报告年份'] == str(target_year)) & (df['公司'].isin(cos))]
        val_col = "(百万)人民币" if "(百万)人民币" in df.columns else df.columns[-1]
        data_map = df_year.set_index(['公司', '字段名'])[val_col].to_dict()
        get_val = lambda co, field: data_map.get((co, field), None) # 🌟 找不到时返回 None 而不是 0，用来判定是否未披露

        rows_config = [("未采用保费分配法的保险合同", None, "header"), ("保险服务收入", "未采用保费分配法计量的保险合同保险服务收入", "data"), ("合同服务边际的释放", "合同服务边际的摊销", "data"), ("非金融风险调整的变动", "非金融风险调整的变动", "data"), ("预期当期发生的保险服务费用", "预计当期发生的保险服务费用", "data"), ("保险获取现金流的摊销", "保险获取现金流的摊销（保险服务收入）", "data"), ("其他", "其他收入调整", "data"), ("保险服务费用", "未采用保费分配法计量的保险合同保险服务费用", "neg_data"), ("保险获取现金流的摊销 ", "保险获取现金流的摊销（保险服务费用）", "neg_data"), ("亏损部分的确认及转回", "亏损部分的确认及转回", "neg_data"), ("当期发生的赔款及其他相关费用", "当期发生的赔款及其他相关费用", "neg_data"), ("已发生赔款负债相关的履约现金流量变动", "已发生赔款负债相关的履约现金流量变动", "neg_data"), ("其他项", "FIXED_ZERO", "data"), ("保险服务业绩", "FORMULA_KPI_1", "subtotal"), ("采用保费分配法的保险合同", None, "header"), ("保险服务收入 ", "采用保费分配法计量的保险合同保险服务收入", "data"), ("保险服务费用 ", "采用保费分配法计量的保险合同保险服务费用", "neg_data"), ("保险服务业绩 ", "采用保费分配法计量的保险合同保险业绩", "subtotal"), ("集团/业务条线-汇总", None, "header"), ("保险服务收入  ", "保险服务收入合计", "data"), ("保险服务费用  ", "保险服务费用合计", "neg_data"), ("保险服务业绩（不含再保收支净额）", "FORMULA_TOTAL_KPI", "total"), ("CSM释放在保险服务业绩中占比", "FORMULA_CSM_RATIO", "percent")]
        current_hl = str(highlight_co).strip()
        html = "<table style='width:96%; border-collapse: collapse; font-family: sans-serif; margin-bottom: 10px; font-size: 10.5px;'><thead>"
        html += f"<tr style='background-color: #00338D; color: white; font-size: 11px; text-align: center; font-weight: bold;'><th style='padding: 3px 2px; text-align: left; border: 1.5px solid white;'>项目名称 (单位: {unit_str}人民币)</th>"
        for co in cos:
            is_hl = (str(co).strip() == current_hl)
            html += f"<th style='padding: 3px 2px; text-align: center; background-color: {'#001A4D' if is_hl else '#00338D'}; border-top: 1px solid {'#00338D' if is_hl else 'white'}; border-left: 1px solid {'#00338D' if is_hl else 'white'}; border-right: 1px solid {'#00338D' if is_hl else 'white'}; border-bottom: none;'>{co}</th>"
        html += "</tr></thead><tbody>"

        total_rows = len(rows_config)
        for row_idx, (row_name, field, r_type) in enumerate(rows_config):
            is_header, is_total = (r_type == "header"), (r_type in ["subtotal", "total"])
            row_bg = "#E8EDF4" if is_header else ("white" if row_idx % 2 == 0 else "#F8F9FA")
            n_weight, n_indent = ("bold" if (is_header or is_total) else "normal"), ("0px" if is_header else "10px")
            html += f"<tr><td style='padding: 2px 3px; padding-left: {n_indent}; text-align: left; font-weight: {n_weight}; background-color: {row_bg}; border: 1px solid #EAEAEA; color: #333333;'>{row_name}</td>"
            
            for co in cos:
                is_hl = (str(co).strip() == current_hl)
                val_str, is_missing = "", False
                if not is_header:
                    if r_type == "data": val = (get_val(co, field) / div) if field != "FIXED_ZERO" and get_val(co, field) is not None else (0 if field == "FIXED_ZERO" else None)
                    elif r_type == "neg_data": val = -(get_val(co, field) / div) if get_val(co, field) is not None else None
                    elif r_type in ["subtotal", "total", "percent"]:
                        v1, v2 = get_val(co, "未采用保费分配法计量的保险合同保险服务收入"), get_val(co, "未采用保费分配法计量的保险合同保险服务费用")
                        if r_type == "subtotal": val = (v1/div - v2/div) if v1 is not None and v2 is not None else None
                        elif r_type == "total": 
                            t1, t2 = get_val(co, "保险服务收入合计"), get_val(co, "保险服务费用合计")
                            val = (t1/div - t2/div) if t1 is not None and t2 is not None else None
                        elif r_type == "percent":
                            csm_rel = get_val(co, "合同服务边际的摊销")
                            denom = (v1 - v2) if v1 is not None and v2 is not None else None
                            val = (csm_rel / denom) if csm_rel is not None and denom is not None and denom != 0 else None
                    
                    if pd.isna(val) or val is None: is_missing = True; val_str = "未披露"
                    elif r_type == "percent": val_str = f"{val*100:.0f}%"
                    elif val < 0: val_str = f"({abs(val):,.0f})"
                    elif val == 0: val_str = "0"
                    else: val_str = f"{val:,.0f}"

                c_bg = "#CDCDCD" if is_missing else (HL_BOX_FILL   if is_hl else row_bg)
                c_color = "white" if is_missing else ("#00338D" if is_hl else ("#333333" if n_weight == "bold" else "#444444"))
                c_weight = "bold" if (is_hl or is_total) and not is_missing else "normal"
                borders = f"border-left: 2.5px solid #00338D; border-right: 1.5px solid #00338D; border-top: none; border-bottom: {'1.5px solid #00338D' if row_idx == total_rows - 1 else 'none'};" if is_hl else "border: 1px solid #EAEAEA;"
                
                html += f"<td style='padding: 2px 3px; text-align: center; background-color: {c_bg}; {borders} color: {c_color}; font-weight: {c_weight};'>{val_str}</td>"
            html += "</tr>"
            
        return html + "</tbody></table>"
    
# --- 30.新业务价值及增长分析表 (纯HTML生成不变) ---
    def create_nbv_summary_table(df, cos, div, unit_str, target_year, target_prev_year, highlight_co="无"):
        cy, py = target_year, target_prev_year
        cy_str, py_str = str(cy)[-2:], str(py)[-2:]
        
        # 🌟 修复 1：动态构建分析指标列表，如果是前一年的表，就不加增长率这一行
        metrics = ["新业务盈利合同(CSM)"]
        if str(cy) != str(prev_year): metrics.append(f"新业务CSM增长率 ({cy_str}YE/{py_str}YE-1)")
        metrics.extend(["新业务亏损合同(LC)", "新业务未来现金流入现值", "新业务IFRS利润率", "新业务增长 (新业务CSM/CSM摊销)"])
        
        col_data = {}
        for co in cos:
            df_co = df[df['公司'] == co]
            get_val = lambda y, kw: df_co[(df_co['报告年份'].astype(str) == str(y)) & (df_co['字段名'] == kw)].drop_duplicates(subset=['公司', '报告年份', '字段名'])['(百万)人民币'].sum() if not df_co[(df_co['报告年份'].astype(str) == str(y)) & (df_co['字段名'] == kw)].drop_duplicates(subset=['公司', '报告年份', '字段名'])['(百万)人民币'].empty else 0
            nb_csm_cy, nb_csm_py, lc_cy, pv_in_cy, amort_cy = get_val(cy, '新业务CSM（集团口径）'), get_val(py, '新业务CSM（集团口径）'), get_val(cy, '新业务亏损合同（LC）——非PAA'), get_val(cy, '新业务未来现金流入现值'), get_val(cy, 'CSM摊销')
            
            # 🌟 修复 2：将缺失值统一标记为 "未披露"
            fmt_num = lambda x: "未披露" if pd.isna(x) or x == 0 else f"{x / div:.1f}"
            fmt_pct = lambda x: "未披露" if pd.isna(x) or x == 0 else f"{x:.0%}"
            
            v1, v3, v4 = fmt_num(nb_csm_cy), fmt_num(lc_cy), fmt_num(pv_in_cy)
            v2 = "未披露" if nb_csm_py == 0 or nb_csm_cy == 0 else fmt_pct((nb_csm_cy / nb_csm_py) - 1)
            v5 = "未披露" if pv_in_cy == 0 else fmt_pct((nb_csm_cy - lc_cy) / pv_in_cy)
            v6 = fmt_pct(nb_csm_cy / (-amort_cy)) if amort_cy != 0 else "未披露"
            
            col_data[co] = {
                "新业务盈利合同(CSM)": v1, f"新业务CSM增长率 ({cy_str}YE/{py_str}YE-1)": v2, 
                "新业务亏损合同(LC)": v3, "新业务未来现金流入现值": v4, 
                "新业务IFRS利润率": v5, "新业务增长 (新业务CSM/CSM摊销)": v6
            }
            
        current_hl = str(highlight_co).strip()
        html = "<table style='width:100%; border-collapse: collapse; font-family: sans-serif; font-size: 12px; margin-bottom: 15px;'><tr style='background-color: #00338D; color: white; text-align: center; font-weight: bold;'><th style='padding: 6px 4px; border: 1px solid #EAEAEA; width: 22%; text-align: left;'>分析指标</th>"
        for co in cos:
            is_hl = (co == current_hl)
            html += f"<th style='padding: 6px 4px; background-color: {'#001A4D' if is_hl else '#00338D'}; border: 1px solid #EAEAEA;'>{co}</th>"
        html += "</tr>"
        
        for i, metric in enumerate(metrics):
            html += f"<tr style='background-color: {'white' if i % 2 == 0 else '#F9F9F9'};'><td style='padding: 6px 4px; text-align: left; font-weight: bold; color: #333333; border: 1px solid #EAEAEA;'>{metric}</td>"
            for co in cos:
                val, is_hl = col_data[co].get(metric, "未披露"), (co == current_hl)
                # 🌟 修复 3：HTML 渲染时，如果拿到 "未披露"，直接置灰并写白色字
                if val == "未披露":
                    html += "<td style='padding: 6px 4px; text-align: center; background-color: #CDCDCD; color: white; border: 1px solid #EAEAEA;'>未披露</td>"
                else:
                    html += f"<td style='padding: 6px 4px; text-align: center; background-color: {'rgba(0, 51, 141, 0.08)' if is_hl else 'transparent'}; font-weight: {'bold' if is_hl else 'normal'}; color: {'#00338D' if is_hl else '#444444'}; border: 1px solid #EAEAEA;'>{val}</td>"
            html += "</tr>"
        return html + "</table>"


    # ==========================================
    # 🎛️ 终极模块路由引擎：绑定专属 UI 控件并出图
    # ==========================================
    def render_pure_chart_entity(m_id, print_mode):
        current_hl = highlight_co if highlight_co else "无"
        
        # 1. 关键数据概览表
        if m_id == "summary_table":
            html_table = create_financial_summary_table(df_filtered, selected_cos, current_hl)
            if html_table: st.markdown(html_table, unsafe_allow_html=True)
            
        # 2. 利润贡献与保险和投资利润占比
        elif m_id == "prof_mix":
            if not print_mode:
                c1, c2 = st.columns([1, 2])
                with c1: lab = st.toggle("显示数据标签", value=True, key=f"lab_{m_id}")
                with c2: gap = st.slider("柱子间距", 0.2, 0.7, 0.4, key=f"gap_{m_id}")
            else:
                lab, gap = True,0.4 
            fig = create_profit_mixed_chart(df_filtered, selected_cos, "", lab, gap, divisor, unit_label, current_hl)
            show_chart(fig, print_mode,m_id=m_id)
            
        # 3. 单一指标趋势 (保险收入/费用/业绩)
        elif m_id in ["inc_total", "exp_total", "perf_total"]:
            field_map = {"inc_total": "保险服务收入合计", "exp_total": "保险服务费用合计", "perf_total": "保险服务业绩"}
            if not print_mode:
                c1, c2, c3 = st.columns([1.5, 2, 3])
                with c1: lab = st.toggle("显示柱状图数据标签", value=True, key=f"lab_{m_id}")
                with c2: psz = st.slider("涨幅文字大小", 8, 24, 12, key=f"psz_{m_id}")
                with c3: gap = st.slider("调整公司间距与柱子粗细", 0.1, 0.8, 0.3, key=f"gap_{m_id}")
            else:
                lab, psz, gap=True, 13, 0.3
            fig = create_kpmg_chart(df_filtered, field_map[m_id], "", lab, psz, gap)
            show_chart(fig, print_mode,m_id=m_id)
            
        # 4. 收入结构分析 1/2
        elif m_id == "comp_1":
            if not print_mode:
                c1, c2, c3, c4 = st.columns(4)
                with c1: lab = st.toggle("显示数据标签", value=True, key=f"lab_{m_id}")
                with c2: lsz = st.slider("标签大小", 5, 20, 12, key=f"lsz_{m_id}")
                with c3: wid = st.slider("柱子宽度", 0.1, 1.0, 0.6, 0.05, key=f"wid_{m_id}")
                with c4: cfs = st.slider("公司名称大小", 10, 20, 12, key=f"cfs_{m_id}")
            else:
                # 使用 .get() 从缓存中读取用户在网页上调整过的值，如果没有就用默认值
                lab = st.session_state.get(f"lab_{m_id}", True)
                lsz = st.session_state.get(f"lsz_{m_id}", 12)
                wid = st.session_state.get(f"wid_{m_id}", 0.6)
                cfs = st.session_state.get(f"cfs_{m_id}", 12)
            comp_fields = ["采用保费分配法计量的保险合同保险服务收入", "未采用保费分配法计量的保险合同保险服务收入"]
            fig = create_kpmg_composition_chart(df_filtered, comp_fields, "", lab, lsz, wid, cfs, current_hl)
            show_chart(fig, print_mode,m_id=m_id)
            
# 5. 收入结构分析 2/2 (多组成带表格)
        elif m_id == "comp_2":
            if not print_mode:
                c1, c2, c3, c4 = st.columns(4)
                with c1: lab = st.toggle("显示数据标签", value=True, key=f"lab_{m_id}")
                with c2: lsz = st.slider("标签大小", 5, 20, 10, key=f"lsz_{m_id}") # 🌟 默认值调到10
                with c3: wid = st.slider("柱子宽度", 0.1, 1.0, 0.6, 0.05, key=f"wid_{m_id}")
                with c4: cfs = st.slider("公司名称大小", 10, 20, 11, key=f"cfs_{m_id}") # 🌟 默认值调到11
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                lsz = st.session_state.get(f"lsz_{m_id}", 10)
                wid = st.session_state.get(f"wid_{m_id}", 0.6)
                cfs = st.session_state.get(f"cfs_{m_id}", 11)
                
            f_m2 = {"合同服务边际的摊销":"合同服务边际的释放", "非金融风险调整的变动":"非金融风险调整的变动", "预计当期发生的保险服务费用":"预期当期发生的保险服务费用", "保险获取现金流的摊销（保险服务收入）":"保险获取现金流的摊销", "与当期服务或过去服务相关得保费经验调整":"与当期服务或过去服务相关的保费经验调整", "其他收入调整":"其他"}
            c_m2 = {"合同服务边际的摊销":"rgb(30, 73, 226)", "非金融风险调整的变动":"rgb(254, 174, 215)", "预计当期发生的保险服务费用":"rgb(0, 163, 161)", "保险获取现金流的摊销（保险服务收入）":"rgb(1, 184, 245)", "与当期服务或过去服务相关得保费经验调整":"rgb(0, 219, 214)", "其他收入调整":"rgb(114, 19, 234)"}
            
            fig, df_avg = create_kpmg_multi_composition_chart(df_filtered, f_m2, c_m2, "", lab, lsz, wid, cfs, current_hl)
            
            # 🌟 修复：极致压缩表格体积，防止 PDF 换页
            if fig and not df_avg.empty:
                st.markdown("<div style='margin-bottom: 2px;'><div style='font-size: 12px; font-weight: bold; margin-bottom: 4px; color:#333;'>各公司平均占比情况 </div>", unsafe_allow_html=True)
                html = "<table style='width:100%; border-collapse: collapse; font-family: sans-serif; font-size: 10px; margin-bottom: 4px;'><tr style='background-color: #00338D; color: white; text-align: center; font-weight: bold;'><th style='padding: 2px 4px; border: 1px solid white;'>报告年份</th>"
                for orig_k, display_name in f_m2.items(): 
                    html += f"<th style='padding: 2px 4px; background-color: {c_m2[orig_k]}; color: white; border: 1px solid white;'>{display_name}</th>"
                html += "</tr>"
                for yr, row in df_avg.iterrows():
                    html += f"<tr><td style='padding: 2px 4px; font-weight: bold; background-color: #F8F9FA; border: 1px solid #EAEAEA;'>{yr}</td>"
                    for orig_k, display_name in f_m2.items(): 
                        html += f"<td style='padding: 2px 4px; text-align: center; background-color: white; border: 1px solid #EAEAEA;'>{row[display_name]:.1f}%</td>"
                    html += "</tr>"
                html += "</table></div>"
                st.markdown(html, unsafe_allow_html=True)
                
            show_chart(fig, print_mode,m_id=m_id)


# 6.新加费用的图
        elif m_id == "exp_1":
            if not print_mode:
                c1, c2, c3, c4 = st.columns(4)
                with c1: lab = st.toggle("显示数据标签", value=True, key=f"lab_{m_id}")
                with c2: lsz = st.slider("标签大小", 5, 20, 10, key=f"lsz_{m_id}") # 默认值调小点
                with c3: wid = st.slider("柱子宽度", 0.1, 1.0, 0.6, 0.05, key=f"wid_{m_id}")
                with c4: cfs = st.slider("公司名称大小", 10, 20, 11, key=f"cfs_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                lsz = st.session_state.get(f"lsz_{m_id}", 10)
                wid = st.session_state.get(f"wid_{m_id}", 0.6)
                cfs = st.session_state.get(f"cfs_{m_id}", 11)
        
            f_m2 = {
                "保险获取现金流的摊销（保险服务费用）": "保险获取现金流的摊销",
                "亏损部分的确认及转回": "亏损部分的确认及转回",
                "当期发生的赔款及其他相关费用": "当期发生的赔款及费用",
                "已发生赔款负债相关的履约现金流量变动": "已发生赔款负债变动",
            }
            c_m2 = {
                "保险获取现金流的摊销（保险服务费用）": "rgb(30, 73, 226)",
                "亏损部分的确认及转回":               "rgb(254, 174, 215)",
                "当期发生的赔款及其他相关费用":        "rgb(0, 163, 161)",
                "已发生赔款负债相关的履约现金流量变动": "rgb(1, 184, 245)",
            }
        
            fig, df_avg = create_kpmg_exp_chart(df_filtered, f_m2, c_m2, "", lab, lsz, wid, cfs, current_hl)
            
            # 🌟 修复：将表格包裹在一个 div 中，防止 margin 溢出导致打印换页。同时极大地压缩表格字体和 padding
            if fig and not df_avg.empty:
                st.markdown("<div style='margin-bottom: 2px;'><div style='font-size:12px; font-weight:bold; margin-bottom:4px; color:#333;'>各公司平均占比情况 </div>", unsafe_allow_html=True)
                html = "<table style='width:100%; border-collapse:collapse; font-family:sans-serif; font-size:10px; margin-bottom:4px;'><tr style='background-color:#00338D; color:white; text-align:center; font-weight:bold;'><th style='padding:2px 4px; border:1px solid white;'>报告年份</th>"
                for orig_k, display_name in f_m2.items():
                    html += f"<th style='padding:2px 4px; background-color:{c_m2[orig_k]}; color:white; border:1px solid white;'>{display_name}</th>"
                html += "</tr>"
                for yr, row in df_avg.iterrows():
                    html += f"<tr><td style='padding:2px 4px; font-weight:bold; background-color:#F8F9FA; border:1px solid #EAEAEA;'>{yr}</td>"
                    for orig_k, display_name in f_m2.items():
                        html += f"<td style='padding:2px 4px; text-align:center; background-color:white; border:1px solid #EAEAEA;'>{row[display_name]:.1f}%</td>"
                    html += "</tr>"
                html += "</table></div>"
                st.markdown(html, unsafe_allow_html=True)
                
            show_chart(fig, print_mode, m_id=m_id)
            
            
        # 6. 利润构成拆解
        elif m_id in ["prof_2025", "prof_2024"]:
            y_val = latest_year if m_id == "prof_2025" else prev_year
            if not print_mode:
                c1, c2, c3, c4 = st.columns(4)
                with c1: lab = st.toggle("显示利润标签", value=True, key=f"lab_{m_id}")
                with c2: psz = st.slider("标签字号", 8, 16, 11, key=f"psz_{m_id}")
                with c3: wid = st.slider("柱宽", 0.2, 0.8, 0.4, key=f"wid_{m_id}")
                with c4: cfs = st.slider("公司字号", 10, 20, 11, key=f"cfs_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                psz = st.session_state.get(f"psz_{m_id}", 11)
                wid = st.session_state.get(f"wid_{m_id}", 0.5)
                cfs = st.session_state.get(f"cfs_{m_id}", 11)
            fig, df_p = create_profit_composition_chart(df_filtered, selected_cos, y_val, lab, psz, wid, cfs, current_hl)
            show_chart(fig, print_mode,m_id=m_id)
            
        # 7. 投资相关简单柱图
        elif m_id in ["inv_return", "uw_profit", "inv_profit"]:
            field_map = {"inv_return": "净投资回报", "uw_profit": "承保财务净损益", "inv_profit": "投资利润"}
            if not print_mode:
                c1, c2, c3 = st.columns([2, 2, 3])
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: psz = st.slider("文字大小", 8, 20, 14, key=f"psz_{m_id}")
                with c3: gap = st.slider("柱子间距", 0.1, 0.8, 0.3, key=f"gap_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                psz = st.session_state.get(f"psz_{m_id}", 13)
                gap = st.session_state.get(f"gap_{m_id}", 0.3)
            fig = create_simple_kpmg_chart(df_plot_final, field_map[m_id], "", lab, psz, gap, current_hl)
            show_chart(fig, print_mode,m_id=m_id)
            
        # 8. 税前利润
        elif m_id == "tax_profit":
            if not print_mode:
                c1, c2, c3, c4, c5 = st.columns(5)
                with c1: lab = st.toggle("显示标签", value=True, key=f"lab_{m_id}")
                with c2: wid = st.slider("柱宽", 0.1, 0.8, 0.4, key=f"wid_{m_id}")
                with c3: f_sz = st.slider("字号", 8, 16, 10, key=f"fs_{m_id}")
                with c4: cfs = st.slider("公司名字号", 10, 24, 13, key=f"cfs_{m_id}")
                with c5: hoy = st.slider("高度比例", 1.0, 1.2, 1.02, step=0.01, key=f"hoy_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                wid = st.session_state.get(f"wid_{m_id}", 0.4)
                f_sz = st.session_state.get(f"fs_{m_id}", 10)
                cfs = st.session_state.get(f"cfs_{m_id}", 13)
                hoy = st.session_state.get(f"hoy_{m_id}", 1.02)
            fig = create_tax_subplot_chart(df_tax_pivot, selected_cos, lab, wid, f_sz, cfs, hoy, current_hl)
            show_chart(fig, print_mode,m_id=m_id)
            
        # 9. 综合收益变动趋势
        elif m_id in ["net_profit", "oci_profit", "total_profit"]:
            field_map = {"net_profit": "净利润", "oci_profit": "其他综合收益", "total_profit": "综合收益总额"}
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: psz = st.slider("文字大小", 8, 20, 12, key=f"psz_{m_id}")
                with c3: gap = st.slider("柱宽", 0.1, 0.8, 0.3, key=f"gap_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                psz = st.session_state.get(f"psz_{m_id}", 12)
                gap = st.session_state.get(f"gap_{m_id}", 0.3)
            fig = create_financial_trend_chart_v5(df_fin_raw, field_map[m_id], "", lab, psz, gap, current_hl)
            show_chart(fig, print_mode,m_id=m_id)
            
        # 10. 资产端分类
        elif m_id == "asset_struct":
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: wid = st.slider("柱宽", 0.1, 1.0, 0.6, key=f"wid_{m_id}")
                with c3: sz = st.slider("字号", 8, 20, 12, key=f"sz_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                wid = st.session_state.get(f"wid_{m_id}", 0.6)
                sz = st.session_state.get(f"sz_{m_id}", 12)
            f_map = {"AC": "债权投资", "FVOCI": "其他债权投资", "FVTPL": "交易性金融资产", "指定FVOCI": "其他权益工具投资"}
            c_map = {"AC": "rgb(0, 184, 245)", "FVOCI": "rgb(114, 19, 234)", "FVTPL": "rgb(253, 52, 156)", "指定FVOCI": "rgb(181, 2, 95)"}
            fig = create_asset_composition_chart(df_filtered, selected_cos, f_map, c_map, "", lab, sz, wid, 12, current_hl)
            show_chart(fig, print_mode,m_id=m_id)
            
        # 11. 两年的 OCI 变动分析
        elif m_id in ["oci_year_lat", "oci_year_pre"]:
            y = latest_year if m_id == "oci_year_lat" else prev_year
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: gap = st.slider("间距", 0.05, 0.5, 0.15, key=f"gap_{m_id}")
                with c3: sz = st.slider("公司字号", 8, 20, 12, key=f"sz_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                gap = st.session_state.get(f"gap_{m_id}", 0.15)
                sz = st.session_state.get(f"sz_{m_id}", 12)
            fig = create_oci_chart(df_filtered, y, "", lab, sz, gap, selected_cos, current_hl)
            show_chart(fig, print_mode,m_id=m_id)
            
        # 12. 资负 OCI
        elif m_id == "oci_deep":
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: gap = st.slider("间距", 0.05, 0.5, 0.15, key=f"gap_{m_id}")
                with c3: sz = st.slider("公司字号", 8, 20, 12, key=f"sz_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                gap = st.session_state.get(f"gap_{m_id}", 0.15)
                sz = st.session_state.get(f"sz_{m_id}", 12)
                
            fig = create_asset_liab_oci_chart(df_filtered, selected_cos, gap, sz, lab, current_hl)
            show_chart(fig, print_mode, m_id=m_id)
            
            # 🌟 直接接收并渲染已经在函数内排版好的精美 HTML 表格，下面一堆旧的拼表代码全删了！
            html_table, c_y, p_y = calculate_oci_analysis_table(df_filtered, selected_cos, current_hl)
            if html_table:
                st.markdown(f"<p style='font-size:13px; font-weight:bold; color:#00338D; margin-top:-30px; margin-bottom:6px;'>资负 OCI 变动分析表 ({p_y}YE - {c_y}YE)</p>", unsafe_allow_html=True)
                st.markdown(html_table, unsafe_allow_html=True)

# 13. 净资产变动与总资产变动
        elif m_id in ["equity_trend", "asset_trend"]:
            field_map = {"equity_trend": "期末股东权益", "asset_trend": "总资产"}
            df_asset_raw = df_filtered[df_filtered['字段名'].isin(['期末股东权益', '总资产']) & df_filtered['公司'].isin(selected_cos)].drop_duplicates(subset=['公司', '报告年份', '字段名']).copy()
            df_asset_raw.rename(columns={'字段名': '指标名称', '(百万)人民币': 'value'}, inplace=True)
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: sz = st.slider("字号", 8, 24, 11, key=f"sz_{m_id}")
                with c3: gap = st.slider("间距", 0.1, 0.8, 0.3, key=f"gap_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                gap = st.session_state.get(f"gap_{m_id}", 0.3)
                sz = st.session_state.get(f"sz_{m_id}", 12)
            fig = create_asset_trend_chart(df_asset_raw, field_map[m_id], "", lab, sz, gap, current_hl)
            show_chart(fig, print_mode,m_id=m_id)
            
        # 14. CSM余额变动趋势
        elif m_id == "csm_bal":
            df_csm_raw = df_filtered[(df_filtered['字段名'] == 'CSM期末余额') & df_filtered['公司'].isin(selected_cos)].drop_duplicates(subset=['公司', '报告年份', '字段名']).copy()
            df_csm_raw.rename(columns={'字段名': '指标名称', '(百万)人民币': 'value'}, inplace=True)    
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: sz = st.slider("字号", 8, 24, 12, key=f"sz_{m_id}")
                with c3: gap = st.slider("柱宽", 0.1, 0.8, 0.35, key=f"gap_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                gap = st.session_state.get(f"gap_{m_id}", 0.35)
                sz = st.session_state.get(f"sz_{m_id}", 12)
            fig = create_csm_trend_chart(df_csm_raw, 'CSM期末余额', "", lab, sz, gap, current_hl)
            show_chart(fig, print_mode,m_id=m_id)

        # 15. CSM期初余额占比分析
        elif m_id == "csm_ratio":
            if not print_mode:
                c1, c2 = st.columns(2)
                with c1: lab = st.toggle("显示占比百分比标签", True, key=f"lab_{m_id}") 
                with c2: gap = st.slider("柱距调节 (越小柱子越粗)", 0.1, 0.8, 0.3, key=f"gap_{m_id}")   
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                gap = st.session_state.get(f"gap_{m_id}", 0.3)
            fig = create_csm_ratio_chart(df_filtered, selected_cos, lab, gap, current_hl)
            show_chart(fig, print_mode,m_id=m_id)

# 16. CSM 概览明细表
        elif m_id == "csm_table":
            # 🌟 核心修改：现在函数直接返回排版好的 HTML 文本
            html_table = show_csm_summary_table(df_filtered, latest_year, selected_cos, current_hl)
            
            # 如果成功生成了 HTML，就直接用 st.markdown 渲染它，原先那堆拼接逻辑全删了！
            if html_table:
                st.markdown(html_table, unsafe_allow_html=True)

        # 17. CSM过渡期拆分
        elif m_id == "csm_trans":
            if not print_mode:
                c1, c2 = st.columns(2)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: wid = st.slider("柱宽", 0.1, 0.8, 0.4, key=f"wid_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                wid = st.session_state.get(f"wid_{m_id}", 0.4)
            fig = create_csm_transition_chart(df_filtered, selected_cos, lab, wid, current_hl)
            show_chart(fig, print_mode,m_id=m_id)

        # 18. 两年的摊销前CSM占比
        elif m_id in ["csm_comp_lat", "csm_comp_pre"]:
            y = latest_year if m_id == "csm_comp_lat" else prev_year
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: wid = st.slider("柱宽", 0.1, 1.0, 0.5, key=f"wid_{m_id}")
                with c3: sz = st.slider("字号", 8, 20, 11, key=f"sz_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                wid = st.session_state.get(f"wid_{m_id}", 0.5)
                sz = st.session_state.get(f"sz_{m_id}", 11)
            fig = create_csm_composition_chart(df_filtered, selected_cos, y, lab, sz, wid, current_hl, title_text="")
            show_chart(fig, print_mode,m_id=m_id)

        # 19. CSM比率折线图 (摊销比率 + 持续率双图并行)
        elif m_id == "csm_ratio_trend":
            if not print_mode:
                c1, c2 = st.columns(2)
                with c1: lab = st.toggle("显示数值", True, key=f"lab_{m_id}")
                with c2: sz = st.slider("点大小", 4, 15, 6, key=f"sz_{m_id}")
            else:
                lab, sz = st.session_state.get(f"lab_{m_id}", True), st.session_state.get(f"sz_{m_id}", 6)
        
            color_map = get_color_map(selected_cos)
            df_c_sub = df_filtered[df_filtered['字段名'].isin(['CSM摊销', 'CSM期末余额', '新业务CSM（集团口径）'])].pivot_table(
                index=['公司', '报告年份'], columns='字段名', values='(百万)人民币').reset_index().fillna(0)
            denom1 = (df_c_sub['CSM期末余额'] - df_c_sub['CSM摊销']).replace(0, np.nan)
            denom2 = df_c_sub['CSM摊销'].replace(0, np.nan)
            
            df_c_sub['摊销比率'] = -df_c_sub['CSM摊销'] / denom1
            df_c_sub['持续率'] = -df_c_sub['新业务CSM（集团口径）'] / denom2
            
            df_c_sub.replace([np.inf, -np.inf], np.nan, inplace=True)
            df_c_sub.replace([np.inf, -np.inf], np.nan, inplace=True)
            df_c_sub['报告年份'] = df_c_sub['报告年份'].astype(str).str.replace(".0", "", regex=False) + "YE"
        
# ✅ 自定义圆点图例（HTML），两个子图共用：自动侦测缺失/为0的数据并标注(未披露)
            legend_items = []
            for co in selected_cos:
                # 1. 抓取该公司的全部数据
                d_co = df_c_sub[df_c_sub['公司'] == co]
                
                # 2. 判断是否未披露（没查到数据，或者摊销/持续率全为 0 或 NaN）
                is_missing = False
                if d_co.empty:
                    is_missing = True
                else:
                    v_tanxiao = d_co['摊销比率'].fillna(0)
                    v_chixu = d_co['持续率'].fillna(0)
                    if (v_tanxiao == 0).all() and (v_chixu == 0).all():
                        is_missing = True
                
                # 3. 如果未披露，加上后缀，顺便把字变成浅灰色，视觉上更直观
                display_name = f"{co} (未披露)" if is_missing else co
                text_color = "#999999" if is_missing else "#333333"
                bg_color = HIGHLIGHT_COLOR if co == current_hl else color_map.get(co, "#333")
                
                # 4. 组装单个图例项
                legend_items.append(
                    f'<div style="display:flex;align-items:center;gap:4px;">'
                    f'<div style="width:9px;height:9px;border-radius:50%;background:{bg_color};flex-shrink:0;"></div>'
                    f'<span style="color:{text_color};">{display_name}</span></div>'
                )

            st.markdown(
                f'<div style="display:flex;flex-wrap:wrap;gap:12px;margin-bottom:8px;font-size:11px;font-family:Microsoft YaHei,sans-serif;">{"".join(legend_items)}</div>',
                unsafe_allow_html=True
            )
        
            # ✅ 合并成一个 subplot，打印网页都稳定
            from plotly.subplots import make_subplots
            fig = make_subplots(rows=1, cols=2,
                                subplot_titles=["CSM摊销比率趋势", "CSM持续率趋势"],
                                horizontal_spacing=0.08)
        
            for col_idx, metric in enumerate(['摊销比率', '持续率'], 1):
                df_m = df_c_sub[['公司', '报告年份', metric]].rename(columns={metric: 'value'})
                latest_yr = df_m['报告年份'].max()
                # ✅ 改成：只从有效（非0非NaN）的数据里找最大最小
                df_latest = df_m[df_m['报告年份'] == latest_yr].dropna(subset=['value'])
                df_latest_valid = df_latest[df_latest['value'] != 0]  # 👈 排除0值
                max_co = df_latest_valid.loc[df_latest_valid['value'].idxmax(), '公司'] if not df_latest_valid.empty else None
                min_co = df_latest_valid.loc[df_latest_valid['value'].idxmin(), '公司'] if not df_latest_valid.empty else None
        
                for co in selected_cos:
                    d_co = df_m[df_m['公司'] == co].sort_values('报告年份')
                    
                    # 🌟 核心拦截逻辑：检查数据是否真实有效
                    if d_co.empty: 
                        continue
                    
                    # 取出所有的非空数值
                    valid_values = d_co['value'].dropna()
                    
                    # 如果全是 NaN，或者所有算出来的值都是 0，就不画这条线
                    if valid_values.empty or (valid_values == 0).all():
                        continue
                    is_hl  = (co == current_hl)
                    is_ext = (co in [max_co, min_co])
                    fig.add_trace(go.Scatter(
                        x=d_co['报告年份'], y=d_co['value'], name=co,
                        mode='lines+markers+text' if lab else 'lines+markers',
                        line=dict(
                            color=HIGHLIGHT_COLOR if is_hl else color_map.get(co, "#333"),
                            width=4 if is_hl else (2.5 if is_ext else 2),
                            dash="solid" if (is_hl or is_ext) else "dot"
                        ),
                        marker=dict(size=sz * 1.5 if is_hl else sz, symbol='circle'),
                        text=[f"{v*100:.1f}%" for v in d_co['value']] if lab else None,
                        textposition="top center", cliponaxis=False,
                        showlegend=False  # ✅ 关掉Plotly自带图例，用上面HTML圆点图例
                    ), row=1, col=col_idx)
        
                fig.update_yaxes(tickformat=".1%", showgrid=False, zeroline=False, row=1, col=col_idx)
                fig.update_xaxes(showgrid=False, showline=False, zeroline=False, row=1, col=col_idx)
        
            fig.update_layout(
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                showlegend=False,
                height=380 if print_mode else 430,
                margin=dict(t=40, b=20, l=30, r=30),
            )
        
            show_chart(fig, print_mode, m_id=m_id)

        # 20. 综合净资产指标
        elif m_id == "csm_equity":
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: wid = st.slider("柱宽", 0.1, 1.0, 0.5, key=f"wid_{m_id}")
                with c3: sz = st.slider("字号", 8, 20, 12, key=f"sz_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                wid = st.session_state.get(f"wid_{m_id}", 0.5)
                sz = st.session_state.get(f"sz_{m_id}", 12)
            fig = create_csm_equity_analysis(df_filtered, selected_cos, lab, sz, wid, 12, 20, current_hl, title_text="")
            show_chart(fig, print_mode,m_id=m_id)

        # 21. 新业务盈利合同(CSM)
        elif m_id == "nb_csm":
            df_nb_csm_raw = df_filtered[(df_filtered['字段名'] == '新业务CSM（集团口径）') & df_filtered['公司'].isin(selected_cos)].drop_duplicates(subset=['公司', '报告年份', '字段名']).copy()
            df_nb_csm_raw.rename(columns={'字段名': '指标名称', '(百万)人民币': 'value'}, inplace=True)
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: sz = st.slider("字号", 8, 24, 12, key=f"sz_{m_id}")
                with c3: gap = st.slider("间距", 0.1, 0.8, 0.3, key=f"gap_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                gap = st.session_state.get(f"gap_{m_id}", 0.3)
                sz = st.session_state.get(f"sz_{m_id}", 12)
            fig = create_new_biz_csm_chart(df_nb_csm_raw, '新业务CSM（集团口径）', "", lab, sz, gap, current_hl)
            show_chart(fig, print_mode,m_id=m_id)


        # 22. 新业务亏损(CSM-非PAA)
        elif m_id == "nb_lost":
            df_nb_csm_raw = df_filtered[(df_filtered['字段名'] == '新业务亏损合同（LC）——非PAA') & df_filtered['公司'].isin(selected_cos)].drop_duplicates(subset=['公司', '报告年份', '字段名']).copy()
            df_nb_csm_raw.rename(columns={'字段名': '指标名称', '(百万)人民币': 'value'}, inplace=True)
            # ✅ 取负数
            df_nb_csm_raw['value'] = -df_nb_csm_raw['value']
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: sz = st.slider("字号", 8, 24, 12, key=f"sz_{m_id}")
                with c3: gap = st.slider("间距", 0.1, 0.8, 0.3, key=f"gap_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                gap = st.session_state.get(f"gap_{m_id}", 0.3)
                sz = st.session_state.get(f"sz_{m_id}", 12)
            fig = create_new_lost_csm_chart(df_nb_csm_raw, '新业务亏损合同（LC）——非PAA', "", lab, sz, gap, current_hl)
            show_chart(fig, print_mode, m_id=m_id)
            
        # 22. 新业务指标拆解三图
        elif m_id == "nb_struct":
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: wid = st.slider("柱宽", 0.2, 1.0, 0.6, key=f"wid_{m_id}")
                with c3: sz  = st.slider("字号", 8, 20, 12, key=f"sz_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                wid = st.session_state.get(f"wid_{m_id}", 0.6)
                sz  = st.session_state.get(f"sz_{m_id}", 11)
        
            f1, f2, f3 = create_new_business_metrics_charts(
                df_filtered, selected_cos, lab, sz, wid, 11, current_hl, is_print_mode=print_mode
            )
            if f1:
                show_chart(f1, print_mode, m_id=m_id)
                # 👇 第2、3个图前插续标题（只在打印模式）
                # if print_mode: render_continue_title(m_id)
                show_chart(f2, print_mode, m_id=m_id)
                show_chart(f3, print_mode, m_id=m_id)

        # 23. 新业务IFRS利润率趋势
        elif m_id == "nb_margin_trend":
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: mk = st.slider("节点大小", 4, 15, 6, key=f"mk_{m_id}")
                with c3: fs = st.slider("图例字号", 8, 16, 12, key=f"fs_{m_id}")
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                mk = st.session_state.get(f"mk_{m_id}", 6)
                fs = st.session_state.get(f"fs_{m_id}", 12)
                
            color_map = get_color_map(selected_cos)
            fig = create_nb_margin_trend_chart(df_filtered, selected_cos, "", color_map, lab, mk, fs, current_hl)  
            
            # 🌟 修复：去掉 col_l 和 col_r 的分列，直接渲染，它就会自动居中并撑满页面！
            if fig:
                col_left, col_center, col_right = st.columns([1, 8, 1]) 
                with col_center:
                    show_chart(fig, print_mode, m_id=m_id)

        # 24. 费用结构及HTML统计表
        elif m_id == "exp_struct":
            if not print_mode:
                c1, c2, c3 = st.columns(3)
                with c1: lab = st.toggle("显示标签", True, key=f"lab_{m_id}")
                with c2: wid = st.slider("柱宽", 0.2, 0.8, 0.6, key=f"wid_{m_id}")
                with c3: sz = st.slider("字号", 8, 20, 10, key=f"sz_{m_id}")  
            else:
                lab = st.session_state.get(f"lab_{m_id}", True)
                wid = st.session_state.get(f"wid_{m_id}", 0.6)
                sz = st.session_state.get(f"sz_{m_id}", 11)               
            fig, stats = create_expense_breakdown_chart(df_filtered, selected_cos, lab, sz, wid, 12, current_hl)
            p_acq, p_maint, p_non = stats      
            st.markdown(f"""
            <table style='width:100%; border-collapse: collapse; font-family: sans-serif; margin-bottom: 15px; font-size: 11px;'>
                <tr style='background-color: #00338D; color: white; font-size: 12px; text-align: center; font-weight: bold;'>
                    <th style='padding: 6px 4px; text-align: left; border: 1px solid white;'>项目</th>
                    <th style='padding: 6px 4px; text-align: center; border: 1px solid white;'>获取费用均值</th>
                    <th style='padding: 6px 4px; text-align: center; border: 1px solid white;'>维持费用均值</th>
                    <th style='padding: 6px 4px; text-align: center; border: 1px solid white;'>非履约费用均值</th>
                </tr>
                <tr>
                    <td style='padding: 4px 4px; text-align: left; font-weight: bold; background-color: #F8F9FA; border: 1px solid #EAEAEA;'>各公司平均占比</td>
                    <td style='padding: 4px 4px; text-align: center; background-color: white; border: 1px solid #EAEAEA; color: rgb(30, 73, 226); font-weight: bold;'>{p_acq:.0f}%</td>
                    <td style='padding: 4px 4px; text-align: center; background-color: white; border: 1px solid #EAEAEA; color: rgb(118, 210, 255); font-weight: bold;'>{p_maint:.0f}%</td>
                    <td style='padding: 4px 4px; text-align: center; background-color: white; border: 1px solid #EAEAEA; color: rgb(114, 19, 234); font-weight: bold;'>{p_non:.0f}%</td>
                </tr>
            </table>
            """, unsafe_allow_html=True)        
            show_chart(fig, print_mode,m_id=m_id)

        # 25. 关键披露：折现率假设与非金融风险置信水平表与摊销比
        elif m_id == "discount_rate":
            html = create_discount_rate_table(df_filtered, latest_year, selected_cos, current_hl)
            if html: st.markdown(html, unsafe_allow_html=True)

        elif m_id == "confidence_level":
            html = create_confidence_level_table(df_filtered, latest_year, selected_cos, current_hl)
            if html: st.markdown(html, unsafe_allow_html=True)


        elif m_id == "csm_maturity_table":
            if not print_mode:
                c1, c2 = st.columns(2)
                with c1: lab = st.toggle("显示摊销累积比例", True, key=f"lab_{m_id}")
                with c2: mk  = st.slider("数据点大小", 4, 15, 8, key=f"mk_{m_id}")
            else:
                lab, mk = True, 8
        
            PRESET_COLORS = [
                "#C00000", "#0865EE", "#FEAED7", "#92D050", "#7030A0",
                "#EF9867", "#61CBF4", "#C7A0F7", "#FF4500", "#008B8B",
                "#FFD700", "#8B008B"
            ]
        
            # ✅ 每次都根据当前 selected_cos 重建 color_map，避免缓存旧状态导致撞色
            color_map_key = tuple(selected_cos)  # 用公司列表作为缓存 key
            if st.session_state.get('_color_map_key') != color_map_key:
                st.session_state['company_color_map'] = {
                    co: PRESET_COLORS[i % len(PRESET_COLORS)]
                    for i, co in enumerate(selected_cos)
                }
                st.session_state['_color_map_key'] = color_map_key
        
            color_map = st.session_state['company_color_map']
        
            fig = create_csm_maturity_table(
                df_filtered, latest_year, selected_cos,
                color_map, lab, mk, current_hl, title_text=""
            )
            if fig:
                show_chart(fig, print_mode,m_id=m_id)
                
# 26. 附录：业绩明细表与新业务明细表 (双表同打)
        elif m_id == "report_detail":  
            try: cy_int = int(latest_year)
            except: cy_int = 2011 # 兜底最新年份
            py_int = cy_int - 1
            
            # ====== 第一页：渲染最新一年 (cy_int) ======
            st.markdown(f"<div style='font-size:13px; font-weight:bold; color:#00338D; margin-bottom: 0px;'>▪ {cy_int}年度分析表</div>", unsafe_allow_html=True)
            html_cy = create_financial_report_table(df_filtered, cy_int, selected_cos, divisor, unit_label, current_hl)
            
            if html_cy: 
                # 🌟 核心技巧：只对最新年的表进行“极致压缩”，强行削减 padding 和 行高，防止 PDF 跨页
                html_cy = html_cy.replace("padding: 3px 4px;", "padding: 3px 4px; line-height: 1.1;") # 压扁数据行
                html_cy = html_cy.replace("padding: 4px 3px;", "padding: 4px 3px;") # 压扁表头
                html_cy = html_cy.replace("font-size: 10.5px;", "font-size: 10.5px;") # 字号微缩
                html_cy = html_cy.replace("margin-bottom: 10px;", "margin-bottom: 2px;") # 去掉底部留白
                st.markdown(html_cy, unsafe_allow_html=True)
            
            # ====== 第二页：强制断页，并渲染去年 (py_int) ======
            # 引入专业的 pdf-page-break 强行另起一页
            st.markdown(f"""
            <div class="pdf-page-break"></div>
            <div style='font-size:13px;font-weight:bold;color:#00338D;margin-top:15px;margin-bottom:5px;'>
            ▪ {py_int}年度分析表
            </div>
            """, unsafe_allow_html=True)
            
            html_py = create_financial_report_table(df_filtered, py_int, selected_cos, divisor, unit_label, current_hl)
            
            if html_py: 
                # 🌟 去年的表不做任何替换，原汁原味渲染，保持原有大小
                st.markdown(html_py, unsafe_allow_html=True)

        elif m_id == "nbv_table":
            try: cy_int = int(latest_year)
            except: cy_int = 2024 
            py_int, ppy_int = cy_int - 1, cy_int - 2
            st.markdown(f"<div style='font-size:13px; font-weight:bold; color:#00338D; margin-bottom: 5px;'>▪ {cy_int}年度分析表</div>", unsafe_allow_html=True)
            html_cy = create_nbv_summary_table(df_filtered, selected_cos, divisor, unit_label, cy_int, py_int, current_hl)
            if html_cy: st.markdown(html_cy, unsafe_allow_html=True)
            st.markdown(f"<div style='font-size:13px; font-weight:bold; color:#00338D; margin-bottom: 5px;'>▪ {py_int}年度分析表</div>", unsafe_allow_html=True)
            html_py = create_nbv_summary_table(df_filtered, selected_cos, divisor, unit_label, py_int, ppy_int, current_hl)
            if html_py: st.markdown(html_py, unsafe_allow_html=True)

        elif m_id == "six_dimensional_charts":
            figs = create_six_dimensional_charts(
                df_raw=df_filtered, target_year=latest_year, cos=selected_cos,
                divisor=divisor, unit_label=unit_label, highlight_co=current_hl,
                label_size=11, show_labels=False, dot_size=11
            )
            valid_figs = [f for f in figs if f is not None]
        
            if valid_figs:
                legend_fig = create_six_dimensional_legend(cos=selected_cos, highlight_co=current_hl)
                st.plotly_chart(legend_fig, use_container_width=True, config={"displayModeBar": False})
        
                from plotly.subplots import make_subplots
                pairs = [(valid_figs[i], valid_figs[i+1] if i+1 < len(valid_figs) else None)
                         for i in range(0, len(valid_figs), 2)]
        
                for pair_idx, (left_fig, right_fig) in enumerate(pairs):
                    # ✅ 只在第二对（pair_idx==1）前加续标题，后面不再重复
                    if print_mode and pair_idx == 1:
                        render_continue_title(m_id)
                
                    if right_fig is None:
                        left_fig.update_layout(
                            height=260 if print_mode else 300,  # ✅ 同步缩小
                            margin=dict(t=60, b=10, l=20, r=15),
                            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
                        st.plotly_chart(left_fig, use_container_width=False, config={"displayModeBar": False})
                        continue
                
                    combined = make_subplots(rows=1, cols=2, horizontal_spacing=0.08,
                        subplot_titles=[
                            left_fig.layout.title.text if left_fig.layout.title.text else "",
                            right_fig.layout.title.text if right_fig.layout.title.text else ""
                        ])
                    for trace in left_fig.data:
                        trace.showlegend = False
                        combined.add_trace(trace, row=1, col=1)
                    for trace in right_fig.data:
                        trace.showlegend = False
                        combined.add_trace(trace, row=1, col=2)
                    combined.update_xaxes(left_fig.layout.xaxis.to_plotly_json(), row=1, col=1)
                    combined.update_xaxes(right_fig.layout.xaxis.to_plotly_json(), row=1, col=2)
                    combined.update_yaxes(left_fig.layout.yaxis.to_plotly_json(), row=1, col=1)
                    combined.update_yaxes(right_fig.layout.yaxis.to_plotly_json(), row=1, col=2)
                    combined.update_layout(
                        height=260 if print_mode else 320,  # ✅ 同步缩小
                        margin=dict(t=60, b=30, l=20, r=15),
                        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                        showlegend=False
                    )
                    st.plotly_chart(combined, use_container_width=True, config={"displayModeBar": False})
            else:
                st.info("暂无可用于绘制六维图的数据。")


    # ==========================================
    # 🌟 统一调用出口：根据 m_id 自动组装全部零配件
    # ==========================================
    def render_report_module(m_id, print_mode, is_first=False):
        mod_data = notes_dict.get(m_id, {})
    
        full_title = mod_data.get('title', m_id)
        if 'df_notes' in st.session_state and isinstance(st.session_state['df_notes'], pd.DataFrame):
            df_n = st.session_state['df_notes']
            if '模块ID' in df_n.columns:
                match = df_n[df_n['模块ID'] == m_id]
                if not match.empty:
                    r = match.iloc[0]
                    title_parts = []
                    for field in ['一级分类', '二级分类', '对应图表名称']:
                        val = str(r.get(field, '')).strip()
                        if val and val.lower() != 'nan' and val != '全部':
                            title_parts.append(val)
                    if title_parts:
                        full_title = " - ".join(title_parts)
    
        if print_mode:
            st.markdown("<div class='page-break-container' style='margin:0;padding:0;'>", unsafe_allow_html=True)
        if not print_mode:
            st.markdown(
                "<div class='no-print' style='height:2px; background:linear-gradient(...)'></div>",
                unsafe_allow_html=True
            )    
        # ====== 第 1 步：标题 ======（移出else，打印和网页都执行）
        title_cls = "page-break-title" if (print_mode and not is_first) else ""
        mt = "0px" if (print_mode and is_first) else "20px"
        font_size = "35px" if print_mode else "30px"
        st.markdown(
            f"<h3 class='{title_cls}' style='"
            f"text-align:left; color:#00338D; font-size:{font_size}; font-weight:900; "
            f"font-family:Microsoft YaHei, 微软雅黑, sans-serif; "
            f"margin-top:{mt}; margin-bottom:20px; border:none; padding-bottom:0px;'>"
            f"{full_title}</h3>",
            unsafe_allow_html=True
        )
    
        # ====== 第 2 步：手动截图覆盖 ======
        if 'manual_upload_images' in st.session_state and m_id in st.session_state.manual_upload_images:
            if print_mode:
                st.image(st.session_state.manual_upload_images[m_id], use_column_width=True)
            else:
                img_col_left, img_col_center, img_col_right = st.columns([1, 8, 1])
                with img_col_center:
                    st.image(st.session_state.manual_upload_images[m_id], use_column_width=True)
            _, nt = display_notes(m_id)
            display_bottom_note(nt)
            if print_mode:
                st.markdown("</div>", unsafe_allow_html=True)
            return
    
        # ====== 第 3 步：AI / 注释框 ======
        an, nt = display_notes(m_id, ai_df=df_filtered, ai_field=mod_data.get('title', m_id))
    
        # ====== 第 4 步：图表 ======
        if print_mode:
            if m_id not in ["csm_amortization", "discount_rate", "confidence_level", "csm_maturity_table"]:
                unit_text = "百分比 (%)" if "comp" in m_id or m_id == "asset_struct" or "ratio" in m_id or "margin" in m_id or "csm_trans"in m_id or "struct" in m_id else f"{unit_label}人民币"
                st.markdown(f"<p style='text-align:right; font-size:11px; margin-bottom:2px; color:#666;'>单位：{unit_text}</p>", unsafe_allow_html=True)
            render_pure_chart_entity(m_id, print_mode)
        else:
            chart_col_left, chart_col_center, chart_col_right = st.columns([1, 10, 1])
            with chart_col_center:
                if m_id not in ["csm_amortization", "discount_rate", "confidence_level", "csm_maturity_table"]:
                    unit_text = "百分比 (%)" if "comp" in m_id or m_id == "asset_struct" or "ratio" in m_id or "margin" in m_id or "csm_trans"in m_id or "struct" in m_id else f"{unit_label}人民币"
                    st.markdown(f"<p style='text-align:right; font-size:12px; margin-bottom:2px; color:#666;'>单位：{unit_text}</p>", unsafe_allow_html=True)
                render_pure_chart_entity(m_id, print_mode)
    
        # ====== 第 5 步：底部注释 ======
        display_bottom_note(nt)
    
        if print_mode:
            st.markdown("</div>", unsafe_allow_html=True)
            
#-----------跨页标题重复---------
    def render_continue_title(m_id):
        title = ""
        if 'df_notes' in st.session_state and isinstance(st.session_state['df_notes'], pd.DataFrame):
            df_n = st.session_state['df_notes']
            match = df_n[df_n['模块ID'] == m_id]
            if not match.empty:
                r = match.iloc[0]
                parts = [str(r.get(f,'')).strip() for f in ['一级分类','二级分类','对应图表名称']
                         if str(r.get(f,'')).strip() not in ['','nan','全部']]
                title = " - ".join(parts)
        st.markdown(
            f"<h3 class='print-only page-break-title' style='text-align:left; color:#00338D; font-size:32px; "  # ✅ 加 print-only
            f"font-weight:900; font-family:Microsoft YaHei,微软雅黑,sans-serif; "
            f"margin-top:10px; margin-bottom:0px; border:none;'>"
            f"{title}（续）</h3>",
            unsafe_allow_html=True
        )

    # ==========================================
    # 🌐 网页模式 / 🖨️ 打印模式 的最终执行器
    # ==========================================
    if not print_mode:
        st.markdown(
            "<hr class='no-print' style='border:none;border-top:1px solid #EAEAEA;margin:10px 0;'>",
            unsafe_allow_html=True
        )
    if print_mode:
        if 'ordered_modules' not in locals() or not ordered_modules:
            st.warning("⚠️ 报告顺序由【模块ID】的先后顺序决定，请先在上方传入有模块ID的注释表文件。")
        else:
            import datetime
            today = datetime.date.today()
            date_str = f"{today.year}年{today.month}月"
            type_str = f"（{selected_type}系）" if selected_type and selected_type != "全部" else ""
            
            cover_url = "https://raw.githubusercontent.com/z-xylym/my-actuary-tool/main/%E6%A0%87%E9%A2%98%E9%A1%B5.png"
            back_url  = "https://raw.githubusercontent.com/z-xylym/my-actuary-tool/main/%E5%B0%81%E5%BA%95%E9%A1%B5.png"
            
            # ✅ 封面：用st.image撑满，components.html叠加绝对定位文字
            st.markdown(f"""
            <div style="position:relative; width:100%; aspect-ratio:16/9;
                page-break-after:always; overflow:hidden; margin:0; padding:0;
                -webkit-print-color-adjust:exact; print-color-adjust:exact; forced-color-adjust:none;">
                <img src="{cover_url}" style="width:100%; height:100%; object-fit:cover; display:block;"/>
                <div style="position:absolute; top:0; left:0; width:100%; height:100%;
                    display:flex; flex-direction:column; justify-content:center; align-items:flex-start;
                    padding:0 8%; box-sizing:border-box; margin-top:-30px; z-index:10;
                    forced-color-adjust:none; -webkit-print-color-adjust:exact; print-color-adjust:exact;">
                    <div style="font-size:52px; font-weight:900; line-height:1.4; margin-bottom:16px;
                        font-family:Microsoft YaHei,微软雅黑,sans-serif;
                        color:white; -webkit-text-fill-color:white;
                        text-shadow:2px 2px 4px rgba(0,0,0,0.5), 0 0 20px rgba(0,0,0,0.3);
                        forced-color-adjust:none; -webkit-print-color-adjust:exact;">
                        保险公司{latest_year}年{type_str}<br>新会计准则业绩表现和洞察
                    </div>
                    <div style="font-size:22px; font-weight:500; margin:0;
                        font-family:Microsoft YaHei,微软雅黑,sans-serif;
                        color:white; -webkit-text-fill-color:white;
                        text-shadow:1px 1px 3px rgba(0,0,0,0.5);
                        forced-color-adjust:none; -webkit-print-color-adjust:exact;">{date_str}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            
            # ✅ 图表
            for i, mod in enumerate(ordered_modules):
                render_report_module(mod, print_mode=True, is_first=(i == 0))
            
            st.markdown(f"""
            <div style="position:relative; width:100%; aspect-ratio:16/9;
                page-break-before:always; overflow:hidden; margin:0; padding:0;
                -webkit-print-color-adjust:exact; print-color-adjust:exact; forced-color-adjust:none;">
                <img src="{back_url}" style="width:100%; height:100%; object-fit:cover; display:block;"/>
            </div>
            """, unsafe_allow_html=True)
    else:
        # 如果是网页查看模式，只渲染侧边栏选中的那一个图表！
        if active_m_id:
            # 网页模式下不需要断页，直接传 is_first=True 就行
            render_report_module(active_m_id, print_mode=False, is_first=True)












#____________________________Step8_____________________________#
def show_step_8_content():
    st.markdown("行业分析部分正在开发中。。。敬请期待")




#_______________________________________________________________________
# 1. 先定义这个名单（全局变量）
DEFAULT_COMPANIES = [
    {"类别": "上市", "公司": "中国人寿", "类型": "寿险", "链接地址": "https://www.chinalife.com.cn/chinalife/xxpl/gkxxpl/ndxx/"},
    {"类别": "上市", "公司": "平安寿险", "类型": "寿险", "链接地址": "https://life.pingan.com/p/#/list-page?disclosureChannel=iLifeInfoDisclosure&disclosureName=%E5%B9%B4%E5%BA%A6%E4%BF%A1%E6%81%AF&disclosureId=200c389bb2d2463c8e90dbcfb7965bb2"},
    {"类别": "上市", "公司": "太保寿险", "类型": "寿险", "链接地址": "https://www.cpic.com.cn/xrsbx/gkxxpl/ndxx/?subMenu=2&inSub=1"},
    {"类别": "上市", "公司": "泰康人寿", "类型": "寿险", "链接地址": "https://www.taikanglife.com/publicinfonew/annualinfonew/list_402_1.html"},
    {"类别": "上市", "公司": "新华人寿", "类型": "寿险", "链接地址": "https://www.newchinalife.com/node/398"},
    {"类别": "上市", "公司": "太平人寿", "类型": "寿险", "链接地址": "https://life.cntaiping.com/info-ndxxpl/"},
    {"类别": "上市", "公司": "人保寿险", "类型": "寿险", "链接地址": "https://www.picclife.com/picclifewebsite/webfile/ComprehensiveInformation/index.html"},
    {"类别": "上市", "公司": "阳光人寿", "类型": "寿险", "链接地址": "https://www.sinosig.com/v/pilu?type=sx&tabIndex=10001_31"},
    {"类别": "上市", "公司": "友邦人寿", "类型": "寿险", "链接地址": "https://www.aia.com.cn/zh-cn/gongkaixinxipilou/nianduxinxi"},
    {"类别": "银保", "公司": "中邮人寿", "类型": "寿险", "链接地址": "https://www.chinapost-life.com/publish/publish2/"},
    {"类别": "银保", "公司": "工银安盛", "类型": "寿险", "链接地址": "https://www.icbc-axa.com/public/public_year/public_year1/publicfirstindex.jsp"},
    {"类别": "银保", "公司": "交银人寿", "类型": "寿险", "链接地址": "https://www.bocommlife.com/101843/index.html"},
    {"类别": "银保", "公司": "建信人寿", "类型": "寿险", "链接地址": "https://www.ccb-life.com.cn/html/6182/3230/index.html"},
    {"类别": "银保", "公司": "中信保诚", "类型": "寿险", "链接地址": "https://www.citic-prudential.com.cn/annualinformation/list.html"},
    {"类别": "银保", "公司": "农银人寿", "类型": "寿险", "链接地址": "https://www.abchinalife.com/xxpl/ndxx/index.shtml"},
    {"类别": "银保", "公司": "招商信诺", "类型": "寿险", "链接地址": "https://www.cignacmb.com/xinxi/niandubaogao/ndxibg.html"},
    {"类别": "银保", "公司": "中银三星", "类型": "寿险", "链接地址": "https://www.boc-samsunglife.cn/information?code=GW647"},
    {"类别": "养老健康", "公司": "太平养老", "类型": "养老", "链接地址": "https://www.cntaiping.com/about-ndxx/"},
    {"类别": "养老健康", "公司": "恒安标准养老", "类型": "养老", "链接地址": "https://www.haslpension.com/henganyanglao/gkxxpl/ndxx/index.html"},
    {"类别": "养老健康", "公司": "人保健康", "类型": "健康", "链接地址": "https://www.picc.com.cn/xwzx/gkxx/ndxx/"},
    {"类别": "养老健康", "公司": "平安养老", "类型": "养老", "链接地址": "https://yl.pingan.com/branding/disclosure/annual-info"},
    {"类别": "养老健康", "公司": "泰康养老", "类型": "养老", "链接地址": "https://www.tkpension.com/ndxx/"},
    {"类别": "养老健康", "公司": "平安健康", "类型": "健康", "链接地址": "https://health.pingan.com/P/pubInfoDisclosure?menuName=public_information_disclosure&tagId=9223372036854806087&menuId=4035225266123994694"},
    {"类别": "养老健康", "公司": "太保健康", "类型": "健康", "链接地址": "https://www.cpic.com.cn/jkx/gkxxpl/ndxx/?subMenu=2&inSub=1"},
    {"类别": "外资", "公司": "中英人寿", "类型": "寿险", "链接地址": "https://www.aviva-cofco.com.cn/website/xxzx/gkxxpl/ndxxpl/list-1.shtml"},
    {"类别": "外资", "公司": "中意人寿", "类型": "寿险", "链接地址": "https://www.generalichina.com/ndxx/"},
    {"类别": "外资", "公司": "中荷人寿", "类型": "寿险", "链接地址": "https://www.bob-cardif.com/xinxipilu/nianduxinxi/index.html"},
    {"类别": "外资", "公司": "同方全球", "类型": "寿险", "链接地址": "https://www.aegonthtf.com/gkxxpl/ndxx/ndxx/"},
    {"类别": "外资", "公司": "陆家嘴国泰", "类型": "寿险", "链接地址": "https://www.cathaylife.cn/ndxxplbg/index.html"},
    {"类别": "外资", "公司": "复星保德信", "类型": "寿险", "链接地址": "https://www.pflife.com.cn/fbofficialweb/Annual?submenu=Annual"},
    {"类别": "外资", "公司": "恒安标准", "类型": "寿险", "链接地址": "https://www.hengansl.com/hengan/gkxxpl/ndxx/index.html"},
    {"类别": "小型", "公司": "财信吉祥人寿", "类型": "寿险", "链接地址": "https://life.hnchasing.com/annual_info/annual_info_disclosure/"},
    {"类别": "小型", "公司": "东吴人寿", "类型": "寿险", "链接地址": "https://www.soochowlife.net/cs/gkxxpl/ndxx/index.html"},
    {"类别": "小型", "公司": "国富人寿", "类型": "寿险", "链接地址": "https://www.e-guofu.com/yearsInfo/index.html"},
    {"类别": "小型", "公司": "瑞泰人寿", "类型": "寿险", "链接地址": "https://www.oldmutual-chnenergy.com/InfoPublish/anualReport/"},
    {"类别": "小型", "公司": "东方嘉富人寿", "类型": "寿险", "链接地址": "https://www.sinokorealife.com.cn/ndxx.jhtml"}


]


# ==========================================
# 全局配置：勾稽校验规则（动态年份版）
# ==========================================

DEFAULT_RULES = [
    # === 跨年规则 ===
    {"规则名称": "1. 股东权益期初平衡", "公式": "curr_year['期初股东权益'] == prev_year['期末股东权益']", "类型": "cross", "描述": "本期期初=上期期末"},
    {"规则名称": "2. CSM期初平衡", "公式": "curr_year['CSM期初余额'] == prev_year['CSM期末余额']", "类型": "cross", "描述": "本期期初=上期期末"},
    
    # === 年度内规则 ===
    {"规则名称": "3. CSM余额勾稽", "公式": "abs(curr['CSM期末余额']-(curr['CSM期初余额']+curr['新业务CSM（集团口径）']+curr['CSM计息']+curr['CSM调整']+curr['CSM摊销']+curr['CSM其他']))<5", "类型": "single"},
    {"规则名称": "4. 新业务现值拆分", "公式": "abs(curr['新业务未来现金流入现值']-(curr['新业务未来现金流入现值（盈利）']+curr['新业务未来现金流入现值（亏损）']))<1", "类型": "single"},
    {"规则名称": "5. 保险服务收入合计", "公式": "abs(curr['保险服务收入合计']-(curr['未采用保费分配法计量的保险合同保险服务收入']+curr['采用保费分配法计量的保险合同保险服务收入']))<1", "类型": "single"},
    {"规则名称": "6. 非PAA收入拆分", "公式": "abs(curr['未采用保费分配法计量的保险合同保险服务收入']-(curr['合同服务边际的摊销']+curr['非金融风险调整的变动']+curr['预计当期发生的保险服务费用']+curr['与当期服务或过去服务相关得保费经验调整']+curr['其他收入调整']+curr['保险获取现金流的摊销（保险服务收入）']))<5", "类型": "single"},
    {"规则名称": "7. 保险服务费用合计", "公式": "abs(curr['保险服务费用合计']-(curr['未采用保费分配法计量的保险合同保险服务费用']+curr['采用保费分配法计量的保险合同保险服务费用']))<1", "类型": "single"},
    {"规则名称": "8. 非PAA费用拆分", "公式": "abs(curr['未采用保费分配法计量的保险合同保险服务费用']-(curr['保险获取现金流的摊销（保险服务费用）']+curr['亏损部分的确认及转回']+curr['当期发生的赔款及其他相关费用']+curr['已发生赔款负债相关的履约现金流量变动']))<5", "类型": "single"},
    {"规则名称": "9. 获取费用摊销一致性", "公式": "curr['保险获取现金流的摊销（保险服务收入）']==curr['保险获取现金流的摊销（保险服务费用）']", "类型": "single"},
    {"规则名称": "10. PAA保险业绩", "公式": "abs(curr['采用保费分配法计量的保险合同保险业绩']-(curr['采用保费分配法计量的保险合同保险服务收入']-curr['采用保费分配法计量的保险合同保险服务费用']))<1", "类型": "single"},
    {"规则名称": "11. 服务收入总表一致性", "公式": "curr['保险服务收入']==curr['保险服务收入合计']", "类型": "single"},
    {"规则名称": "12. 税前利润勾稽", "公式": "abs(curr['税前利润总额']-(curr['保险利润']+curr['投资利润']+curr['其他利润']))<5", "类型": "single"},
    {"规则名称": "13. 投资利润勾稽", "公式": "abs(curr['投资利润']-(curr['净投资回报']+curr['承保财务损益']+curr['分出再保险财务损益']))<5", "类型": "single"},
    {"规则名称": "14. 费用分类一致性", "公式": "abs((curr['获取费用']+curr['维持费用']+curr['非履约费用'])-(curr['职工薪酬']+curr['物业及设备支出']+curr['业务投入及监管费用支出']+curr['行政办公支出']+curr['其他支出']))<10", "类型": "single"},
    {"规则名称": "15. 期初保险合同负债总额", "公式": "abs(curr['期初保险合同负债总额']-(curr['CSM期初余额']+curr['BEL期初余额']+curr['RA期初余额']))<1", "类型": "single"},
    {"规则名称": "16. 期末保险合同负债总额", "公式": "abs(curr['期末保险合同负债总额']-(curr['CSM期末余额']+curr['BEL期末余额']+curr['RA期末余额']))<1", "类型": "single"},
    {"规则名称": "17. 非PAA期初余额合计", "公式": "abs(curr['非PAA期初余额合计']-(curr['LRC非亏损部分期初余额（非PAA）']+curr['LRC亏损部分期初余额（非PAA）']+curr['LIC期初余额（非PAA）']))<1", "类型": "single"},
    {"规则名称": "18. 非PAA期末余额合计", "公式": "abs(curr['非PAA期末余额合计']-(curr['LRC非亏损部分期末余额（非PAA）']+curr['LRC亏损部分期末余额（非PAA）']+curr['LIC期末余额（非PAA）']))<1", "类型": "single"},
    {"规则名称": "19. PAA期初余额合计", "公式": "abs(curr['PAA期初余额合计']-(curr['LRC非亏损部分期初余额（PAA）']+curr['LRC亏损部分期初余额（PAA）']+curr['LIC期初BEL余额（PAA）']+curr['LIC期初RA余额（PAA）']))<1", "类型": "single"},
    {"规则名称": "20. PAA期末余额合计", "公式": "abs(curr['PAA期末余额合计']-(curr['LRC非亏损部分期末余额（PAA）']+curr['LRC亏损部分期末余额（PAA）']+curr['LIC期末BEL余额（PAA）']+curr['LIC期末RA余额（PAA）']))<1", "类型": "single"},
    
    # === 合同服务边际来源校验 ===
    {"规则名称": "21. CSM期初余额来源校验", "公式": "abs(curr['CSM期初余额']-(curr['采用公允价值法计量的合同期初负债']+curr['采用修正追溯调整法计量的合同期初负债']+curr['其他合同期初负债']))<1", "类型": "single"},
    {"规则名称": "22. CSM期末余额来源校验", "公式": "abs(curr['CSM期末余额']-(curr['采用公允价值法计量的合同期末负债']+curr['采用修正追溯调整法计量的合同期末负债']+curr['其他合同期末负债']))<1", "类型": "single"},
    
    # === 投资资产勾稽 ===
    {"规则名称": "23. 投资资产勾稽", "公式": "abs(curr['投资资产']-(curr['现金及其他']+curr['投资性房地产']+curr['固定收益类金融资产']+curr['权益类金融资产']+curr['联营及合营企业投资']))<1", "类型": "single"},
    {"规则名称": "26. CSM期初余额来源校验", "公式": "abs(curr['CSM期初余额']-(curr['采用公允价值法计量的合同期初负债']+curr['采用修正追溯调整法计量的合同期初负债']+curr['其他合同期初负债']))<1", "类型": "single"},
    {"规则名称": "27. CSM期末余额来源校验", "公式": "abs(curr['CSM期末余额']-(curr['采用公允价值法计量的合同期末负债']+curr['采用修正追溯调整法计量的合同期末负债']+curr['其他合同期末负债']))<1", "类型": "single"},
    {"规则名称": "28. 投资资产勾稽", "公式": "abs(curr['投资资产']-(curr['现金及其他']+curr['投资性房地产']+curr['固定收益类金融资产']+curr['权益类金融资产']+curr['联营及合营企业投资']))<1", "类型": "single"}
]
# ==================== 1. 页面基础配置 (必须在最顶端且仅此一份) ====================
st.set_page_config(
    page_title="DigiLife 寿研数智 | 年报处理平台", 
    page_icon="Digi.png",  # ✨ 修改点：加上 .png 后缀
    layout="wide"
)

# ==================== 2. 设置左上角 Logo ====================
# 注意：确保 Digi.png 文件和 app.py 在同一个文件夹里
try:
    st.logo(
        "Digi.png",       # ✨ 修改点：加上 .png 后缀
        icon_image="Digi.png" # ✨ 修改点：加上 .png 后缀
    )
except Exception as e:
    st.error(f"Logo 加载失败，请检查文件名是否为 Digi.png。错误信息: {e}")


# ==================== 辅助函数：单位嗅探 ====================
def get_report_unit(text):
    """该公司披露年报的金额单位"""
    if not text: return None
    m1 = re.search(r'(?:金额)?单位[：:\s]*((?:人民币)?[元十百千万亿]+)', text)
    if m1: return m1.group(1).strip()
    m2 = re.search(r'(人民币[十百千万亿]+元)', text)
    if m2: return m2.group(1).strip()
    if "人民币元" in text: return "人民币元"
    return None

# ==================== 后台引擎：调用 DeepSeek 提取单页表格 ====================
import base64

# ==================== 新增后台引擎：调用 Vision VLM 提取图片表格 ====================
def extract_single_page_vision(pdf_bytes, page_num, expected_name, api_key, base_url, model_name):
    """将 PDF 转为高清图片，使用视觉大模型提取表格"""
    try:
        # 1. 内存中将 PDF 页面转为 Base64 高清图
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        if page_num < 1 or page_num > len(doc):
            return None, ""
        page = doc.load_page(page_num - 1)
        
        # 放大 3倍截图，保证文字清晰
        zoom_matrix = fitz.Matrix(3.0, 3.0) 
        pix = page.get_pixmap(matrix=zoom_matrix, alpha=False)
        img_data = pix.tobytes("png")
        base64_image = base64.b64encode(img_data).decode('utf-8')
        doc.close()

        # 2. 设置读图的 Prompt
        prompt = f"""你是一个四大会计师事务所的资深数字化审计专家。
我为你提供了一张【保险公司年报的单页高清截图】，目标是精准提取：【{expected_name}】。
表格中的【数字是由图片构成的】，请发挥强大的视觉识别能力，提取所有文字和表格。
【强制要求】：
1. 所有的科目名、数字金额之间，必须被 "|" 隔开。严禁使用连续空格。
2. 精准对齐多级表头！如果有留白单元格，必须用 "|" 占位补齐！
3. 绝不能漏掉任何一行数据，保留金额里的逗号和括号。
4. 纯文本输出，不要使用 Markdown 代码块。不需要输出 SHEET_NAME 标签，直接排版。"""

        # 3. 呼叫视觉大模型
        client = OpenAI(api_key=api_key, base_url=base_url)
        response = client.chat.completions.create(
            model=model_name,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_image}"}}
                    ]
                }
            ],
            temperature=0.0
        )
        
        # 4. 解析结果
        result_text = response.choices[0].message.content.strip()
        result_text = re.sub(r'^```(csv|text)?\n?', '', result_text, flags=re.MULTILINE).replace("```", "").strip()
        
        lines = result_text.split('\n')
        parsed_data = []
        max_cols = 0
        for row in lines:
            clean_row = row.strip()
            if not clean_row: continue
            if re.match(r'^[\s\|-]+$', clean_row) and '-' in clean_row: continue
            
            if clean_row.startswith('|'): clean_row = clean_row[1:]
            if clean_row.endswith('|'): clean_row = clean_row[:-1]
            
            cols = [col.strip() for col in clean_row.split('|')] if '|' in clean_row else [clean_row]
            parsed_data.append(cols)
            max_cols = max(max_cols, len(cols))
            
        for row in parsed_data:
            if len(row) < max_cols:
                row.extend([''] * (max_cols - len(row)))
                
        return pd.DataFrame(parsed_data), "【提示：该页已使用 OCR 视觉引擎处理，原文本为图片结构】"
        
    except Exception as e:
        return None, f"图像处理失败: {str(e)}"
def extract_single_page(pdf_bytes, page_num, expected_name, api_key, base_url, model_name):
    """使用 pdfplumber 提取文本，并用 LLM 转换为结构化数据"""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        if page_num < 1 or page_num > len(pdf.pages):
            return None, ""
        page = pdf.pages[page_num - 1]
        raw_text = page.extract_text(layout=True)
        
    if not raw_text:
        return None, ""

    prompt = f"""你是一个四大会计师事务所的资深数字化审计专家。
当前任务：提取【{expected_name}】。
请将以下纯文本中的【段落文字】原样保留为一行，将【财务表格】转化为用 "|" 严格分隔的标准网格格式。
【强制要求】：
1. 所有的科目名、附注编号、数字金额之间，必须被 "|" 隔开。严禁使用连续空格。
2. 精准对齐多级表头！如果某个单元格是空的，必须用 "|" 占位补齐！
3. 绝不能漏掉任何一行数据，保留金额里的逗号和括号。
4. 纯文本输出，不要使用 Markdown 代码块。不需要输出 SHEET_NAME，直接输出转化后的数据。

以下是原始文本，请开始输出：
{raw_text}"""


    client = OpenAI(api_key=api_key, base_url=base_url)
    response = client.chat.completions.create(
        model=model_name,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.0
    )
    
    result_text = response.choices[0].message.content.strip()
    result_text = re.sub(r'^```(csv|text)?\n?', '', result_text, flags=re.MULTILINE)
    result_text = result_text.replace("```", "").strip()
    
    # 转换为 DataFrame
    lines = result_text.split('\n')
    parsed_data = []
    max_cols = 0
    for row in lines:
        clean_row = row.strip()
        if not clean_row: continue
        # 忽略大模型可能生成的 markdown 分割线
        if re.match(r'^[\s\|-]+$', clean_row) and '-' in clean_row: continue
        
        if clean_row.startswith('|'): clean_row = clean_row[1:]
        if clean_row.endswith('|'): clean_row = clean_row[:-1]
        
        if '|' not in clean_row:
            cols = [clean_row]
        else:
            cols = [col.strip() for col in clean_row.split('|')]
        parsed_data.append(cols)
        max_cols = max(max_cols, len(cols))
        
    for row in parsed_data:
        if len(row) < max_cols:
            row.extend([''] * (max_cols - len(row)))
            
    return pd.DataFrame(parsed_data), raw_text

# ==================== 后台引擎：调用 DeepSeek (混合双引擎) ====================
def ai_find_pages(pdf_bytes, api_key, target_tables, base_url, model_name):
    """混合双引擎：AI负责读目录推算主表，Python雷达负责深潜寻找附注表"""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    
    # 1. 提取前 60 页（扩大视野，确保覆盖大型公司报表区）
    toc_text = ""
    for i in range(min(60, len(doc))):
        page = doc.load_page(i)
        # 获取文本并去掉多余换行，保留前 1200 个字（通常页眉和目录都在这里）
        page_text = page.get_text("text").replace("\n", " ")
        # 压缩多余空格，节省 AI 的 Token，让它看得更多
        page_text = " ".join(page_text.split())
        toc_text += f"---第{i+1}页---\n{page_text[:1200]}\n"
    # 2. ⚡ 附注专属雷达特征矩阵
    feature_matrix = {
        "保险合同-未采用保费分配法按计量组成部分分析 (原保险)": [
            ["未采用保费分配法", "不采用保费分配法"], 
            ["计量组成部分", "履约现金流量"], 
            ["合同服务边际", "非金融风险调整", "未来现金流量"] 
        ],
        "保险合同-非PAA按未到期责任负债和已发生赔款负债分析 (原保险)": [
            ["未采用保费分配法", "不采用保费分配法"], 
            ["未到期责任负债"], 
            ["已发生赔款负债", "已发生赔款"]
        ],
        "保险合同-PAA按未到期责任负债和已发生赔款负债分析 (原保险)": [
            ["采用保费分配法"], 
            ["未到期责任负债"], 
            ["已发生赔款负债", "已发生赔款"]
        ],
        "保险合同-当期初始确认对资产负债表的影响": [
            ["初始确认的合同", "初始确认的保险合同", "初始确认的影响","如下表所示","如表所示"],  
            ["非亏损合同", "亏损合同"],  
            ["未来现金流出", "未来现金流入", "保险获取现金流量"] 
        ],
        "业务及管理费 (附注)": [
            ["业务及管理费"],
            ["职工薪酬", "折旧", "办公费", "宣传费", "租赁"] # 增加二级科目关键词
        ],
        "评估假设-折现率假设": [
            ["折现率假设", "折现率曲线","精算假设"]
        ],
        "投资资产": [
            [
                "投资组合情况",
                "投资资产",
                "投资组合",
                "投资资产组合",
                "按投资对象分类"
            ],
            [
                "现金及现金等价物",
                "定期存款",
                "债券",
                "债务",
                "股票",
                "基金",
                "长期股权投资",
                "投资性房地产"
            ]
        ],
    }
    
    simple_title_OR = {
        "保险服务收入表 (附注)": ["保险服务收入"],
        "保险服务支出/费用表 (附注)": ["保险服务支出", "保险服务费用"],
        "其他综合收益/损益表 (附注)": ["其他综合收益", "其他综合损益"]
    }
    
    context_anchors = ["项目注释", "财务报表附注", "报表附注", "项目附注"]
    currency_indicators = ["人民币百万元", "人民币千元", "人民币元", "单位：千元", "单位：百万元"]
    time_indicators = ["本年", "上年", "本期", "上期", "期末", "期初", "年初", "年末"]
    
    found_hints = {table: [] for table in target_tables if table in feature_matrix or table in simple_title_OR}
    page_type_map = {} 
    
    for i in range(20, len(doc)):
        page = doc.load_page(i)
        text_raw = page.get_text("text")
        text_clean = text_raw.replace("\n", "").replace(" ", "")
        header_text = text_clean[:800]
        
        is_group = "本集团" in header_text or "合并" in header_text
        is_company = "本公司" in header_text or "母公司" in header_text
        if is_group:
            page_type_map[i+1] = 'group'
        elif is_company:
            page_type_map[i+1] = 'company'
        
        is_in_notes = any(anchor in text_clean for anchor in context_anchors)
        has_currency = any(c in text_clean for c in currency_indicators)
        has_time = any(t in text_clean for t in time_indicators)
        is_continued_table = "(续)" in header_text or "（续）" in header_text
        
        try:
            has_actual_table = len(page.find_tables().tables) > 0
        except:
            has_actual_table = False
            
        is_table_page = is_continued_table or has_actual_table or (has_currency and has_time)
        
        if is_in_notes and is_table_page:
            for table, condition_groups in feature_matrix.items():
                if table in target_tables:
                    if "(原保险)" in table and ("再保险" in header_text or "分出" in header_text):
                        continue
                    if all(any(kw in text_clean for kw in group) for group in condition_groups):
                        if any(kw in header_text for kw in condition_groups[0]) or \
                           (len(condition_groups)>1 and any(kw in header_text for kw in condition_groups[1])):
                            if (i + 1) not in found_hints[table]:
                                found_hints[table].append(i + 1)
                                
            for table, any_of_words in simple_title_OR.items():
                if table in target_tables:
                    if "保险" in table and ("再保险" in header_text or "分出" in header_text):
                        continue
                    if any(kw in header_text for kw in any_of_words):
                        if (i + 1) not in found_hints[table]:
                            found_hints[table].append(i + 1)

    doc.close()
    
    for table, pages in found_hints.items():
        if not pages:
            continue
        group_pages = [p for p in pages if page_type_map.get(p) == 'group']
        company_pages = [p for p in pages if page_type_map.get(p) == 'company']
        
        if group_pages and company_pages:
            found_hints[table] = [p for p in pages if p not in company_pages]
    
    hint_text = "\n\n【⚡ Python程序附注雷达线索】\n"
    for table, pages in found_hints.items():
        if pages:
            pages = sorted(list(set(pages)))
            hint_text += f"- {table} 真实表格物理页码位于: {pages[:4]}\n"
    
    prompt = f"""你是一个顶级的寿险审计专家，任务是定位 PDF 中的核心财务报表页码。

【关键定位逻辑（双模式自适应）】
模式 A：目录驱动（适用于有目录的报表）
- 如果前几页存在“目录”或“Contents”，请识别目标报表对应的页码。
- 计算物理偏移量（封面/说明页占用的页数），得出真实的物理页码。

模式 B：标题驱动（适用于无目录或目录不全的报表，如太平人寿）
- 如果没有明确目录，请直接扫描文本中 `---第N页---` 标记下方的页眉或正文大标题。
- 例如：若在 `---第8页---` 下方看到“合并利润表”，则物理页码即为 8。

【通用执行准则】
1. 🎯 如果需求表单名称包含“（合并）”，优先寻找：
   合并资产负债表
   合并利润表
   合并现金流量表
   合并股东权益变动表

2. 🎯 如果需求表单名称包含“（公司）”，优先寻找：
   公司资产负债表
   公司利润表
   公司现金流量表
   公司股东权益变动表

3. 🎯 如果年报中没有区分“合并”和“公司”，仅出现：
   资产负债表
   利润表
   现金流量表
   股东权益变动表

则同一页码同时返回给：
   对应的（合并）
   对应的（公司）
4. 🎯 跨页逻辑：由于财务报表通常包含“(续)”，请务必返回所有相关的物理页码数组。
   - 示例：若第 3 页是资产负债表，第 4 页是其负债部分的续表，返回 [3, 4]。
5. 🎯 附注表线索（最高优先级）：对于带有“(附注)”或“保险合同”字样的复杂底表，请【无脑完全信任】下方提供的【Python程序附注雷达线索】，直接返回线索中的页码。
6. 🎯 特别提醒：现金流量表有时在目录中比较隐蔽，请仔细搜寻；若无目录，请在资产负债表和利润表之后的 5-10 页内搜寻其标题。

需求表单列表：
{json.dumps(target_tables, ensure_ascii=False)}

【待扫描文本内容（前 60 页）】：
{toc_text}

【⚡ Python程序附注雷达辅助线索】：
{hint_text}

【强制输出格式】
只输出合法 JSON，键为表单名，值为物理页码数组。找不到填 [0]。
⚠️ 警告：输出的 JSON 的键（Key）必须【完全包含】《需求表单列表》中的所有字段，【绝对不能有任何遗漏】！如果文中真的找不到，请严格对应填写 [0]。
格式示例：{{"合并利润表": [8, 9], "资产负债表 (合并优先)": [2, 3, 4]}}"""


    client = OpenAI(api_key=api_key, base_url=base_url) # 网址动态化
    response = client.chat.completions.create(
        model=model_name, # 模型动态化
        messages=[{"role": "user", "content": prompt}],
        temperature=0.1
    )
    
    result_text = response.choices[0].message.content.strip()
    result_text = result_text.replace("```json", "").replace("```", "").strip()
    return json.loads(result_text)

# ==================== 全量 CSS 美化 ====================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@300;400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: "PingFang SC", "HarmonyOS Sans", "Microsoft YaHei", "Noto Sans SC", sans-serif !important; color: #4A5568; }
    .stApp { background: linear-gradient(160deg, #F8FAFC 0%, #F1F5F9 50%, #E2E8F0 100%); }
    ::-webkit-scrollbar { width: 6px; height: 6px; }
    ::-webkit-scrollbar-track { background: transparent; }
    ::-webkit-scrollbar-thumb { background: #B0BEC5; border-radius: 3px; }
    ::-webkit-scrollbar-thumb:hover { background: #90A4AE; }
    ::selection { background: #B8C5D6; color: #1E293B; }
    [data-testid="stSidebar"] { background: linear-gradient(180deg, #E2E8F0 0%, #CBD5E1 100%) !important; border-right: 1px solid #94A3B8; }
    [data-testid="stSidebar"] .stTextInput label, [data-testid="stSidebar"] .stMarkdown { color: #334155 !important; font-weight: 500; }
    h1 { color: #111827 !important; font-weight: 600 !important; letter-spacing: 2px; border-bottom: none !important; padding-bottom: 10px; background: transparent !important; }
    h3 { color: #1E293B !important; font-size: 17px !important; font-weight: 600 !important; letter-spacing: 1px; background: rgba(255, 255, 255, 0.45) !important; backdrop-filter: blur(10px); -webkit-backdrop-filter: blur(10px); border: 1px solid rgba(255, 255, 255, 0.8); border-radius: 8px !important; padding: 8px 18px !important; box-shadow: 0 4px 12px rgba(0, 0, 0, 0.03), inset 0 1px 0 rgba(255, 255, 255, 0.5) !important; width: fit-content; margin-bottom: 18px !important; margin-top: 10px !important; }
    .stButton > button { font-size: 14px !important; font-weight: 600 !important; letter-spacing: 1.5px; background-color: #94A3B8; color: #FFFFFF; border: 1px solid #64748B; border-radius: 4px; padding: 8px 24px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1), 0 2px 4px -1px rgba(0,0,0,0.06); transition: all 0.2s ease; }
    .stButton > button:hover { background-color: #64748B; border-color: #475569; color: #F8FAFC; transform: translateY(-1px); box-shadow: 0 10px 15px -3px rgba(0,0,0,0.1); }
    .stButton > button:active { transform: translateY(1px); box-shadow: none; }
    .stButton > button:disabled { background-color: #CBD5E1 !important; color: #94A3B8 !important; border-color: #CBD5E1 !important; box-shadow: none !important; cursor: not-allowed; }
    .stTabs [data-baseweb="tab-list"] { gap: 4px; background-color: transparent; border-bottom: 2px solid #CBD5E1; }
    .stTabs [data-baseweb="tab"] { font-size: 14px; letter-spacing: 0.5px; background-color: #EDF2F7; color: #64748B; border-radius: 6px 6px 0 0; border: 1px solid #CBD5E1; border-bottom: none; padding: 10px 22px; transition: all 0.15s ease; }
    .stTabs [data-baseweb="tab"]:hover { background-color: #E2E8F0; color: #475569; }
    .stTabs [aria-selected="true"] { background-color: #94A3B8 !important; color: #FFFFFF !important; font-weight: 600 !important; border-color: #64748B !important; }
    .stTextInput > div > div > input, .stNumberInput > div > div > input { background-color: #FFFFFF !important; border: 1px solid #CBD5E1 !important; border-radius: 4px !important; color: #1E293B !important; font-weight: 500; transition: border-color 0.2s ease; }
    .stTextInput > div > div > input:focus, .stNumberInput > div > div > input:focus { border-color: #64748B !important; box-shadow: 0 0 0 2px rgba(100,116,139,0.15) !important; }
    .stMultiSelect [data-baseweb="tag"] { background-color: #CBD5E1 !important; color: #1E293B !important; border-radius: 4px !important; font-size: 13px !important; }
    [data-testid="stFileUploader"] section { border: 2px dashed #B0BEC5 !important; border-radius: 8px !important; background-color: #FAFBFC !important; transition: border-color 0.2s ease; }
    [data-testid="stFileUploader"] section:hover { border-color: #78909C !important; background-color: #F5F7FA !important; }
    hr { border: none !important; height: 1px !important; background: linear-gradient(to right, transparent, #B0BEC5, transparent) !important; margin: 20px 0 !important; }
    .info-card { background: #FFFFFF; border: 1px solid #E2E8F0; border-left: 4px solid #94A3B8; border-radius: 6px; padding: 20px 24px; margin: 12px 0; box-shadow: 0 1px 3px rgba(0,0,0,0.04); }
    .info-card h4 { color: #475569; margin: 0 0 8px 0; font-weight: 600; font-size: 15px; }
    .info-card p, .info-card li { color: #64748B; font-size: 14px; line-height: 1.8; }
    .info-card.pink { border-left-color: #D4A5A5; }
    .info-card.green { border-left-color: #A5C4B5; }
    .info-card.blue { border-left-color: #94A3B8; }
    .sidebar-brand { text-align: center; padding: 8px 0 16px 0; border-bottom: 1px solid #B0BEC5; margin-bottom: 20px; }
    .sidebar-brand .logo-text { font-size: 20px; font-weight: 700; color: #475569; letter-spacing: 3px; }
    .sidebar-brand .logo-sub { font-size: 11px; color: #94A3B8; letter-spacing: 4px; margin-top: 4px; }
    .placeholder-section { text-align: center; padding: 40px 20px; color: #94A3B8; }
    .placeholder-section .big-icon { font-size: 48px; margin-bottom: 16px; opacity: 0.6; }
    .placeholder-section p { font-size: 14px; line-height: 2; color: #64748B; }
</style>
""", unsafe_allow_html=True)



# ==========================================
# 🔐 1. 初始化系统状态变量 (紧凑写法)
# ==========================================
for k, v in {'logged_in':False, 'user_role':None, 'api_key':"", 'base_url':"https://api.deepseek.com", 'model_name':"deepseek-chat"}.items():
    if k not in st.session_state: st.session_state[k] = v

# ==========================================
# 🚪 2. 独立登录界面 (极致压缩 + 旗舰科技发光主题)
# ==========================================
if not st.session_state['logged_in']:
    
    # 这段专属暗色 CSS 仅在登录界面生效，登录后自动恢复上方的主系统浅色配置
    st.markdown("""<style>
    header {visibility: hidden;}
    .stApp {background: linear-gradient(135deg, #040B16 0%, #0A1931 50%, #002266 100%); color: #E2E8F0;}
    /* 毛玻璃卡片底座，微微泛青蓝光 */
    [data-testid="column"]:nth-of-type(2) {background: rgba(255,255,255,0.08); backdrop-filter: blur(25px); -webkit-backdrop-filter: blur(25px); border-radius: 20px; border: 1px solid rgba(0,243,255,0.3); box-shadow: 0 15px 35px rgba(0,0,0,0.5), inset 0 0 15px rgba(0,243,255,0.15); padding: 40px 30px; margin-top: 6vh;}
    
    /* 💎 寿研数智年报平台 - 顶级银青渐变全息发光 */
    .title-glow {background: linear-gradient(90deg, #FFFFFF 0%, #76D2FF 50%, #00F3FF 100%); -webkit-background-clip: text; -webkit-text-fill-color: transparent; font-weight: 900; filter: drop-shadow(0 0 15px rgba(0,243,255,0.6));}

    div[data-baseweb="input"]>div, div[data-baseweb="select"]>div {background: rgba(0,0,0,0.4)!important; border: 1px solid rgba(0,243,255,0.3)!important; border-radius: 8px!important;}
    div[data-baseweb="input"]>div:focus-within, div[data-baseweb="select"]>div:focus-within {border-color: #00F3FF!important; box-shadow: 0 0 12px rgba(0,243,255,0.5)!important;}
    
    div[data-testid="stRadio"] label p {color: #FFFFFF!important; font-weight: bold!important; font-size: 14px!important;}
    
    /* 🌟 穿透级 CSS：强制修改 AI 下拉框和 API 框的文字颜色为 #00F3FF */
    label p, .stSelectbox label p, .stTextInput label p {color: #00F3FF!important; letter-spacing: 1px!important;}
    input, .stSelectbox span {color: #FFFFFF!important; font-size: 14px!important;}
    
    /* 大按钮发光样式 */
    button[kind="primary"] {background: linear-gradient(90deg, #0044CC, #0088FF)!important; border: 1px solid rgba(0,243,255,0.5)!important; box-shadow: 0 0 15px rgba(0,136,255,0.4)!important; color: white!important; font-weight: bold!important; letter-spacing: 2px!important; border-radius: 8px!important; margin-top: 5px!important;}
    button[kind="primary"]:hover {box-shadow: 0 0 25px rgba(0,243,255,0.8)!important; transform: scale(1.02);}
    
    /* 💡 将 popover 悬浮窗按钮伪装成靠右的超链接小字 */
    [data-testid="stPopover"] {display: flex; justify-content: flex-end;}
    [data-testid="stPopover"] button {background: transparent!important; border: none!important; box-shadow: none!important; color: #94A3B8!important; font-size: 12px!important; font-weight: normal!important; padding: 0!important; min-height: 0!important; text-decoration: underline; margin-top: 8px; margin-bottom: 15px;}
    [data-testid="stPopover"] button:hover {color: #00F3FF!important;}
    </style>""", unsafe_allow_html=True)

    _, col2, _ = st.columns([1, 1.5, 1])
    with col2:
        st.markdown("<div style='text-align:center; margin-bottom:30px;'><h1 style='font-size:32px; margin:0;'><span class='title-glow'>寿研数智年报平台</span></h1><p style='color:#76D2FF; font-size:11px; letter-spacing:4px; margin-top:8px; font-weight:bold;'>ACTUARIAL INTELLIGENCE</p></div>", unsafe_allow_html=True)
        
        u_type = st.radio("访问权限", ["普通用户", "项目组成员"], label_visibility="collapsed", horizontal=True)
        
        # 🌟 修改 1：不管选什么角色，都显示密码输入框
        sec_code = st.text_input("安全验证", type="password", placeholder="请输入内部安全码") 
        
        # 字典映射压缩代码体积
        ai_map = {"阿里云百炼 (通义千问)": ("https://dashscope.aliyuncs.com/compatible-mode/v1", "qwen-plus"), "DeepSeek (深度求索)": ("https://api.deepseek.com", "deepseek-chat"), "月之暗面 (Kimi)": ("https://api.moonshot.cn/v1", "moonshot-v1-8k"), "智谱AI (GLM-4)": ("https://open.bigmodel.cn/api/paas/v4", "glm-4"), "OpenAI (ChatGPT)": ("https://api.openai.com/v1", "gpt-4o")}
        ai_pr = st.selectbox("选择您将使用的AI", list(ai_map.keys()) + ["自定义私有化节点"])
        
        d_url, d_mod = ai_map.get(ai_pr, ("https://api.deepseek.com", "deepseek-chat"))
        if ai_pr == "自定义私有化节点":
            d_url, d_mod = st.text_input("RPC 接口地址", "https://api.deepseek.com"), st.text_input("指定模型版本", "deepseek-chat")

        api_input = st.text_input("请填写您使用AI的API Key", type="password", placeholder=" sk-... (不填则仅开放年报检测、数据合并及可视化本地分析功能)")
        
        # 条件判断：只有选了 OpenAI 时才显示红字警告，否则用一个很小的空白顶位
        if "OpenAI" in ai_pr:
            st.markdown("<p style='font-size:11px; color:#F87171; text-align:right; margin-top:4px; margin-bottom:0;'>⚠️ 严禁向境外节点传输涉密数据</p>", unsafe_allow_html=True)
        
        # 伪装成超链接小字的帮助菜单
        with st.popover("如何获取API key?"):
            st.markdown("**1.** 前往各大模型官方开放平台注册账号。\n\n**2.** 在控制台生成 `sk-` 开头的密钥。\n\n**3.** 复制填入上方输入框即可体验 AI 功能。")

        # 启动大按钮
        if st.button("启 动 系 统", type="primary", use_container_width=True):
            # 🌟 修改 2：分别校验普通用户和项目组成员的密码
            if u_type == "普通用户" and sec_code != "KPMG1234": 
                st.error("❌ 拒绝访问：普通用户安全码错误")
            elif u_type == "项目组成员" and sec_code != "KPMG666": 
                st.error("❌ 拒绝访问：项目组成员安全码错误")
            else: 
                st.session_state.update({'logged_in':True, 'user_role':u_type, 'api_key':api_input, 'base_url':d_url, 'model_name':d_mod})
                st.rerun()

        st.markdown("<div style='text-align:center; color:#94A3B8; font-size:11px; margin-top:30px; letter-spacing:1px;'>系统版本：v3.0 (Alpha) © 2026<br>Developed by 林友沐Bella@KPMG</div>", unsafe_allow_html=True)

    # 阻止程序继续往下渲染主系统界面
    st.stop()
# ==========================================
# 📊 3. 主系统界面 (根据角色分发权限)
# ==========================================
else:
    # 代码规范锁：0123456789
    api_key = st.session_state.get('api_key', "")
    base_url = st.session_state.get('base_url', "https://api.deepseek.com")
    model_name = st.session_state.get('model_name', "deepseek-chat")
    user_role = st.session_state.get('user_role', "普通用户") 

    # 顶部状态栏与退出按钮
    top_col1, top_col2 = st.columns([8, 1])
    with top_col1:
        st.markdown(
            f"<div class='no-print' style='color: #6c757d; font-size: 14px; margin-top: 8px;'>"
            f"当前身份：<b>{user_role}</b> | 欢迎使用寿研数智分析平台"
            f"</div>", 
            unsafe_allow_html=True
        )
    with top_col2:
        if st.button("退出登录"):
            st.session_state['logged_in'] = False
            st.rerun()
            
    # ==================== 主界面标题 ====================
    st.markdown("<hr style='margin-top: 5px; margin-bottom: 15px;'>", unsafe_allow_html=True)
    st.markdown("<h1 class='no-print' style='font-weight:700; padding-top:10px;'>寿研数智・年报处理平台</h1>", unsafe_allow_html=True)

    # ---------------- 核心权限控制 ----------------
    if st.session_state['user_role'] == "项目组成员":
        # 如果是 KPMG 成员，完整显示 0 到 7 一共 8 个 Tab
        tab0, tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
            " 🌐 Step 0 ／ 官网年报监控 ",
            " 📑 Step 1 ／ 智能页码检索 ",
            " ⚡ Step 2 ／ 表格智能转换 ",
            " 📝 Step 3 ／ 目标表标准填报 ",
            " 🔍 Step 4 ／ 数据勾稽检查 ",
            " ⛓️‍💥 Step 5 ／ 多公司合并 ",
            " 📊 Step 6 ／ 自定义对标分析 ",
            " 🖼️ Step 7 ／ 公司级对标报告 ",
            " 📈 Step 8 ／ 行业分类统计分析 "
        ])
        
        # ─────────── 以下是各 Tab 对应的内容 ───────────
        with tab0:
            st.markdown("### 🌐 寿险公司官网年报监控")
            
            # 1. 年份选择器
            col_t0_1, _ = st.columns([1, 2])
            with col_t0_1:
                target_year = st.number_input("📅 请选择监控年份", min_value=2010, max_value=2050, value=2025, step=1)
            
            st.markdown(f"""
            <div class="info-card green">
                <h4>功能说明</h4>
                <p>系统将自动扫描下列 36 家险企官网，检测网页中是否出现 <b>{target_year}</b> 字样。您可以<b>双击下方表格修改网址</b>。检测到后请点击链接手动下载 PDF 并前往 Step 1。</p>
            </div>
            """, unsafe_allow_html=True)
        
            # 2. 将预设数据加载到 Session State
            if 'company_df' not in st.session_state:
                # 使用之前提供的 36 家公司 DEFAULT_COMPANIES 列表
                st.session_state.company_df = pd.DataFrame(DEFAULT_COMPANIES)
            
            # 3. 交互式数据表 (可直接修改网址)
            st.markdown("#### 🏢 监控目标名单 (双击下方链接可修改)")
            edited_df = st.data_editor(
                st.session_state.company_df,
                column_config={
                    "链接地址": st.column_config.TextColumn("🌐 网页链接地址 (支持双击编辑)", max_chars=1000, width="large"),
                    "类别": st.column_config.TextColumn(disabled=True),
                    "公司": st.column_config.TextColumn(disabled=True),
                    "类型": st.column_config.TextColumn(disabled=True),
                },
                use_container_width=True,
                hide_index=True,
                key="company_data_editor_step0"
            )
            st.session_state.company_df = edited_df
        
            st.markdown("---")
            
            # 4. 扫描按钮逻辑
            col_btn1, col_btn2 = st.columns([1, 1])
            
            with col_btn1:
                target_company = st.selectbox("🎯 单家快速检索：", options=edited_df['公司'].tolist())
                single_scan = st.button(f"🔍 检索【{target_company}】", use_container_width=True)
            
            with col_btn2:
                st.write("") # 对齐
                batch_scan = st.button("🚀 启动全网 36 家批量扫描", type="primary", use_container_width=True)
        
            # 5. 执行扫描逻辑 (通用函数)
            if single_scan or batch_scan:
                # 确定要扫描的任务列表
                if single_scan:
                    tasks = edited_df[edited_df['公司'] == target_company].to_dict('records')
                else:
                    tasks = edited_df.to_dict('records')
                    
                results = []
                my_bar = st.progress(0, text="准备启动扫描...")
                total_tasks = len(tasks)
                
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
                }
        
                for index, row in enumerate(tasks):
                    company_name = row['公司']
                    url = str(row['链接地址'])
                    my_bar.progress((index + 1) / total_tasks, text=f"正在扫描 ({index+1}/{total_tasks}): {company_name}")
                    
                    status = "🔴 未更新 / 无法访问"
                    if url.lower().endswith('.pdf'):
                        status = "⚠️ 直接PDF链接 (需手动核实)"
                    elif "http" in url:
                        try:
                            # 原有的爬虫逻辑开始
                            response = requests.get(url, headers=headers, timeout=8)
                            response.encoding = response.apparent_encoding 
                            soup = BeautifulSoup(response.text, 'html.parser')
                            page_text = soup.get_text()
                            
                            # 检查网页文本中是否包含用户输入的年份
                            if str(target_year) in page_text:
                                status = "🟢 极可能已更新!"
                            # 原有的爬虫逻辑结束
                        except Exception as e:
                            status = "🟡 网站拦截/超时，需手动查看"
                    else:
                        status = "无效链接"
                    
                    results.append({
                        "公司名称": company_name,
                        "监控状态": status,
                        "检测结果描述": f"网页中已检索到 {target_year} 字样" if "🟢" in status else "未发现关键字",
                        "直达链接": url
                    })
                    time.sleep(0.3) # 稍微提速
        
                my_bar.empty()
                st.success(f"🎉 {target_year}年度检测扫描完成！")
                
                # 6. 展示扫描结果
                df_result = pd.DataFrame(results)
                st.data_editor(
                    df_result,
                    column_config={
                        "直达链接": st.column_config.LinkColumn("点击前往网页"),
                        "监控状态": st.column_config.TextColumn("状态", width="medium")
                    },
                    hide_index=True,
                    use_container_width=True,
                    key="scan_result_display"
                )
                st.info("💡 提示：请点击标记为 🟢 的公司链接下载 PDF，下载后请前往 [Step 1] 进行页码定位。")
            
            # ─────────── Step 1：智能页码检索 ───────────
        with tab1:
            st.markdown("### 📑 智能页码检索")
            uploaded_file = st.file_uploader(
                "拖拽或选择一份已经下载好的年报 PDF 文件",
                type="pdf",
                help="推荐上传一份寿险公司的完整年报"
            )
        
            if uploaded_file:
                if 'pdf_bytes' not in st.session_state or st.session_state.get('pdf_name') != uploaded_file.name:
                    st.session_state['pdf_bytes'] = uploaded_file.read()
                    st.session_state['pdf_name'] = uploaded_file.name
                    if 'found_pages' in st.session_state:
                        del st.session_state['found_pages']
                    if 'edited_pages' in st.session_state:
                        del st.session_state['edited_pages']
                
                pdf_bytes = st.session_state['pdf_bytes']
                doc = fitz.open(stream=pdf_bytes, filetype="pdf")
                total_pages = len(doc)
                
                st.caption(f"当前加载文件：{uploaded_file.name}　|　文档共 {total_pages} 页")
                st.markdown("---")
                
                col_left, col_spacer, col_right = st.columns([1, 0.05, 1.2])
        
                with col_left:
                    st.markdown("#### 检索目标设定")
                    target_tables = st.multiselect(
                        "请选择需要定位的报表：",
                        [
                            "资产负债表（合并）", "利润表（合并）","现金流量表（合并）", "股东/所有者权益变动表（合并）", "资产负债表（公司）", "利润表（公司）",
                            "现金流量表（公司）", "股东/所有者权益变动表（公司）", "保险服务收入表 (附注)", "保险服务支出/费用表 (附注)", "其他综合收益/损益表 (附注)", "业务及管理费 (附注)",
                            "保险合同-非PAA按未到期责任负债和已发生赔款负债分析 (原保险)","保险合同-PAA按未到期责任负债和已发生赔款负债分析 (原保险)", "保险合同-未采用保费分配法按计量组成部分分析 (原保险)","保险合同-当期初始确认对资产负债表的影响",
                            "评估假设-折现率假设","投资资产"
                        ],
                        default=[   "资产负债表（合并）","利润表（合并）", "保险合同-非PAA按未到期责任负债和已发生赔款负债分析 (原保险)"]
                    )
                    
                    st.markdown("")  
                    btn_search = st.button("启动智能检索", use_container_width=True)
                    
                    if btn_search:
                        # 🌟 黄金法则：直接从内存里现取，绝对不会报 NameError！
                        current_api_key = st.session_state.get('api_key', "").strip()
                        
                        if not current_api_key:
                            st.error("⚠️ 未检测到 API Key！请刷新页面返回登录界面填写。")
                        elif not target_tables:
                            st.error("请至少选择一张报表。")
                        else:
                            with st.spinner("需要一些时间，请稍等~"):
                                try:
                                    # 这里的传参也要换成刚刚取出来的 current_api_key
                                    result = ai_find_pages(pdf_bytes, current_api_key, target_tables, base_url, model_name)
                                    main_table_pairs = [
                                        ("资产负债表（合并）","资产负债表（公司）"),
                                        ("利润表（合并）","利润表（公司）"),
                                        ("现金流量表（合并）","现金流量表（公司）"),
                                        ("股东/所有者权益变动表（合并）","股东/所有者权益变动表（公司）")
                                    ]
                                    
                                    for merge_key, company_key in main_table_pairs:
                                    
                                        merge_pages = result.get(merge_key, [0])
                                        company_pages = result.get(company_key, [0])
                                    
                                        # AI只找到合并
                                        if merge_pages != [0] and company_pages == [0]:
                                            result[company_key] = merge_pages
                                    
                                        # AI只找到公司
                                        elif company_pages != [0] and merge_pages == [0]:
                                            result[merge_key] = company_pages
                                    st.session_state['found_pages'] = result
                                    st.success("检索完成！请在下方核对物理页码。")
                                except Exception as e:
                                    st.error(f"引擎出现异常：{e}")
        
                    if 'found_pages' in st.session_state:
                        st.markdown("---")
                        st.markdown("#### 结果核对")
                        st.caption("提示：若表格跨页，请以英文逗号分隔页码（如 64, 65）。可结合右侧预览进行校准。")
                        
                        edited_pages = {}
                        # ✅ 关键修改点：强制遍历用户选择的 target_tables！
                        for table_name in target_tables:
                            # 如果 AI 返回了结果就用 AI 的，如果 AI 偷懒漏掉了，就默认给 [0]
                            page_data = st.session_state['found_pages'].get(table_name, [0])
                            
                            if isinstance(page_data, int):
                                page_data = [page_data]
                            str_val = ", ".join(map(str, page_data))
                            user_input = st.text_input(f"{table_name}", value=str_val, key=f"page_{table_name}")
                            
                            try:
                                edited_pages[table_name] = [int(x.strip()) for x in user_input.split(",") if x.strip().isdigit()]
                            except:
                                edited_pages[table_name] = page_data
                        
                        st.session_state['edited_pages'] = edited_pages
                        
                        st.markdown("")
                        if st.button("确认页码，进入下一步", use_container_width=True):
                            st.success("页码已确认！请前往 Step 2 进行表格转换。")
        
                with col_right:
                    st.markdown("#### 页面预览")
                    if 'found_pages' in st.session_state and 'edited_pages' in st.session_state:
                        # ✅ 修改点：换成依赖左侧整理好的完整字典（edited_pages）！
                        table_to_view = st.selectbox("选择要预览的报表：", options=list(st.session_state['edited_pages'].keys()))
                        
                        # 获取该表对应的页码，如果是空，默认给个 [0] 兜底
                        pages_to_preview = st.session_state['edited_pages'].get(table_to_view, [0])
                        if not pages_to_preview:
                            pages_to_preview = [0]
                        
                        if len(pages_to_preview) > 1:
                            current_page = st.radio("该报表包含多页，请切换预览：", options=pages_to_preview, horizontal=True)
                        else:
                            current_page = pages_to_preview[0]
                        
                        preview_idx = current_page - 1
                        
                        if 0 <= preview_idx < total_pages:
                            page = doc.load_page(preview_idx)
                            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                            st.image(pix.tobytes("png"), caption=f"当前预览：第 {current_page} 页 / 共 {total_pages} 页", use_column_width=True)
                        elif current_page == 0:
                            st.info("尚未识别到页码，请在左侧输入框中手动填入。")
                        else:
                            st.warning("该页码超出了文档总页数，请检查左侧修改。")
                    else:
                        st.markdown("")
                        st.info("上传文件并启动检索后，此处将显示对应页面。")
            
    # ─────────── Step 2：表格智能转换 ───────────
        with tab2:
            st.markdown("### ⚡ 表格智能转换")
            st.markdown("""
            <div class="info-card blue">
                <h4>功能说明</h4>
                <p>系统将调用大模型对多页 PDF 进行网格化对齐重构，并自动拼接同表跨页数据，生成标准化 Excel。<b>如果出现提取失败或错位</b>，请开启下方的【图片扫描模式】重试。</p>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown("")
            
            # === 🌟 引擎选择开关 ===
            st.markdown("#### PDF类型选择")
            use_vision = st.toggle("📸 开启图片扫描模式 (适用于扫描版、纯图片PDF)", value=False)
            if use_vision:
                st.caption("提示：图片扫描模式请确保您登录时选择的模型支持视觉功能（如 GPT-4o, 阿里云通义千问等）。DeepSeek 暂不支持直接读图。")
            st.markdown("---")
            
            if 'edited_pages' not in st.session_state or 'pdf_bytes' not in st.session_state:
                st.warning("⚠️ 请先在 Step 1 中上传文件并完成【页码确认】。")
                st.button("开始提取结构化数据", disabled=True, use_container_width=True)
            else:
                # 🌟 从全局内存中直接安全获取用户在登录页填写的配置
                current_api_key = st.session_state.get('api_key', "").strip()
                current_base_url = st.session_state.get('base_url', "https://api.deepseek.com")
                current_model_name = st.session_state.get('model_name', "deepseek-chat")
                
                # === 校验 API Key ===
                can_run = True
                if not current_api_key:
                    st.error("⚠️ 未检测到 API Key！请刷新页面返回登录界面填写授权密钥。")
                    can_run = False
        
                if not can_run:
                    st.button("开始提取结构化数据", disabled=True, use_container_width=True)
                else:
                    valid_tasks = {k: v for k, v in st.session_state['edited_pages'].items() if v != [0]}
                    
                    if not valid_tasks:
                        st.warning("⚠️ 没有找到任何有效的页码，无需提取。")
                    else:
                        if st.button("开始提取", use_container_width=True):
                            with st.spinner("☕ 趁现在喝口水吧，正在后台拼命打工中..."):
                                extracted_sheets = {}
                                global_unit = "未能自动提取，需人工核对"
                                all_tasks = []
                                for table_name, pages in valid_tasks.items():
                                    for page_num in pages:
                                        all_tasks.append({"table": table_name, "page": page_num})
                                
                                total_pages_to_process = len(all_tasks)
                                pages_done = 0
                                
                                progress_bar = st.progress(0)
                                status_box = st.status(f"正在转化 {total_pages_to_process} 个任务...", expanded=True)
                                temp_results = {table_name: {} for table_name in valid_tasks.keys()}
                                
                                with status_box:
                                    # 💡 核心优化：视觉模型并发太高容易被限流，设定为2；普通文本设为5
                                    workers = 2 if use_vision else 5 
                                    
                                    with ThreadPoolExecutor(max_workers=workers) as executor:
                                        future_to_task = {}
                                        
                                        for task in all_tasks:
                                            # 👇 核心分流逻辑：根据开关决定派发哪个函数 👇
                                            if use_vision:
                                                future = executor.submit(
                                                    extract_single_page_vision, 
                                                    st.session_state['pdf_bytes'], 
                                                    task["page"], 
                                                    task["table"], 
                                                    current_api_key,     # 统一使用全局 Key
                                                    current_base_url,    # 统一使用全局 URL
                                                    current_model_name   # 统一使用全局 Model
                                                )
                                            else:
                                                future = executor.submit(
                                                    extract_single_page, 
                                                    st.session_state['pdf_bytes'], 
                                                    task["page"], 
                                                    task["table"], 
                                                    current_api_key,
                                                    current_base_url,    # 统一使用全局 URL
                                                    current_model_name   # 统一使用全局 Model# 统一使用全局 Key
                                                )
                                            future_to_task[future] = task
                                        
                                        for future in as_completed(future_to_task):
                                            task = future_to_task[future]
                                            t_name = task["table"]
                                            p_num = task["page"]
                                            try:
                                                df, raw_text = future.result()
                                                if df is not None and not df.empty:
                                                    temp_results[t_name][p_num] = df
                                                    st.write(f"✅ [{t_name}] - 第 {p_num} 页提取完成！")
                                                    if global_unit == "未能自动提取，需人工核对":
                                                        unit = get_report_unit(raw_text)
                                                        if unit:
                                                            global_unit = unit
                                                            st.write(f"&nbsp;&nbsp;&nbsp;&nbsp;🔎 识别到该公司报表单位：【{global_unit}】")
                                                else:
                                                    st.write(f"⚠️ [{t_name}] - 第 {p_num} 页未提取到有效数据。")
                                            except Exception as e:
                                                st.error(f"❌ [{t_name}] - 第 {p_num} 页提取失败: {str(e)}")
                                            
                                            pages_done += 1
                                            progress_bar.progress(pages_done / total_pages_to_process)
                                
                                # 拼接和导出的代码在 spinner 内部执行
                                for table_name, pages in valid_tasks.items():
                                    table_dfs = []
                                    for p_num in pages: 
                                        if p_num in temp_results[table_name]:
                                            table_dfs.append(temp_results[table_name][p_num])
                                    
                                    if table_dfs:
                                        merged_df = pd.concat(table_dfs, ignore_index=True)
                                        safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '', table_name)[:30]
                                        extracted_sheets[safe_sheet_name] = merged_df
                                            
                                status_box.update(label="🎉 所有任务提取完成！", state="complete", expanded=False)
                                st.session_state['extracted_data'] = extracted_sheets
                                st.session_state['global_unit'] = global_unit
                            
            if 'extracted_data' in st.session_state:
                st.markdown("---")
                st.markdown("#### 📊 提取结果预览")
                extracted_data = st.session_state['extracted_data']
                company_name = st.session_state.get('pdf_name', '未命名').replace(".pdf", "")
                
                unit_df = pd.DataFrame([
                    {"项目": "源文件名称", "信息": company_name},
                    {"项目": "探测到的报表单位", "信息": st.session_state.get('global_unit', '')}
                ])
                
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    unit_df.to_excel(writer, sheet_name="基本信息_单位", index=False)
                    for sheet_name, df in extracted_data.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                excel_data = output.getvalue()
                
                st.download_button(
                    label="⬇️ 一键下载结构化提取表 (Excel)",
                    data=excel_data,
                    file_name=f"{company_name}_数据提取.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                st.markdown("")
                sheet_names = list(extracted_data.keys())
                if sheet_names:
                    tabs = st.tabs(sheet_names)
                    for i, tab in enumerate(tabs):
                        with tab:
                            st.dataframe(extracted_data[sheet_names[i]], use_container_width=True)
         
# ─────────── Step 3：目标表标准填报 ───────────
        with tab3:
            st.markdown("### 📝 目标表标准填报")
            st.markdown('<div class="info-card pink"><h4>功能说明</h4><p>AI 将自动寻找科目数据填充数值，生成 Excel 计算公式。支持内置模板或上传自定义模板。</p></div>', unsafe_allow_html=True)
            
            COMPANY_TYPE_MAP = {item["公司"]: item["类别"] for item in DEFAULT_COMPANIES}
            
            col_t1, col_t2 = st.columns([1, 1])
            with col_t1: use_default = st.toggle("使用系统默认模板", value=True, help="开启后直接使用内置的新华样例表")
            
            template_file = "https://github.com/z-xylym/my-actuary-tool/raw/refs/heads/main/%E6%96%B0%E5%8D%8E%E6%A0%B7%E4%BE%8B%E8%A1%A8.xlsx" if use_default else st.file_uploader("上传自定义目标表模板 (.xlsx)", type="xlsx", key="unique_template_uploader")
        
            if template_file:
                COL_COMPANY, COL_CATEGORY, COL_FIELD_NAME, COL_FIELD_TYPE, COL_NOTE, COL_RULE, COL_CO_TYPE = "公司", "类别", "字段名", "字段类型", "注释", "计算规则", "公司类型"
        
                if 'extracted_data' not in st.session_state:
                    st.warning("⚠️ 尚未找到提取的数据，请先完成 Step 1 和 Step 2。")
                elif st.button("启动智能填报与公式生成", use_container_width=True):
                    with st.status("此过程略慢，请稍等~", expanded=True) as status_box:
                        st.write("📄 正在获取并解析模板表结构...")
                        
                        # 1. 尝试读取文件并解析基础结构
                        try:
                            if use_default:
                                import requests, io
                                res = requests.get(template_file, timeout=15)
                                res.raise_for_status()
                                template_df = pd.read_excel(io.BytesIO(res.content))
                            else:
                                template_df = pd.read_excel(template_file)
                        except Exception as e:
                            st.error(f"❌ 读取模板失败，请检查网络连接或文件格式：{e}")
                            st.stop() # 🌟 发生异常直接强行停止，不再向下执行避免报错
                        
                        # 2. 动态识别年份列
                        year_cols = sorted([c for c in template_df.columns if __import__("re").search(r'20\d{2}', str(c))])
                        if len(year_cols) < 2:
                            st.error("❌ 错误：模板中未能自动识别出两个或以上的年份数据列！请检查表头规范。")
                            st.stop() # 🌟 年份不够也直接停止
                        
                        col_prev, col_curr = year_cols[-2], year_cols[-1]
                        st.session_state['col_prev'] = col_prev
                        st.session_state['col_curr'] = col_curr
                        
                        # 🌟 修复后的正常路径：此时已确保成功拿到 template_df, col_prev, col_curr
                        # --- 🚀 A. 公司名与公司类型自动匹配 ---
                        raw_name = st.session_state.get('pdf_name', '未命名').replace(".pdf", "")
                        company_short = __import__("re").sub(r'202\d年.*', '', raw_name).strip()
                        
                        matched_type = next((c_type for c_name, c_type in COMPANY_TYPE_MAP.items() if c_name[:2] in company_short or company_short[:2] in c_name), "其他")
                        
                        working_df = template_df.copy()
                        if COL_COMPANY in working_df.columns: working_df[COL_COMPANY] = company_short
                        if COL_CO_TYPE in working_df.columns: working_df[COL_CO_TYPE] = matched_type
                        
                        # --- B. 准备 AI 任务 ---
                        st.write("正在准备目标指标清单...")
                        input_items = working_df[working_df[COL_FIELD_TYPE].astype(str).str.strip() == "输入"]
                        ai_target_list = [{"类别": str(r.get(COL_CATEGORY, "")), "标准字段名": str(r.get(COL_FIELD_NAME, "")), "别名参考": str(r.get(COL_NOTE, "")) if pd.notna(r.get(COL_NOTE)) else ""} for _, r in input_items.drop_duplicates(subset=[COL_FIELD_NAME]).iterrows()]
                            
                        st.write("正在分析上下文...")
                        extracted_data = st.session_state['extracted_data']
                        context_text = "".join([f"\n[表名: {name}]\n{df.to_csv(index=False, sep='|')}\n" for name, df in extracted_data.items()])

                            
                        SYSTEM_PROMPT = """你是一个资深的寿险精算审计专家。任务：将 PDF 提取的财务明细精准填入目标底稿，并严格区分上一年度和最新年份。

【通用执行准则：绝对原样提取】
1. ⚠️ 绝对指令：除了下述特殊的加总需求外，所有指标必须【原封不动地照抄】原文里的文本！
   - 必须保留括号（表示负数）、逗号、正负号。例如原文是 "(295,992.00)"，必须输出为 "(295,992.00)"。
   - 严禁擅自删除符号、严禁自行进行四舍五入。
2. 别名匹配：若标准名找不到，请查看“别名参考”列，或利用你的精算知识寻找同义词。
3. 报表锁定：提取“新业务相关指标”时，务必在包含“当期确认”、“初始确认”或“新业务价值”字样的表格中查找。
4. 空值处理：若原文为“—”或“无”请输出 "0"。若完全找不到该指标，请输出 null。

【特殊指令：业务及管理费（业管费）科目拆分】
业务及管理费附注表通常包含“小计/明细”和“减项”。请按以下精算逻辑进行识别：
1. 【获取费用】：必须提取“减：与保险合同履约直接相关的支出”下方，明确标注为“计入...保险获取现金流量”或“保险获取”字样的数值。
2. 【维持费用】：必须提取“减：与保险合同履约直接相关的支出”下方，明确标注为“计入保险服务费用”或“维持费用”字样的数值。
3. 【非履约费用】：提取该附注表最底部的“合计”项数值。逻辑上，该合计 = 总支出 - 获取费用 - 维持费用。
4. 【二级明细归类】（若目标字段属于明细科目且无汇总）：
   - 【职工薪酬】：加总 工资, 奖金, 补贴, 社保, 福利, 公积金, 辞退福利等。
   - 【物业及设备】：加总 折旧, 摊销, 租赁费, 物业费, 维修费等。
   - 【业务投入及监管】：加总 宣传费, 招待费, 保障基金, 监管费, 服务费, 咨询审计费。
   - 【行政办公】：加总 差旅费, 办公费, 邮电印刷, 会议费等。

【特殊指令：投资资产表优先映射】
当目标字段属于“投资资产”类别时：
1. 优先从 [表名: 投资资产] 中寻找最接近的项目名称。
2. 若投资资产表存在对应项目，则禁止使用资产负债表数据覆盖。
3. 若投资资产表不存在该项目，再尝试从资产负债表寻找对应科目。
4. 优先依据目标表中的“别名参考”进行匹配。

【特殊指令：保险合同负债新准则科目映射（最高寻址优先级）】
为了防止名称混淆，当提取以下指标时，【必须严格锁定对应的表名】，并应用精算缩写翻译规则：
1. 锁定表名：[表名: 保险合同-未采用保费分配法按计量组成部分分析 (原保险)]
   - 映射字典：BEL = 未来现金流量现值；RA = 非金融风险调整；总额 = 合计。
   - 目标字段限定：当遇到【BEL期初余额、BEL期末余额、RA期初余额、RA期末余额、期初保险合同负债总额、期末保险合同负债总额】时，仅在此表中寻址取数。

2. 锁定表名：[表名: 保险合同-非PAA按未到期责任负债和已发生赔款负债分析 (原保险)]
   - 映射字典：LRC = 未到期责任负债；LIC = 已发生赔款负债。
   - 目标字段限定：当遇到任何带有【（非PAA）】后缀的字段（如LRC非亏损部分期初余额（非PAA）、LIC期末余额（非PAA）、非PAA期末余额合计等）时，仅在此表中寻址。

3. 锁定表名：[表名: 保险合同-PAA按未到期责任负债和已发生赔款负债分析 (原保险)]
   - 映射字典：LRC = 未到期责任负债；LIC = 已发生赔款负债。
   - 目标字段限定：当遇到任何带有【（PAA）】后缀的字段（如LRC非亏损部分期初余额（PAA）、LIC期初BEL余额（PAA）、PAA期初余额合计等）时，仅在此表中寻址。
4.【时间维度与排版防反转绝对指令（适用任意年份）】
中国企业财报的标准排版规则是：【左边的数据列为当期（最新年份），右边的数据列为上期（历史年份）】。
- 当填报要求中包含“较晚/最新年份”（如当期/本年/年末）的键时：必须从原表表头的“本期”、“本年”、“年末”或【最左侧的数据列】取数。
- 当填报要求中包含“较早/历史年份”（如上期/上年/年初）的键时：必须从原表表头的“上期”、“上年”、“年初”或【紧挨着它的右侧数据列】取数。
⚠️ 警告：大语言模型极易犯“线性思维”错误（误以为小年份排在左边，大年份排在右边）。要求你必须打破定势，仔细核对表头！永远记住：左边的列才是最新当期！
例如：若输出键包含 prev 和 curr，则 curr (最新年) 取报表左侧当期列，prev (历史年) 取报表右侧上期列。
仅输出合法的 JSON 格式，严禁带有任何 Markdown 标记或文字说明。
格式示例：{"字段名": {"prev": "(1,234.0) + 500.0", "curr": "9,876.54"}}"""
                        st.write("正在呼叫大模型进行语义映射与精准取数...")
                        
                        # 🌟 从全局内存中拿取你登录时选择的所有配置！
                        current_api_key = st.session_state.get('api_key', "").strip()
                        current_base_url = st.session_state.get('base_url', "https://api.deepseek.com")
                        current_model_name = st.session_state.get('model_name', "deepseek-chat")
                        
                        client = OpenAI(api_key=current_api_key, base_url=current_base_url)
                        try:
                            response = client.chat.completions.create(
                                model=current_model_name, # 👈 这里也变成了动态的！
                                messages=[
                                    {"role": "system", "content": SYSTEM_PROMPT},
                                    {"role": "user", "content": f"目标指标:\n{json.dumps(ai_target_list, ensure_ascii=False)}\n\n数据:\n{context_text}"}
                                ],
                                temperature=0.0
                            )
                            ai_res = response.choices[0].message.content.strip()
                            ai_res = re.sub(r'^```(json)?\n?', '', ai_res, flags=re.MULTILINE).replace("```", "").strip()
                            ai_data = json.loads(ai_res)
                            st.write("✅ 数据映射成功！正在回填表格...")
                            working_df = template_df.copy()
                            
                            raw_name = st.session_state.get('pdf_name', '未命名').replace(".pdf", "")
                            company_short = re.sub(r'202\d年.*', '', raw_name).strip()
                            
                            matched_type = "其他"
                            for c_name, c_type in COMPANY_TYPE_MAP.items():
                                if c_name[:2] in company_short or company_short[:2] in c_name:
                                    matched_type = c_type
                                    break
                            
                            if COL_COMPANY in working_df.columns:
                                working_df[COL_COMPANY] = company_short
                            
                            if COL_CO_TYPE in working_df.columns:
                                working_df[COL_CO_TYPE] = matched_type
                            
                            def process_final_val(v):
                                if v is None or str(v).lower() in ["null","none","","—","无"]:
                                    return 0.0
                            
                                s = str(v).strip()
                            
                                def to_calc_str(x):
                                    tmp = str(x).replace(',','').strip()
                                    if (tmp.startswith('(') and tmp.endswith(')')) or (tmp.startswith('（') and tmp.endswith('）')):
                                        return "-" + tmp[1:-1].strip()
                                    return tmp
                            
                                if "+" in s:
                                    return "=" + "+".join([to_calc_str(i) for i in s.split("+")])
                            
                                try:
                                    return float(to_calc_str(s))
                                except:
                                    return s
                            
                            for idx,row in working_df.iterrows():
                                fname = str(row.get(COL_FIELD_NAME,"")).strip()
                            
                                if fname not in ai_data:
                                    continue
                            
                                item_data = ai_data[fname]
                            
                                if isinstance(item_data,dict):
                                    working_df.at[idx,col_prev] = process_final_val(item_data.get("prev"))
                                    working_df.at[idx,col_curr] = process_final_val(item_data.get("curr"))
                                else:
                                    working_df.at[idx,col_prev] = 0.0
                                    working_df.at[idx,col_curr] = 0.0
                            
                            st.write("⚙️ 正在生成可追溯的单元格公式...")
                            
                            temp_buffer = io.BytesIO()
                            working_df.to_excel(temp_buffer,index=False)
                            temp_buffer.seek(0)
                            
                            wb = openpyxl.load_workbook(temp_buffer)
                            ws = wb.active
                            
                            cols = {cell.value:cell.column for cell in ws[1]}
                            
                            prev_idx = cols.get(col_prev)
                            curr_idx = cols.get(col_curr)
                            
                            for r in range(2,ws.max_row+1):
                                for c_idx in [prev_idx,curr_idx]:
                                    if not c_idx:
                                        continue
                                    cell = ws.cell(r,c_idx)
                                    if isinstance(cell.value,str) and cell.value.startswith("="):
                                        cell.value = cell.value
                            
                            field_to_row = {
                                str(ws.cell(r,cols[COL_FIELD_NAME]).value).strip():r
                                for r in range(2,ws.max_row+1)
                                if ws.cell(r,cols[COL_FIELD_NAME]).value
                            }
                            
                            sorted_fields = sorted(field_to_row.keys(),key=len,reverse=True)
                            
                            prev_letter = get_column_letter(prev_idx) if prev_idx else ""
                            curr_letter = get_column_letter(curr_idx) if curr_idx else ""
                            
                            for r in range(2,ws.max_row+1):
                            
                                ftype = str(ws.cell(r,cols[COL_FIELD_TYPE]).value).strip()
                                rule = str(ws.cell(r,cols[COL_RULE]).value).strip()
                            
                                if ftype!="计算" or rule in ["nan","None",""]:
                                    continue
                            
                                prev_formula = rule
                                curr_formula = rule
                            
                                for f in sorted_fields:
                                    if f in rule:
                                        row_num = field_to_row[f]
                            
                                        if prev_letter:
                                            prev_formula = prev_formula.replace(f,f"{prev_letter}{row_num}")
                            
                                        if curr_letter:
                                            curr_formula = curr_formula.replace(f,f"{curr_letter}{row_num}")
                            
                                try:
                                    if prev_idx:
                                        ws.cell(r,prev_idx).value = f"={prev_formula}"
                            
                                    if curr_idx:
                                        ws.cell(r,curr_idx).value = f"={curr_formula}"
                                except:
                                    pass
                            
                            for r in range(2,ws.max_row+1):
                                for c_idx in [prev_idx,curr_idx]:
                            
                                    if not c_idx:
                                        continue
                            
                                    cell = ws.cell(r,c_idx)
                                    cell.number_format = '#,##0_ ;(#,##0);"-"'
                            
                                    if cell.value and not str(cell.value).startswith("="):
                                        try:
                                            clean_val = str(cell.value).replace(',','').replace('(','-').replace(')','')
                                            cell.value = float(clean_val)
                                        except:
                                            pass
                            
                            final_buffer = io.BytesIO()
                            wb.save(final_buffer)
                            final_buffer.seek(0)
                            
                            st.session_state['filled_excel'] = final_buffer.getvalue()
                            st.session_state['final_df'] = working_df
                            st.session_state['col_prev'] = col_prev
                            st.session_state['col_curr'] = col_curr

                        except Exception as e:
                            st.error(f"❌ 流程中断: {str(e)}")
                            status_box.update(label="处理失败", state="error", expanded=True)

            if 'filled_excel' in st.session_state:
                st.markdown("---")
                st.markdown("#### 📋 智能填报结果预览")
                company_name = st.session_state.get('pdf_name', '未命名').replace(".pdf", "")
                
                st.download_button(
                    label="⬇️ 下载已填报的底稿 (含公式)",
                    data=st.session_state['filled_excel'],
                    file_name=f"{company_name}_自动填报表.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.markdown("💡 *预览表仅展示数值，下载后的 Excel 已包含自动关联公式。*")
                st.dataframe(st.session_state['final_df'], use_container_width=True)

        # ─────────── Step 4 分析与校验 ───────────
        with tab4:
            # 🎨 注入自定义 CSS，强制缩小按钮和 Popover 的尺寸
            st.markdown("""
                <style>
                /* 缩小 Popover 按钮 */
                div[data-testid="stPopover"] > button {
                    padding: 2px 8px !important;
                    min-height: 26px !important;
                    height: 26px !important;
                    font-size: 12px !important;
                    background-color: #f8f9fa !important;
                    border: 1px solid #ddd !important;
                }
                /* 缩小并修正刷新按钮 */
                div[data-testid="stColumn"] button[kind="secondary"] {
                    padding: 2px 10px !important; /* 左右间距稍微大一点点 */
                    height: 26px !important;
                    font-size: 12px !important;
                    white-space: nowrap !important; /* 强制文字不换行 */
                    min-width: 70px !important;    /* 保证按钮有最小宽度容纳 🔄刷新 */
                    display: flex !important;
                    align-items: center !important;
                    justify-content: center !important;
                }
                /* 调整 subheader 间距 */
                .stSubheader { margin-bottom: -10px !important; }
                </style>
            """, unsafe_allow_html=True)
        
            # 1. 顶部工具栏布局
            # 调整比例，新增一个列用来放下载按钮
            col_title, col_btn_sample, col_btn_refresh, col_manual_area = st.columns([3.5, 1.8, 1.1, 1.8])
            
            with col_title:
                st.subheader("🏁 精算数据勾稽关系检查")
                
            with col_btn_sample:
                # 使用弹窗容纳下载按钮，保持界面 CSS 高度一致
                with st.popover("校验表获取", use_container_width=True):
                    st.caption("将人工校验两列复制到需要校对的表格对应位置")
                    
                    # 🌟 修复：改为从 Github 链接动态获取文件
                    import requests
                    url = "https://github.com/z-xylym/my-actuary-tool/raw/refs/heads/main/%E6%96%B0%E5%8D%8E%E6%A0%B7%E4%BE%8B%E8%A1%A8.xlsx"
                    
                    try:
                        # 添加一个 spinner 防止网络慢的时候用户以为卡死了
                        with st.spinner("获取中..."):
                            response = requests.get(url, timeout=10)
                            response.raise_for_status()  # 检查请求是否成功
                            
                        # 把网络请求拿到的二进制内容直接塞给下载按钮
                        st.download_button(
                            label="⬇️ 下载样例文件", 
                            data=response.content, 
                            file_name="新华人工校验样例.xlsx", 
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    except Exception as e:
                        st.error("网络请求失败，请检查网络或 GitHub 链接是否有效！")
            
            with col_btn_refresh:
                if st.button("🔄 刷新", use_container_width=True):
                    st.rerun() 
        
            # --- 📖 核对重点与语音 (合二为一的区域) ---
            with col_manual_area:
                # 在这一个列里再分两个超微型列，实现“字标同行”
                c1, c2 = st.columns([1, 0.4])
                
                # --- 📖 操作手册弹出框 ---
                manual_content = """
                精算数据人工核对指南
                
                1. 模板准备： 请将新华样例中的“人工校验”两列手动复制到当前需要核对的表格中。
                2. 年度对齐： 重点检查 最新年 和 上一年 的数据是否有填反现象。
                3. PAA 计算： 采用保费分配法保险业绩 = 保费分配法下的服务收入 - 服务费用。
                4. 符号修正： 部分公司利润表营业支出披露为正，需人工对每一项添加负号（原本负号变正）。
                5. 费用拆解： 业务及管理费中，职工薪酬等项有时需人工尝试加总核对，维持费用、获取费用应为正数，保证进出项总和相等。
                6. 外部搜索： 折现率假设、非金融风险置信水平需手动在 PDF 中全文搜索并填入。
                7. CSM 摊销： 1-20年合同服务边际摊销之和应等于 1。请查看“未来现金流现值、风险调整、CSM”页下方注释，或手动计算CSM摊销比例。
                """
        
                with c1:
                    with st.popover("🕶️核对指南", use_container_width=True):
                        # 💡 这里的 HTML 字符串必须紧贴左侧，或者确保没有多余的缩进空格
                        html_code = f"""
        <div style="font-family: 'Microsoft YaHei', sans-serif; padding: 5px;">
            <h4 style="color: #1f4e79; margin: 0 0 10px 0; font-size: 1rem; border-bottom: 2px solid #e1e4e8; padding-bottom: 5px;">
                精算数据人工核对指南
            </h4>
            <div style="line-height: 1.6; font-size: 0.85rem; color: #333333;">
                <p style="margin-bottom: 8px;">
                    <strong style="color: #d32f2f;">1. 模板准备：</strong> 
                    将新华样例的“人工校验”两列手动复制到核对表中。
                </p>
                <p style="margin-bottom: 8px;">
                    <strong style="color: #d32f2f;">2. 年度对齐：</strong> 
                    检查 <span style="background-color: #fff3cd; padding: 1px 3px;">去年</span> 和 <span style="background-color: #fff3cd; padding: 1px 3px;">最新年</span> 数据是否有填反。
                </p>
                <p style="margin-bottom: 8px;">
                    <strong style="color: #d32f2f;">3. PAA 计算：</strong> 
                    采用保费分配法保险业绩 = 保费分配法下的服务收入 - 服务费用。
                </p>
                <p style="margin-bottom: 8px;">
                    <strong style="color: #d32f2f;">4. 符号修正：</strong> 
                    利润表营业支出里的项目如为正，需人工加负号（原本负号变正）。
                </p>
                <p style="margin-bottom: 8px;">
                    <strong style="color: #d32f2f;">5. 费用拆解：</strong> 
                    业管费分类需人工加总，维持费用和获取费用为正，确保进出项总和相等。
                </p>
                <p style="margin-bottom: 8px;">
                    <strong style="color: #d32f2f;">6. 外部搜索：</strong> 
                    折现率和置信水平需在 PDF 中全文搜索填入。
                </p>
                <p style="margin-bottom: 0;">
                    <strong style="color: #d32f2f;">7. CSM 摊销：</strong> 
                    1-20年摊销之和应等于 1。查看合同服务边际页下注释或手动计算摊销比例。
                </p>
            </div>
        </div>
        """
                        # ⚠️ 关键点：unsafe_allow_html=True 必须有，且 html_code 前后不要有导致 Markdown 识别为代码块的空格
                        st.markdown(html_code, unsafe_allow_html=True)
                with c2:
                    # 清洗语音文本
                    clean_voice_text = manual_content.replace('\n', '。').replace('**', '')
                    tts_html = f"""
                    <div style="display: flex; align-items: center; height: 26px;">
                        <button id="tts-btn" onclick="toggleSpeak()" style="
                            border: 1px solid #ddd;
                            background-color: #f8f9fa;
                            border-radius: 4px;
                            cursor: pointer;
                            width: 26px;
                            height: 26px;
                            display: flex;
                            align-items: center;
                            justify-content: center;
                            font-size: 12px;
                        ">📢</button>
                    </div>
                    <script>
                    var msg = new SpeechSynthesisUtterance();
                    msg.text = "{clean_voice_text}";
                    msg.lang = 'zh-CN';
                    msg.rate = 1.2;
                    function toggleSpeak() {{
                        const btn = document.getElementById('tts-btn');
                        if (window.speechSynthesis.speaking) {{
                            window.speechSynthesis.cancel();
                            btn.style.backgroundColor = "#f8f9fa";
                            btn.innerText = "📢";
                        }} else {{
                            window.speechSynthesis.speak(msg);
                            btn.style.backgroundColor = "#ffebeb"; 
                            btn.innerText = "⏹";
                            msg.onend = function() {{
                                btn.style.backgroundColor = "#f8f9fa";
                                btn.innerText = "📢";
                            }};
                        }}
                    }}
                    </script>
                    """
                    st.components.v1.html(tts_html, height=30)
        
                    
            if 'final_df' in st.session_state:
                df = st.session_state['final_df'].copy()
                
                with st.expander("🛠️ 勾稽规则配置 (点击展开/修改公式)"):
                    st.info("curr代表当前年度数据；prev_year代表上一年度；curr_year代表当前年度。")
                    rules_df = st.data_editor(pd.DataFrame(DEFAULT_RULES), num_rows="dynamic", key="rules_editor_v3")
        
                    target_prev = str(st.session_state.get('col_prev'))
                    target_curr = str(st.session_state.get('col_curr'))
                    
                    col_prev_name = next(
                        (col for col in df.columns if str(target_prev) in str(col)),
                        None
                    )
                    
                    col_curr_name = next(
                        (col for col in df.columns if str(target_curr) in str(col)),
                        None
                    )
                    
                    if not col_prev_name or not col_curr_name:
                        st.error(
                            f"❌ 表格中找不到包含 '{target_prev}' 和 '{target_curr}' 的列。"
                        )
                        st.stop()
        
                def parse_finance_num(val):
                    if pd.isnull(val): return 0.0
                    v_str = str(val).strip()
                    if v_str in ['-', '']: return 0.0
                    if v_str.startswith('(') and v_str.endswith(')'): v_str = '-' + v_str[1:-1]
                    v_str = v_str.replace(',', '')
                    try: return float(v_str)
                    except ValueError: return 0.0
        
                # —— 💡 黑科技黑科技：智能容错字典 —— 
                class SmartDict(dict):
                    """一个无视空格、全半角冒号/括号的字典"""
                    def _clean(self, k):
                        if not isinstance(k, str): return k
                        # 统一转成英文标点，并删除所有空格、换行
                        return k.replace('（','(').replace('）',')').replace('：',':').replace(' ','').replace('\n','').replace('\xa0','')
                    
                    def __init__(self, mapping):
                        super().__init__(mapping)
                        # 建立 [清洗后的名字 : 原名] 的映射
                        self._map = {self._clean(k): k for k in mapping.keys()}
                        
                    def __getitem__(self, k):
                        clean_k = self._clean(k)
                        if clean_k in self._map:
                            return super().__getitem__(self._map[clean_k])
                        raise KeyError(k) # 如果连清洗后都找不到，才报错原名
                # ————————————————————————————————
        
                # 3. 提取数据，并使用 SmartDict 包装
                base_prev = {
                    str(k).strip(): parse_finance_num(v)
                    for k, v in zip(df['字段名'], df[col_prev_name])
                }
                
                base_curr = {
                    str(k).strip(): parse_finance_num(v)
                    for k, v in zip(df['字段名'], df[col_curr_name])
                }
                
                prev_year_map = SmartDict(base_prev)
                curr_year_map = SmartDict(base_curr)
                
                check_results = []
                
                # 4. 执行校验与智能诊断
                for _, rule in rules_df.iterrows():
                    rule_name = str(rule.get('规则名称', ''))
                    formula = str(rule.get('公式', ''))
                    rule_type = str(rule.get('类型', 'single')).strip().lower()
                
                    targets = (
                        [(f"{target_prev}-{target_curr}", None)]
                        if rule_type == "cross"
                        else [
                            (target_prev, prev_year_map),
                            (target_curr, curr_year_map)
                        ]
                    )
                
                    for year_label, curr_map in targets:
                
                        ctx = {
                            'curr': curr_map,
                            'prev': prev_year_map,
                            'prev_year': prev_year_map,
                            'curr_year': curr_year_map,
                            'abs': abs,
                            'min': min,
                            'max': max,
                            'round': round
                        }
                
                        try:
                            is_pass = bool(eval(formula, {"__builtins__": None}, ctx))
                            status = "✅ PASS" if is_pass else "❌ FALSE"
                            detail = "勾稽无误" if is_pass else "差额超过允许阈值"
                
                        except KeyError as e:
                            status = "⚠️ ERROR"
                            detail = f"缺失字段: {str(e)}"
                
                        except Exception as e:
                            status = "⚠️ ERROR"
                            detail = f"公式错误: {str(e)}"
                
                        check_results.append({
                            "年度": year_label,
                            "规则名称": rule_name,
                            "检查结果": status,
                            "诊断详情": detail
                        })
        
                # 5. 展示结果报表
                res_display_df = pd.DataFrame(check_results)
                
                def color_status(val):
                    if '✅' in str(val): return 'background-color: #C6EFCE; color: #006100;'
                    elif '❌' in str(val): return 'background-color: #FFC7CE; color: #9C0006;'
                    elif '⚠️' in str(val): return 'background-color: #FFEB9C; color: #9C5700;'
                    return ''
        
                st.dataframe(
                    res_display_df.style.map(color_status, subset=['检查结果']),
                    use_container_width=True, hide_index=True
                )
        
                # 6. 导出逻辑（接力模式：确保保留 Step 3 的所有原始公式）
                if st.button("📥 导出带勾稽结果的报告"):
                    if 'filled_excel' not in st.session_state:
                        st.error("❌ 找不到填报后的数据，请先重新运行 Step 3")
                        st.stop()
                        
                    # --- 💡 接力核心：从 session_state 加载 Step 3 已经写好的 Excel ---
                    input_buffer = io.BytesIO(st.session_state['filled_excel'])
                    workbook = openpyxl.load_workbook(input_buffer)
                    
                    # --- 💡 动作 1：处理数据明细页的格式 ---
                    # 这一步是为了确保即便 Step 3 漏了格式，这里也能补上
                    ws_data = workbook.active  # 默认第一个 Sheet 是数据明细
                    ws_data.title = "数据明细" # 统一命名
                    
                    # --- 💡 动作 2：新建勾稽报告页 ---
                    # 如果已存在则删除，防止重复点击报错
                    if "勾稽报告" in workbook.sheetnames:
                        del workbook["勾稽报告"]
                    ws_res = workbook.create_sheet("勾稽报告")
                    
                    # 写入勾稽报告表头
                    headers = list(res_display_df.columns)
                    for c_idx, header in enumerate(headers, 1):
                        ws_res.cell(row=1, column=c_idx).value = header
                    
                    # 写入勾稽报告内容
                    from openpyxl.styles import PatternFill
                    fill_green = PatternFill(start_color='C6EFCE', fill_type='solid')
                    fill_red = PatternFill(start_color='FFC7CE', fill_type='solid')
                    fill_yellow = PatternFill(start_color='FFEB9C', fill_type='solid')
                    
                    for r_idx, row_data in enumerate(res_display_df.values, 2):
                        for c_idx, value in enumerate(row_data, 1):
                            cell = ws_res.cell(row=r_idx, column=c_idx)
                            cell.value = value
                            # 在结果列（第3列）加颜色
                            if c_idx == 3:
                                val_str = str(value)
                                if "PASS" in val_str: cell.fill = fill_green
                                elif "FALSE" in val_str: cell.fill = fill_red
                                elif "ERROR" in val_str: cell.fill = fill_yellow
                    
                    # 自动调整列宽
                    ws_res.column_dimensions['B'].width = 30
                    ws_res.column_dimensions['C'].width = 15
                    ws_res.column_dimensions['D'].width = 40
        
                    # --- 💡 动作 3：保存并导出 ---
                    output = io.BytesIO()
                    workbook.save(output)
                    
                    st.download_button(
                        label="点击下载精算复核底稿", 
                        data=output.getvalue(), 
                        file_name=f"勾稽复核底稿_{st.session_state.get('pdf_name','未命名').replace('.pdf','')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.info("💡 提取结果将在此处显示。请先在上方点击“开始提取”按钮。")
                
#------step5----------------
        with tab5:
            col_title, col_btn = st.columns([4, 1])
            with col_title:
                st.subheader("⛓️‍💥 多公司数据集成与汇率/单位转换")
            st.info("功能说明：支持上传单文件多Sheet或多文件。系统将自动提取所有公司，请在下方为不同公司配置对应的汇率和原始表格的单位。")
        
            uploaded_files = st.file_uploader("请上传已完成勾稽检查的底稿 (支持多文件或单文件多Sheet)", type="xlsx", accept_multiple_files=True)
        
            @st.cache_data(show_spinner=False)
            def run_data_integration(temp_data_list, rate_cfg, unit_cfg):
                exact_exempt_fields = ["折现率假设","非金融风险的置信水平","1年及1年以内合同服务边际","1-5年合同服务边际","5-10年合同服务边际","10-20年合同服务边际","20年合同服务边际","投资收益率","综合偿付能力充足率"]
                combined_list = []
                def clean_to_float(val):
                    try:
                        if isinstance(val, (int, float)): return float(val)
                        if isinstance(val, str):
                            val = val.replace(',','').replace('(','-').replace(')','').strip()
                            return 0.0 if val in ['-',''] else float(val)
                        return 0.0
                    except: return 0.0
        
                for item in temp_data_list:
                    df_single, comp_name = item["df"], item["comp"]
                    rate, unit_mult = rate_cfg[comp_name], unit_cfg[comp_name]
        
                 
                    year_cols = {}
                    for c in df_single.columns:
                        import re
                        m = re.search(r'(20\d{2})', str(c))
                        if m: year_cols[m.group(1)] = c
        
                    if len(year_cols) < 1: continue
        
                    base_cols = ["公司类型","公司","类别","字段名","字段类型"]
                    existing_base = [c for c in base_cols if c in df_single.columns]
        
                    for year_label, col_name in year_cols.items():
                        df_year = df_single[existing_base + [col_name]].copy()
                        df_year["公司"] = comp_name
                        df_year["报告年份"] = year_label
                        df_year["汇率"] = rate
                        df_year["(百万)原币"] = None
                        df_year["(百万)人民币"] = None
                        df_year["汇率"] = df_year["汇率"].astype(object)
        
                        for idx in df_year.index:
                            raw_val = df_year.loc[idx, col_name]
                            f_name = str(df_year.loc[idx, "字段名"]).strip()
                            if f_name in exact_exempt_fields:
                                if "折现率" in f_name or "非金融风险的置信水平" in f_name:
                                    df_year.at[idx,"(百万)原币"] = str(raw_val) if pd.notna(raw_val) else ""
                                    df_year.at[idx,"(百万)人民币"] = str(raw_val) if pd.notna(raw_val) else ""
                                else:
                                    dec_val = float(raw_val.replace('%','').strip())/100.0 if isinstance(raw_val,str) and '%' in raw_val else clean_to_float(raw_val)
                                    df_year.at[idx,"(百万)原币"] = dec_val
                                    df_year.at[idx,"(百万)人民币"] = dec_val
                                df_year.at[idx,"汇率"] = "豁免换算"
                            else:
                                c_val = clean_to_float(raw_val)
                                df_year.at[idx,"(百万)原币"] = c_val * unit_mult
                                df_year.at[idx,"(百万)人民币"] = c_val * unit_mult * rate
        
                        final_cols = ["公司类型","公司","类别","字段名","字段类型","报告年份","(百万)原币","汇率","(百万)人民币"]
                        actual_cols = [c for c in final_cols if c in df_year.columns]
                        combined_list.append(df_year[actual_cols])
        
                return pd.concat(combined_list, ignore_index=True) if combined_list else pd.DataFrame()
        
            # ✅ 从列名自动识别单位的辅助函数
            def detect_unit_from_col(col_name):
                """从列名里读取单位关键词，返回对应的下拉选项文字"""
                s = str(col_name)
                if "十亿" in s: return "原表为: 十亿元 (× 1,000)"
                if "亿元" in s: return "原表为: 亿元 (× 100)"
                if "百万" in s: return "原表为: 百万元 (无需转换)"
                if "万元" in s or "万" in s: return "原表为: 万元 (÷ 100)"
                if "千元" in s: return "原表为: 千元 (÷ 1,000)"
                if "元" in s:   return "原表为: 元 (÷ 1,000,000)"
                return "原表为: 百万元 (无需转换)"  # 默认
        
            unit_multipliers = {
                "原表为: 百万元 (无需转换)": 1.0,
                "原表为: 万元 (÷ 100)": 0.01,
                "原表为: 元 (÷ 1,000,000)": 0.000001,
                "原表为: 千元 (÷ 1,000)": 0.001,
                "原表为: 亿元 (× 100)": 100.0,
                "原表为: 十亿元 (× 1,000)": 1000.0,
            }
        
            if uploaded_files:
                all_temp_data, found_companies = [], {}
        
                for file in uploaded_files:
                    xl = pd.ExcelFile(file)
                    for sheet_name in xl.sheet_names:
                        df_raw = pd.read_excel(file, sheet_name=sheet_name)
                        current_company = str(df_raw["公司"].iloc[0]) if "公司" in df_raw.columns and not df_raw["公司"].empty else sheet_name
                        if "基本信息" in current_company or "Sheet" in current_company: continue
        
                        # ✅ 从列名自动检测默认单位
                        import re
                        detected_unit = "原表为: 百万元 (无需转换)"
                        for c in df_raw.columns:
                            if re.search(r'20\d{2}', str(c)):
                                detected_unit = detect_unit_from_col(c)
                                break
        
                        found_companies[current_company] = detected_unit
                        all_temp_data.append({"comp": current_company, "df": df_raw, "source": f"{file.name} - {sheet_name}"})
        
                st.markdown("#### 💵 汇率与数值单位配置盘")
                st.caption("目标表统一要求以【百万元人民币】展示。系统已根据列名自动识别默认单位，请人工核对后调整。")
        
                rate_config, unit_config = {}, {}
                rate_cols = st.columns(3)
                for i, (comp, default_unit) in enumerate(sorted(found_companies.items())):
                    with rate_cols[i % 3]:
                        with st.container(border=True):
                            st.markdown(f"**🏢 {comp}**")
                            rate_config[comp] = st.number_input("汇率 (相对于RMB)", value=1.0, step=0.0001, format="%.4f", key=f"rate_{comp}")
                            # ✅ 默认值用自动检测的单位
                            unit_options = list(unit_multipliers.keys())
                            default_idx = unit_options.index(default_unit) if default_unit in unit_options else 0
                            unit_choice = st.selectbox("原表数值单位", unit_options, index=default_idx, key=f"unit_{comp}")
                            unit_config[comp] = unit_multipliers[unit_choice]
        
                if st.button("开始集成并换算数据", type="primary", use_container_width=True):
                    with st.spinner("正在后台执行极速数据合并与换算..."):
                        final_all_df = run_data_integration(all_temp_data, rate_config, unit_config)
        
                    if not final_all_df.empty:
                        st.session_state['integrated_data'] = final_all_df
                        years_found = sorted(final_all_df['报告年份'].unique().tolist())
                        st.success(f"✅ 集成与换算完毕！共处理 {len(found_companies)} 家公司，识别到年份：{' / '.join(years_found)}，生成 {len(final_all_df)} 条对标数据。")
                        st.dataframe(final_all_df, use_container_width=True)
        
                        import io
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            final_all_df.to_excel(writer, index=False, sheet_name='行业集成分析表')
                        st.download_button(label="下载行业集成目标表 (长表格式)", data=output.getvalue(), file_name="行业集成目标数据表.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    else:
                        st.error("未能从上传的文件中提取到有效数据，请检查列名中是否包含年份信息（如 2024YE、2025YE 等）。")


        # ─────────── KPMG 官方色卡 ───────────
        KPMG_CATEGORIES = {
            "Primary Colors": {"KPMG Blue": "#00338D", "Cobalt Blue": "#1E49E2", "Dark Blue": "#0C233C", "Light Blue": "#ACEAFF", "Pacific Blue": "#00B8F5", "Purple": "#7213EA", "Pink": "#FD349C"},
            "Accent Colors": {"Blue": "#76D2FF", "Dark Purple": "#510DBC", "Light Purple": "#B497FF", "Dark Pink": "#AB0D82", "Light Pink": "#FFA3DA", "Dark Green": "#098E7E", "Green": "#00C0AE", "Light Green": "#63EBB2"},
            "Traffic Light": {"Red": "#ED2124", "Amber": "#F1C44D", "Positive Green": "#269924"}
        }
        DEFAULT_COLORS = list(KPMG_CATEGORIES["Primary Colors"].values()) + list(KPMG_CATEGORIES["Accent Colors"].values())
        
        # ─────────── Step 6 可视化分析面板 ───────────
        with tab6:
            st.markdown("### 📊 自定义对标分析")
        
            # ── 数据源选择 ───────────────────────────────────────────────────────
            if 'integrated_data' not in st.session_state:
                st.session_state['integrated_data'] = None
        
            source_choice = st.radio(
                "数据源选择", ["直接引用集成后的数据", "上传集成表 Excel"], horizontal=True
            )
            df_raw = None
        
            if source_choice == "上传集成表 Excel":
                viz_file = st.file_uploader("上传行业集成目标表", type=["xlsx"])
                if viz_file:
                    df_raw = pd.read_excel(viz_file)
                    st.session_state['integrated_data'] = df_raw
            else:
                df_raw = st.session_state.get('integrated_data')
        
            if df_raw is None:
                st.info("💡 请先完成数据集成或上传目标底稿。")
                st.stop()
        
            # ==========================================
            # 🌟 提速秘籍 1：锁定数据预处理 (只算一次)
            # ==========================================
            @st.cache_data(show_spinner=False)
            def prepare_viz_data(df_in):
                df_clean = df_in.copy()
                df_clean.columns = (
                    df_clean.columns.astype(str)
                    .str.strip()
                    .str.replace('\n', '', regex=False)
                    .str.replace('\r', '', regex=False)
                    .str.replace('\ufeff', '', regex=False)
                )
                df_clean['报告年份'] = df_clean['报告年份'].astype(str).str.replace('.0', '', regex=False)
                val_col = "(百万)人民币" if "(百万)人民币" in df_clean.columns else df_clean.columns[-1]
        
                # 去重
                dedup_cols = ['公司', '报告年份', '字段名']
                if '类别' in df_clean.columns:
                    dedup_cols.append('类别')
                df_clean = df_clean.drop_duplicates(subset=dedup_cols, keep='first')
        
                # 透视表
                df_pivot = (
                    df_clean.groupby(['公司', '报告年份', '字段名'])[val_col]
                    .sum().unstack('字段名').reset_index()
                )
        
                # 提取选项列表
                all_fields   = sorted(df_clean['字段名'].unique().tolist())
                all_types    = sorted(df_clean['类别'].dropna().astype(str).unique().tolist()) if '类别'    in df_clean.columns else []
                all_co_types = sorted(df_clean['公司类型'].dropna().astype(str).unique().tolist()) if '公司类型' in df_clean.columns else []
        
                all_years_sorted = sorted(
                    [y for y in df_clean['报告年份'].unique() if str(y).isdigit()],
                    key=lambda x: int(x)
                )
                
                return df_clean, df_pivot, val_col, all_fields, all_types, all_co_types, all_years_sorted
        
            # 调用缓存的数据处理函数
            df_clean, df_pivot, val_col, all_fields, all_types, all_co_types, all_years_sorted = prepare_viz_data(df_raw)
        
            # ── 动态年份颜色计算 (极速，无需缓存) ──────────────────────────────────
            KPMG_COLOR_MAP = {
                "Lightest": "#BFE8FF",     # Pacific Blue (light)
                "Light": "#76D2FF",      # Blue (original Light Blue)
                "Primary": "#1E49E2",     # Cobalt Blue
                "Dark": "#00338D",       # KPMG Blue (original Primary Blue)
            }
            
            n_years = len(all_years_sorted)
            dynamic_year_colors = {}
            target_colors = []
            
            if n_years == 3:
                target_colors = [KPMG_COLOR_MAP["Light"], KPMG_COLOR_MAP["Primary"], KPMG_COLOR_MAP["Dark"]]
            elif n_years == 2:
                target_colors = [KPMG_COLOR_MAP["Primary"], KPMG_COLOR_MAP["Dark"]]
            else:
                target_colors = [KPMG_COLOR_MAP["Lightest"], KPMG_COLOR_MAP["Light"], KPMG_COLOR_MAP["Primary"], KPMG_COLOR_MAP["Dark"]]
            
            num_colors_to_use = len(target_colors)
            if n_years > 0:
                for idx, year in enumerate(all_years_sorted):
                    color_index = 0
                    if n_years > 1:
                        proportion = idx / (n_years - 1)
                        color_index = int(proportion * (num_colors_to_use - 1))
                    color_index = min(color_index, num_colors_to_use - 1)
                    dynamic_year_colors[str(year)] = target_colors[color_index]
        
            # ── KPMG 色卡 ────────────────────────────────────────────────────────
            with st.expander("🎨 查看 KPMG 官方色卡"):
                # 确保全局存在 KPMG_CATEGORIES 变量
                if 'KPMG_CATEGORIES' in globals():
                    for cat_name, cat_colors in KPMG_CATEGORIES.items():
                        st.markdown(f"**{cat_name}**")
                        html_str = "".join([
                            f'<div style="display:inline-block;margin-right:15px;margin-bottom:8px;">'
                            f'<div style="width:14px;height:14px;background-color:{c};display:inline-block;'
                            f'border-radius:3px;vertical-align:middle;border:1px solid #ddd;"></div>'
                            f'<span style="font-size:13px;vertical-align:middle;"> {n} <b>({c})</b></span></div>'
                            for n, c in cat_colors.items()
                        ])
                        st.markdown(html_str, unsafe_allow_html=True)
                else:
                    st.caption("未检测到 KPMG_CATEGORIES 色卡定义。")
        
            st.divider()
        
            # ==========================================
            # 🌟 提速秘籍 2：锁定 UI 与绘图更新范围 (片段刷新)
            # ==========================================
            @st.fragment
            def render_viz_dashboard(df_clean, df_pivot, val_col, all_fields, all_types, all_co_types, dynamic_year_colors):
                # ── 核心配置面板 ─────────────────────────────────────────────────────
                with st.expander("🛠️ 核心配置面板", expanded=True):
                    r1c1, r1c2, r1c3 = st.columns([1.5, 1, 1])
                    is_pct_stack_mode = False
        
                    with r1c1:
                        chart_type = st.selectbox(
                            "📈 1. 图表类型",
                            ["簇状柱状图", "堆积柱状图", "折线对比图", "饼图", "内外环结构对比图"]
                        )
        
                        # ── 饼图 / 环图 ──────────────────────────────────────────────
                        if "环" in chart_type or chart_type == "饼图":
                            calc_mode = "结构分析"
                            selected_multi_fields = st.multiselect(
                                "🎯 选择构成指标", all_fields,
                                default=all_fields[:min(3, len(all_fields))]
                            )
                            selected_cos = st.selectbox("🏗️ 选择展示公司", sorted(df_pivot['公司'].unique().tolist()))
                            plot_df_base = (
                                df_clean[
                                    df_clean['字段名'].isin(selected_multi_fields) &
                                    (df_clean['公司'] == selected_cos)
                                ].copy().rename(columns={val_col: 'final_val'})
                            )
        
                        else:
                            # ── 堆积图子模式 ─────────────────────────────────────────
                            if chart_type == "堆积柱状图":
                                stack_sub = st.radio(
                                    "堆积模式", ["单指标 / 公式", "多指标占比"],
                                    horizontal=True, key="stack_sub_mode"
                                )
                                is_pct_stack_mode = (stack_sub == "多指标占比")
        
                            # ── 多指标占比 ───────────────────────────────────────────
                            if is_pct_stack_mode:
                                calc_mode = "多指标占比"
                                if all_types:
                                    pct_type_f = st.selectbox(
                                        "🗂️ 按类别筛选指标", ["全部类别"] + all_types, key="pct_type_filter"
                                    )
                                    pct_field_pool = (
                                        sorted(df_clean[df_clean['类别'] == pct_type_f]['字段名'].unique().tolist())
                                        if pct_type_f != "全部类别" else all_fields
                                    )
                                else:
                                    pct_field_pool = all_fields
        
                                selected_pct_fields = st.multiselect(
                                    "🎯 选择堆积指标", pct_field_pool,
                                    default=pct_field_pool[:min(3, len(pct_field_pool))],
                                    key="pct_stack_fields"
                                )
                                if selected_pct_fields:
                                    plot_df_base = (
                                        df_clean[df_clean['字段名'].isin(selected_pct_fields)]
                                        [['公司', '报告年份', '字段名', val_col]]
                                        .copy().rename(columns={val_col: 'final_val'})
                                    )
                                else:
                                    plot_df_base = pd.DataFrame(columns=['公司', '报告年份', '字段名', 'final_val'])
        
                            # ── 普通单指标 / 公式 ────────────────────────────────────
                            else:
                                calc_mode = st.radio("数据模式", ["单指标直显", "自定义公式运算"], horizontal=True)
        
                                if calc_mode == "单指标直显":
                                    if all_types:
                                        type_f = st.selectbox(
                                            "🗂️ 按类别筛选指标", ["全部类别"] + all_types, key="type_filter_single"
                                        )
                                        filtered_fields = (
                                            sorted(df_clean[df_clean['类别'] == type_f]['字段名'].unique().tolist())
                                            if type_f != "全部类别" else all_fields
                                        )
                                    else:
                                        filtered_fields = all_fields
        
                                    target_field = st.selectbox("🎯 选择显示指标", filtered_fields)
                                    plot_df_base = (
                                        df_pivot[['公司', '报告年份', target_field]]
                                        .rename(columns={target_field: 'final_val'}).fillna(0)
                                    )
        
                                else:  # 公式模式
                                    v1, v2 = st.columns(2)
                                    var_a = v1.selectbox("变量 A", all_fields, index=0)
                                    var_b = v2.selectbox("变量 B", ["无"] + all_fields, index=1 if len(all_fields) > 1 else 0)
                                    formula_input = st.text_input("✏️ 自定义公式 (使用 A 和 B)", value="A - B")
                                    try:
                                        A_data = df_pivot[var_a].fillna(0)
                                        B_data = df_pivot[var_b].fillna(0) if var_b != "无" else 0
                                        res = pd.eval(formula_input, local_dict={'A': A_data, 'B': B_data})
                                        df_pivot['final_val'] = (
                                            pd.Series(res).replace([np.inf, -np.inf], 0).fillna(0).values
                                        )
                                        plot_df_base = df_pivot[['公司', '报告年份', 'final_val']].copy()
                                    except Exception as e:
                                        st.error(f"公式无效: {e}")
                                        plot_df_base = pd.DataFrame(columns=['公司', '报告年份', 'final_val'])
        
                            # ── 公司类型筛选 ─────────────────────────────────────────
                            all_cos_list = sorted(df_pivot['公司'].unique().tolist())
                            if all_co_types:
                                co_type_sel = st.selectbox(
                                    "🏢 按公司类型快速选择", ["不按公司类型筛选"] + all_co_types, key="co_type_sel"
                                )
                                if (co_type_sel != "不按公司类型筛选" and
                                        st.session_state.get('_prev_co_type_sel') != co_type_sel):
                                    st.session_state['selected_cos_ms'] = sorted(
                                        df_clean[df_clean['公司类型'] == co_type_sel]['公司'].unique().tolist()
                                    )
                                    st.session_state['_prev_co_type_sel'] = co_type_sel
        
                            selected_cos = st.multiselect(
                                "🏗️ 选择对比公司", all_cos_list,
                                default=all_cos_list[:min(2, len(all_cos_list))],
                                key="selected_cos_ms"
                            )
        
                    with r1c2:
                        x_axis_mode = (
                            st.radio("🔍 布局视角", ["以公司为横轴", "以年份为横轴"])
                            if "环" not in chart_type else "结构视角"
                        )
                        decimals   = st.number_input("🔢 小数位数", min_value=0, max_value=4, value=0)
                        show_value = st.toggle("✅ 显示数据标签", value=True)
        
                    with r1c3:
                        unit_options  = {"原始数值": 1.0, "亿元": 0.01, "十亿元": 0.001, "百分比(%)": 100.0}
                        selected_unit = st.selectbox("📏 数值单位换算", list(unit_options.keys()))
                        multiplier    = unit_options[selected_unit]
                        y_axis_title  = st.text_input("📝 Y轴单位显示修改", value=f"单位: {selected_unit.split(' ')[0]}")
                        is_transparent = st.toggle("🌈 开启透明背景模式")
                        show_avg  = st.checkbox("平均值线") if "柱" in chart_type else False
                        avg_color = st.color_picker("基准线颜色", value="#ED2124") if show_avg else "#ED2124"
        
                # ── 绘图数据准备 ─────────────────────────────────────────────────────
                if plot_df_base.empty:
                    st.warning("暂无数据，请检查筛选条件。")
                    st.stop()
        
                if calc_mode in ("结构分析", "多指标占比"):
                    cos_filter = [selected_cos] if isinstance(selected_cos, str) else selected_cos
                    plot_df  = plot_df_base[plot_df_base['公司'].isin(cos_filter)].copy()
                    color_val = "字段名"
                else:
                    plot_df  = plot_df_base[plot_df_base['公司'].isin(selected_cos)].copy()
                    color_val = "报告年份" if x_axis_mode == "以公司为横轴" else "公司"
        
                plot_df['绘制金额'] = plot_df['final_val'] * multiplier
        
                # ── 图例颜色配置 ─────────────────────────────────────────────────────
                st.markdown("#### 🎨 自定义图例标签与颜色")
                unique_items = sorted(plot_df[color_val].unique().tolist())
                rename_map, color_map = {}, {}
                c_cols = st.columns(4)
                for i, item in enumerate(unique_items):
                    with c_cols[i % 4]:
                        st.caption(f"原始值: {item}")
                        new_label = st.text_input("显示名称", value=str(item), key=f"rename_{item}")
                        # 兼容外部如果定义了 DEFAULT_COLORS
                        default_color_fallback = DEFAULT_COLORS[i % len(DEFAULT_COLORS)] if 'DEFAULT_COLORS' in globals() else "#1E49E2"
                        default_c = dynamic_year_colors.get(str(item), default_color_fallback)
                        new_color = st.color_picker("选择颜色", value=default_c, key=f"color_{item}")
                        rename_map[item]     = new_label
                        color_map[new_label] = new_color
        
                plot_df[color_val] = plot_df[color_val].map(rename_map)
        
                # ── 绘图引擎 ─────────────────────────────────────────────────────────
                import plotly.graph_objects as go
                import plotly.express as px
                
                fig       = go.Figure()
                fmt       = f'.{decimals}f'
                suffix    = "%" if selected_unit == "百分比(%)" else ""
                label_fmt = f'%{{y:{fmt}}}{suffix}'
        
                if "柱状图" in chart_type:
                    barmode = 'group' if "簇状" in chart_type else 'relative'
        
                    if is_pct_stack_mode:
                        plot_df['x_label'] = plot_df['公司'] + ' ' + plot_df['报告年份']
                        plot_df['_total']  = plot_df.groupby('x_label')['绘制金额'].transform('sum')
                        plot_df['绘制占比'] = (plot_df['绘制金额'] / plot_df['_total'] * 100).fillna(0)
                        for item in rename_map.values():
                            d = plot_df[plot_df[color_val] == item]
                            fig.add_trace(go.Bar(
                                x=d['x_label'], y=d['绘制占比'], name=str(item),
                                marker_color=color_map[item],
                                text=(
                                    d.apply(lambda r: f"{r['绘制占比']:.{decimals}f}%<br>{r['绘制金额']:.{decimals}f}", axis=1)
                                    if show_value else None
                                ),
                                textposition='inside'
                            ))
                        fig.update_yaxes(range=[0, 105])
                        barmode = 'relative'
        
                    else:
                        for item in rename_map.values():
                            d = plot_df[plot_df[color_val] == item]
                            fig.add_trace(go.Bar(
                                x=d["公司" if color_val == "报告年份" else "报告年份"],
                                y=d["绘制金额"], name=str(item),
                                marker_color=color_map[item],
                                text=d["绘制金额"] if show_value else None,
                                texttemplate=label_fmt if show_value else None,
                                textposition='outside', textangle=0
                            ))
        
                    fig.update_layout(barmode=barmode)
        
                elif "折线对比图" in chart_type:
                    for item in rename_map.values():
                        d = plot_df[plot_df[color_val] == item]
                        fig.add_trace(go.Scatter(
                            x=d["公司" if color_val == "报告年份" else "报告年份"],
                            y=d["绘制金额"], name=str(item),
                            mode='lines+markers+text' if show_value else 'lines+markers',
                            marker_color=color_map[item],
                            text=d["绘制金额"], texttemplate=label_fmt, textposition="top center"
                        ))
        
                elif chart_type == "饼图":
                    latest = plot_df['报告年份'].max()
                    d = plot_df[plot_df['报告年份'] == latest]
                    fig = px.pie(d, values='绘制金额', names=color_val, hole=0.4,
                                 color=color_val, color_discrete_map=color_map)
                    fig.update_traces(textinfo='percent+label' if show_value else 'percent')
        
                elif chart_type == "内外环结构对比图":
                    years_ring = sorted(plot_df['报告年份'].unique().tolist())
                    if len(years_ring) < 2:
                        st.warning("环形图对比需要至少两年的数据。")
                    else:
                        d_outer = plot_df[plot_df['报告年份'] == years_ring[-1]]
                        d_inner = plot_df[plot_df['报告年份'] == years_ring[0]]
                        fig.add_trace(go.Pie(
                            labels=d_outer[color_val], values=d_outer['绘制金额'],
                            hole=0.7, name=years_ring[-1],
                            marker=dict(colors=[color_map[f] for f in d_outer[color_val]]),
                            textinfo='percent+label' if show_value else 'percent'
                        ))
                        fig.add_trace(go.Pie(
                            labels=d_inner[color_val], values=d_inner['绘制金额'],
                            hole=0.4, name=years_ring[0],
                            domain={'x': [0.15, 0.85], 'y': [0.15, 0.85]},
                            marker=dict(colors=[color_map[f] for f in d_inner[color_val]]),
                            textinfo='percent' if show_value else 'none'
                        ))
                        fig.update_layout(annotations=[dict(
                            text=f'内:{years_ring[0]}<br>外:{years_ring[-1]}',
                            x=0.5, y=0.5, font_size=12, showarrow=False
                        )])
        
                # ── 全局样式 ─────────────────────────────────────────────────────────
                bg_color = "rgba(0,0,0,0)" if is_transparent else "white"
        
                if show_avg and "柱" in chart_type and not is_pct_stack_mode:
                    avg_v = plot_df['绘制金额'].mean()
                    fig.add_hline(y=avg_v, line_dash="dash", line_color=avg_color,
                                  annotation_text=f"平均: {avg_v:{fmt}}{suffix}",
                                  annotation_font=dict(color=avg_color))
        
                fig.update_layout(
                    font_family="Microsoft YaHei",
                    plot_bgcolor=bg_color, paper_bgcolor=bg_color,
                    margin=dict(t=120, l=10, r=10, b=20),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                    annotations=[dict(
                        x=0, y=1.18, xref='paper', yref='paper',
                        text=f"<b>{y_axis_title}</b>", showarrow=False,
                        font=dict(size=14, color="#333333"), xanchor='left'
                    )]
                )
                fig.update_xaxes(showgrid=False)
                fig.update_yaxes(showgrid=True, gridcolor="#f0f0f0")
        
                st.plotly_chart(fig, use_container_width=True)
        
                with st.expander("📄 查看底层数据明细"):
                    st.dataframe(plot_df, use_container_width=True)
        
            # ==========================================
            # 在页面上调用封装好的交互展示区
            # ==========================================
            render_viz_dashboard(df_clean, df_pivot, val_col, all_fields, all_types, all_co_types, dynamic_year_colors)
                
         
        # ─────────── Step 7 固定图的展示和 PPT ───────────
        with tab7:
            show_step_7_content()
            
            
        # ─────────── Step 8 固定图的展示和 PPT ───────────
        with tab8:
            show_step_8_content()
        
        
        
        
        
            
    else:
        # 如果是 普通用户，只开放 6 和 7 两个 Tab
        st.warning("🔒 您当前仅拥有数据可视化与报告生成模块的访问权限。")
        tab6, tab7, tab8 = st.tabs([
            " 📊 Step 6 ／ 自定义对标分析 ",
            " 🖼️ Step 7 ／ 公司级对标报告 ",
            " 📈 Step 8 ／ 行业分类统计分析 "
        ])
        
        # ─────────── Step 6 可视化分析面板 ───────────
        with tab6:
            st.markdown("### 📊 自定义对标分析")
        
            # ── 数据源选择 ───────────────────────────────────────────────────────
            if 'integrated_data' not in st.session_state:
                st.session_state['integrated_data'] = None
        
            source_choice = st.radio(
                "数据源选择", ["直接引用集成后的数据", "上传集成表 Excel"], horizontal=True
            )
            df_raw = None
        
            if source_choice == "上传集成表 Excel":
                viz_file = st.file_uploader("上传行业集成目标表", type=["xlsx"])
                if viz_file:
                    df_raw = pd.read_excel(viz_file)
                    st.session_state['integrated_data'] = df_raw
            else:
                df_raw = st.session_state.get('integrated_data')
        
            if df_raw is None:
                st.info("💡 请先完成数据集成或上传目标底稿。")
                st.stop()
        
            # ==========================================
            # 🌟 提速秘籍 1：锁定数据预处理 (只算一次)
            # ==========================================
            @st.cache_data(show_spinner=False)
            def prepare_viz_data(df_in):
                df_clean = df_in.copy()
                df_clean.columns = (
                    df_clean.columns.astype(str)
                    .str.strip()
                    .str.replace('\n', '', regex=False)
                    .str.replace('\r', '', regex=False)
                    .str.replace('\ufeff', '', regex=False)
                )
                df_clean['报告年份'] = df_clean['报告年份'].astype(str).str.replace('.0', '', regex=False)
                val_col = "(百万)人民币" if "(百万)人民币" in df_clean.columns else df_clean.columns[-1]
        
                # 去重
                dedup_cols = ['公司', '报告年份', '字段名']
                if '类别' in df_clean.columns:
                    dedup_cols.append('类别')
                df_clean = df_clean.drop_duplicates(subset=dedup_cols, keep='first')
        
                # 透视表
                df_pivot = (
                    df_clean.groupby(['公司', '报告年份', '字段名'])[val_col]
                    .sum().unstack('字段名').reset_index()
                )
        
                # 提取选项列表
                all_fields   = sorted(df_clean['字段名'].unique().tolist())
                all_types    = sorted(df_clean['类别'].dropna().astype(str).unique().tolist()) if '类别'    in df_clean.columns else []
                all_co_types = sorted(df_clean['公司类型'].dropna().astype(str).unique().tolist()) if '公司类型' in df_clean.columns else []
        
                all_years_sorted = sorted(
                    [y for y in df_clean['报告年份'].unique() if str(y).isdigit()],
                    key=lambda x: int(x)
                )
                
                return df_clean, df_pivot, val_col, all_fields, all_types, all_co_types, all_years_sorted
        
            # 调用缓存的数据处理函数
            df_clean, df_pivot, val_col, all_fields, all_types, all_co_types, all_years_sorted = prepare_viz_data(df_raw)
        
            # ── 动态年份颜色计算 (极速，无需缓存) ──────────────────────────────────
            KPMG_COLOR_MAP = {
                "Lightest": "#BFE8FF",     # Pacific Blue (light)
                "Light": "#76D2FF",      # Blue (original Light Blue)
                "Primary": "#1E49E2",     # Cobalt Blue
                "Dark": "#00338D",       # KPMG Blue (original Primary Blue)
            }
            
            n_years = len(all_years_sorted)
            dynamic_year_colors = {}
            target_colors = []
            
            if n_years == 3:
                target_colors = [KPMG_COLOR_MAP["Light"], KPMG_COLOR_MAP["Primary"], KPMG_COLOR_MAP["Dark"]]
            elif n_years == 2:
                target_colors = [KPMG_COLOR_MAP["Primary"], KPMG_COLOR_MAP["Dark"]]
            else:
                target_colors = [KPMG_COLOR_MAP["Lightest"], KPMG_COLOR_MAP["Light"], KPMG_COLOR_MAP["Primary"], KPMG_COLOR_MAP["Dark"]]
            
            num_colors_to_use = len(target_colors)
            if n_years > 0:
                for idx, year in enumerate(all_years_sorted):
                    color_index = 0
                    if n_years > 1:
                        proportion = idx / (n_years - 1)
                        color_index = int(proportion * (num_colors_to_use - 1))
                    color_index = min(color_index, num_colors_to_use - 1)
                    dynamic_year_colors[str(year)] = target_colors[color_index]
        
            # ── KPMG 色卡 ────────────────────────────────────────────────────────
            with st.expander("🎨 查看 KPMG 官方色卡"):
                # 确保全局存在 KPMG_CATEGORIES 变量
                if 'KPMG_CATEGORIES' in globals():
                    for cat_name, cat_colors in KPMG_CATEGORIES.items():
                        st.markdown(f"**{cat_name}**")
                        html_str = "".join([
                            f'<div style="display:inline-block;margin-right:15px;margin-bottom:8px;">'
                            f'<div style="width:14px;height:14px;background-color:{c};display:inline-block;'
                            f'border-radius:3px;vertical-align:middle;border:1px solid #ddd;"></div>'
                            f'<span style="font-size:13px;vertical-align:middle;"> {n} <b>({c})</b></span></div>'
                            for n, c in cat_colors.items()
                        ])
                        st.markdown(html_str, unsafe_allow_html=True)
                else:
                    st.caption("未检测到 KPMG_CATEGORIES 色卡定义。")
        
            st.divider()
        
            # ==========================================
            # 🌟 提速秘籍 2：锁定 UI 与绘图更新范围 (片段刷新)
            # ==========================================
            @st.fragment
            def render_viz_dashboard(df_clean, df_pivot, val_col, all_fields, all_types, all_co_types, dynamic_year_colors):
                # ── 核心配置面板 ─────────────────────────────────────────────────────
                with st.expander("🛠️ 核心配置面板", expanded=True):
                    r1c1, r1c2, r1c3 = st.columns([1.5, 1, 1])
                    is_pct_stack_mode = False
        
                    with r1c1:
                        chart_type = st.selectbox(
                            "📈 1. 图表类型",
                            ["簇状柱状图", "堆积柱状图", "折线对比图", "饼图", "内外环结构对比图"]
                        )
        
                        # ── 饼图 / 环图 ──────────────────────────────────────────────
                        if "环" in chart_type or chart_type == "饼图":
                            calc_mode = "结构分析"
                            selected_multi_fields = st.multiselect(
                                "🎯 选择构成指标", all_fields,
                                default=all_fields[:min(3, len(all_fields))]
                            )
                            selected_cos = st.selectbox("🏗️ 选择展示公司", sorted(df_pivot['公司'].unique().tolist()))
                            plot_df_base = (
                                df_clean[
                                    df_clean['字段名'].isin(selected_multi_fields) &
                                    (df_clean['公司'] == selected_cos)
                                ].copy().rename(columns={val_col: 'final_val'})
                            )
        
                        else:
                            # ── 堆积图子模式 ─────────────────────────────────────────
                            if chart_type == "堆积柱状图":
                                stack_sub = st.radio(
                                    "堆积模式", ["单指标 / 公式", "多指标占比"],
                                    horizontal=True, key="stack_sub_mode"
                                )
                                is_pct_stack_mode = (stack_sub == "多指标占比")
        
                            # ── 多指标占比 ───────────────────────────────────────────
                            if is_pct_stack_mode:
                                calc_mode = "多指标占比"
                                if all_types:
                                    pct_type_f = st.selectbox(
                                        "🗂️ 按类别筛选指标", ["全部类别"] + all_types, key="pct_type_filter"
                                    )
                                    pct_field_pool = (
                                        sorted(df_clean[df_clean['类别'] == pct_type_f]['字段名'].unique().tolist())
                                        if pct_type_f != "全部类别" else all_fields
                                    )
                                else:
                                    pct_field_pool = all_fields
        
                                selected_pct_fields = st.multiselect(
                                    "🎯 选择堆积指标", pct_field_pool,
                                    default=pct_field_pool[:min(3, len(pct_field_pool))],
                                    key="pct_stack_fields"
                                )
                                if selected_pct_fields:
                                    plot_df_base = (
                                        df_clean[df_clean['字段名'].isin(selected_pct_fields)]
                                        [['公司', '报告年份', '字段名', val_col]]
                                        .copy().rename(columns={val_col: 'final_val'})
                                    )
                                else:
                                    plot_df_base = pd.DataFrame(columns=['公司', '报告年份', '字段名', 'final_val'])
        
                            # ── 普通单指标 / 公式 ────────────────────────────────────
                            else:
                                calc_mode = st.radio("数据模式", ["单指标直显", "自定义公式运算"], horizontal=True)
        
                                if calc_mode == "单指标直显":
                                    if all_types:
                                        type_f = st.selectbox(
                                            "🗂️ 按类别筛选指标", ["全部类别"] + all_types, key="type_filter_single"
                                        )
                                        filtered_fields = (
                                            sorted(df_clean[df_clean['类别'] == type_f]['字段名'].unique().tolist())
                                            if type_f != "全部类别" else all_fields
                                        )
                                    else:
                                        filtered_fields = all_fields
        
                                    target_field = st.selectbox("🎯 选择显示指标", filtered_fields)
                                    plot_df_base = (
                                        df_pivot[['公司', '报告年份', target_field]]
                                        .rename(columns={target_field: 'final_val'}).fillna(0)
                                    )
        
                                else:  # 公式模式
                                    v1, v2 = st.columns(2)
                                    var_a = v1.selectbox("变量 A", all_fields, index=0)
                                    var_b = v2.selectbox("变量 B", ["无"] + all_fields, index=1 if len(all_fields) > 1 else 0)
                                    formula_input = st.text_input("✏️ 自定义公式 (使用 A 和 B)", value="A - B")
                                    try:
                                        A_data = df_pivot[var_a].fillna(0)
                                        B_data = df_pivot[var_b].fillna(0) if var_b != "无" else 0
                                        res = pd.eval(formula_input, local_dict={'A': A_data, 'B': B_data})
                                        df_pivot['final_val'] = (
                                            pd.Series(res).replace([np.inf, -np.inf], 0).fillna(0).values
                                        )
                                        plot_df_base = df_pivot[['公司', '报告年份', 'final_val']].copy()
                                    except Exception as e:
                                        st.error(f"公式无效: {e}")
                                        plot_df_base = pd.DataFrame(columns=['公司', '报告年份', 'final_val'])
        
                            # ── 公司类型筛选 ─────────────────────────────────────────
                            all_cos_list = sorted(df_pivot['公司'].unique().tolist())
                            if all_co_types:
                                co_type_sel = st.selectbox(
                                    "🏢 按公司类型快速选择", ["不按公司类型筛选"] + all_co_types, key="co_type_sel"
                                )
                                if (co_type_sel != "不按公司类型筛选" and
                                        st.session_state.get('_prev_co_type_sel') != co_type_sel):
                                    st.session_state['selected_cos_ms'] = sorted(
                                        df_clean[df_clean['公司类型'] == co_type_sel]['公司'].unique().tolist()
                                    )
                                    st.session_state['_prev_co_type_sel'] = co_type_sel
        
                            selected_cos = st.multiselect(
                                "🏗️ 选择对比公司", all_cos_list,
                                default=all_cos_list[:min(2, len(all_cos_list))],
                                key="selected_cos_ms"
                            )
        
                    with r1c2:
                        x_axis_mode = (
                            st.radio("🔍 布局视角", ["以公司为横轴", "以年份为横轴"])
                            if "环" not in chart_type else "结构视角"
                        )
                        decimals   = st.number_input("🔢 小数位数", min_value=0, max_value=4, value=0)
                        show_value = st.toggle("✅ 显示数据标签", value=True)
        
                    with r1c3:
                        unit_options  = {"原始数值": 1.0, "亿元": 0.01, "十亿元": 0.001, "百分比(%)": 100.0}
                        selected_unit = st.selectbox("📏 数值单位换算", list(unit_options.keys()))
                        multiplier    = unit_options[selected_unit]
                        y_axis_title  = st.text_input("📝 Y轴单位显示修改", value=f"单位: {selected_unit.split(' ')[0]}")
                        is_transparent = st.toggle("🌈 开启透明背景模式")
                        show_avg  = st.checkbox("平均值线") if "柱" in chart_type else False
                        avg_color = st.color_picker("基准线颜色", value="#ED2124") if show_avg else "#ED2124"
        
                # ── 绘图数据准备 ─────────────────────────────────────────────────────
                if plot_df_base.empty:
                    st.warning("暂无数据，请检查筛选条件。")
                    st.stop()
        
                if calc_mode in ("结构分析", "多指标占比"):
                    cos_filter = [selected_cos] if isinstance(selected_cos, str) else selected_cos
                    plot_df  = plot_df_base[plot_df_base['公司'].isin(cos_filter)].copy()
                    color_val = "字段名"
                else:
                    plot_df  = plot_df_base[plot_df_base['公司'].isin(selected_cos)].copy()
                    color_val = "报告年份" if x_axis_mode == "以公司为横轴" else "公司"
        
                plot_df['绘制金额'] = plot_df['final_val'] * multiplier
        
                # ── 图例颜色配置 ─────────────────────────────────────────────────────
                st.markdown("#### 🎨 自定义图例标签与颜色")
                unique_items = sorted(plot_df[color_val].unique().tolist())
                rename_map, color_map = {}, {}
                c_cols = st.columns(4)
                for i, item in enumerate(unique_items):
                    with c_cols[i % 4]:
                        st.caption(f"原始值: {item}")
                        new_label = st.text_input("显示名称", value=str(item), key=f"rename_{item}")
                        # 兼容外部如果定义了 DEFAULT_COLORS
                        default_color_fallback = DEFAULT_COLORS[i % len(DEFAULT_COLORS)] if 'DEFAULT_COLORS' in globals() else "#1E49E2"
                        default_c = dynamic_year_colors.get(str(item), default_color_fallback)
                        new_color = st.color_picker("选择颜色", value=default_c, key=f"color_{item}")
                        rename_map[item]     = new_label
                        color_map[new_label] = new_color
        
                plot_df[color_val] = plot_df[color_val].map(rename_map)
        
                # ── 绘图引擎 ─────────────────────────────────────────────────────────
                import plotly.graph_objects as go
                import plotly.express as px
                
                fig       = go.Figure()
                fmt       = f'.{decimals}f'
                suffix    = "%" if selected_unit == "百分比(%)" else ""
                label_fmt = f'%{{y:{fmt}}}{suffix}'
        
                if "柱状图" in chart_type:
                    barmode = 'group' if "簇状" in chart_type else 'relative'
        
                    if is_pct_stack_mode:
                        plot_df['x_label'] = plot_df['公司'] + ' ' + plot_df['报告年份']
                        plot_df['_total']  = plot_df.groupby('x_label')['绘制金额'].transform('sum')
                        plot_df['绘制占比'] = (plot_df['绘制金额'] / plot_df['_total'] * 100).fillna(0)
                        for item in rename_map.values():
                            d = plot_df[plot_df[color_val] == item]
                            fig.add_trace(go.Bar(
                                x=d['x_label'], y=d['绘制占比'], name=str(item),
                                marker_color=color_map[item],
                                text=(
                                    d.apply(lambda r: f"{r['绘制占比']:.{decimals}f}%<br>{r['绘制金额']:.{decimals}f}", axis=1)
                                    if show_value else None
                                ),
                                textposition='inside'
                            ))
                        fig.update_yaxes(range=[0, 105])
                        barmode = 'relative'
        
                    else:
                        for item in rename_map.values():
                            d = plot_df[plot_df[color_val] == item]
                            fig.add_trace(go.Bar(
                                x=d["公司" if color_val == "报告年份" else "报告年份"],
                                y=d["绘制金额"], name=str(item),
                                marker_color=color_map[item],
                                text=d["绘制金额"] if show_value else None,
                                texttemplate=label_fmt if show_value else None,
                                textposition='outside', textangle=0
                            ))
        
                    fig.update_layout(barmode=barmode)
        
                elif "折线对比图" in chart_type:
                    for item in rename_map.values():
                        d = plot_df[plot_df[color_val] == item]
                        fig.add_trace(go.Scatter(
                            x=d["公司" if color_val == "报告年份" else "报告年份"],
                            y=d["绘制金额"], name=str(item),
                            mode='lines+markers+text' if show_value else 'lines+markers',
                            marker_color=color_map[item],
                            text=d["绘制金额"], texttemplate=label_fmt, textposition="top center"
                        ))
        
                elif chart_type == "饼图":
                    latest = plot_df['报告年份'].max()
                    d = plot_df[plot_df['报告年份'] == latest]
                    fig = px.pie(d, values='绘制金额', names=color_val, hole=0.4,
                                 color=color_val, color_discrete_map=color_map)
                    fig.update_traces(textinfo='percent+label' if show_value else 'percent')
        
                elif chart_type == "内外环结构对比图":
                    years_ring = sorted(plot_df['报告年份'].unique().tolist())
                    if len(years_ring) < 2:
                        st.warning("环形图对比需要至少两年的数据。")
                    else:
                        d_outer = plot_df[plot_df['报告年份'] == years_ring[-1]]
                        d_inner = plot_df[plot_df['报告年份'] == years_ring[0]]
                        fig.add_trace(go.Pie(
                            labels=d_outer[color_val], values=d_outer['绘制金额'],
                            hole=0.7, name=years_ring[-1],
                            marker=dict(colors=[color_map[f] for f in d_outer[color_val]]),
                            textinfo='percent+label' if show_value else 'percent'
                        ))
                        fig.add_trace(go.Pie(
                            labels=d_inner[color_val], values=d_inner['绘制金额'],
                            hole=0.4, name=years_ring[0],
                            domain={'x': [0.15, 0.85], 'y': [0.15, 0.85]},
                            marker=dict(colors=[color_map[f] for f in d_inner[color_val]]),
                            textinfo='percent' if show_value else 'none'
                        ))
                        fig.update_layout(annotations=[dict(
                            text=f'内:{years_ring[0]}<br>外:{years_ring[-1]}',
                            x=0.5, y=0.5, font_size=12, showarrow=False
                        )])
        
                # ── 全局样式 ─────────────────────────────────────────────────────────
                bg_color = "rgba(0,0,0,0)" if is_transparent else "white"
        
                if show_avg and "柱" in chart_type and not is_pct_stack_mode:
                    avg_v = plot_df['绘制金额'].mean()
                    fig.add_hline(y=avg_v, line_dash="dash", line_color=avg_color,
                                  annotation_text=f"平均: {avg_v:{fmt}}{suffix}",
                                  annotation_font=dict(color=avg_color))
        
                fig.update_layout(
                    font_family="Microsoft YaHei",
                    plot_bgcolor=bg_color, paper_bgcolor=bg_color,
                    margin=dict(t=120, l=10, r=10, b=20),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                    annotations=[dict(
                        x=0, y=1.18, xref='paper', yref='paper',
                        text=f"<b>{y_axis_title}</b>", showarrow=False,
                        font=dict(size=14, color="#333333"), xanchor='left'
                    )]
                )
                fig.update_xaxes(showgrid=False)
                fig.update_yaxes(showgrid=True, gridcolor="#f0f0f0")
        
                st.plotly_chart(fig, use_container_width=True)
        
                with st.expander("📄 查看底层数据明细"):
                    st.dataframe(plot_df, use_container_width=True)
        
            # ==========================================
            # 在页面上调用封装好的交互展示区
            # ==========================================
            render_viz_dashboard(df_clean, df_pivot, val_col, all_fields, all_types, all_co_types, dynamic_year_colors)
                
         
            
        # ─────────── Step 7 固定图的展示和 PPT ───────────
        with tab7:
            show_step_7_content()            

        # ─────────── Step 8 固定图的展示和 PPT ───────────
        with tab8:
            show_step_8_content()        


 

# ==================== 页脚 ====================
st.markdown("""
<div style="text-align: center; color: #94A3B8; font-size: 13px; letter-spacing: 1px; margin-top: 50px; padding: 20px; border-top: 1px solid #CBD5E1;">
    Actuarial Data Intelligence · Built for KPMG Actuary Team
</div>
""", unsafe_allow_html=True)
